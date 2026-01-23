# riskapp/app.py

from flask import (
    Flask, render_template, request, redirect, url_for,
    session, flash, current_app, Response, jsonify, abort,send_file
)
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, date, timedelta
import os
from sqlalchemy.exc import OperationalError
from decimal import Decimal, InvalidOperation
from sqlalchemy import desc
from functools import wraps
from sqlalchemy import text, or_, func
from collections import Counter
import csv
from io import StringIO
import io, csv as _csv, os, re, json
from werkzeug.utils import secure_filename
from pathlib import Path
from collections import defaultdict
from flask import current_app
from flask import request, redirect, url_for, flash, current_app
from .models import db, Risk, Comment
from .ai_local.commenter import make_ai_risk_comment, _propose_actions
from io import BytesIO
from .models import CostItem, CostTemplate
from flask import session
from riskapp.models import db, Risk, Evaluation, CostItem
from riskapp.models import CostItem, CostTemplate
import re
from sqlalchemy.exc import IntegrityError
from flask import request
from .models import db, Risk, Mitigation
from datetime import datetime
from io import BytesIO
from datetime import date
from flask import send_file
import time
from datetime import datetime, timedelta
from sqlalchemy.exc import IntegrityError
from flask import request, redirect, url_for, render_template, jsonify, flash, abort
from sqlalchemy import or_
from sqlalchemy.exc import IntegrityError
from flask import request, jsonify
from sqlalchemy.exc import IntegrityError
from flask import request, redirect, url_for, abort, flash
from riskapp.models import Cost

import json
import os as _os, sys as _sys
PKG_ROOT = _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__)))
if PKG_ROOT not in _sys.path:
    _sys.path.insert(0, PKG_ROOT)

import os, smtplib
from email.message import EmailMessage

from urllib.parse import urlparse, quote

from dotenv import load_dotenv
load_dotenv()  # proje kökündeki .env dosyasını okur

from riskapp.ai_local.ps_estimator import PSEstimator
from riskapp.ai_local.engine import AILocal
from riskapp.models import db, Risk, Mitigation   

from sqlalchemy.exc import IntegrityError
import re

from flask import Blueprint
# --- Proje içi paket-absolute importlar ---
from riskapp.models import (
     db, Risk, Evaluation, Comment, Suggestion,
     Account, ProjectInfo, RiskCategory, RiskCategoryRef,
     CostItem
)

from riskapp.seeder import seed_if_empty
from riskapp.ai_utils import ai_complete, ai_json, best_match

# === AI P/S & RAG için ek importlar ===
from riskapp.ai_local.ps_estimator import PSEstimator
from riskapp.ai_local.engine import AILocal

# --- Çok formatlı içe aktarma için opsiyonel bağımlılık ---
try:
    import pandas as _pd
except Exception:
    _pd = None  # pandas yoksa Excel içe aktarmada uyarı veririz

# --- PDF backend'leri opsiyonel olarak yükle ---
try:
    from weasyprint import HTML, CSS  # type: ignore
except Exception:
    HTML = CSS = None  # type: ignore
    # Not: Windows'ta GTK/Pango/Cairo eksikse burada düşecek, sorun değil.

try:
    import pdfkit  # fallback
except Exception:
    pdfkit = None

import re as _re  # importlar arasında yoksa ekle
from flask import jsonify
from sqlalchemy import or_

# Ref No formatı (örn: R-PRJ12-2025-0034)
_REF_PATTERN = _re.compile(r"^R-[A-Z0-9]{2,10}-\d{4}-\d{3,6}$")

from random import choices
import string
COST_CATEGORIES = ["İş Gücü", "Ekipman", "Yazılım", "Eğitim", "Hizmet", "Operasyon"]
# Basit TTL cache (external cache yoksa bile iş görür)
_PARETO_AI_CACHE = {}  # key -> (ts, payload)
_CACHE_TTL_SEC = 30

def _parse_ym(s):
    """'YYYY-MM' ya da 'YYYY-MM-DD' -> (y, m) | None"""
    try:
        if not s:
            return None
        s = str(s).strip()[:7]   # 'YYYY-MM-DD' gelirse ilk 7'yi al
        y, m = s.split("-")
        y, m = int(y), int(m)
        if 1 <= m <= 12:
            return (y, m)
    except Exception:
        pass
    return None

def _ym_to_str(y, m):
    return f"{int(y):04d}-{int(m):02d}"

def _next_ym(y, m):
    y, m = int(y), int(m)
    return (y + (m // 12), 1 if m == 12 else m + 1)

import unicodedata as _ud

def _normcat(s: str) -> str:
    # boşlukları kırp + Unicode'u NFC'ye getir + casefold ile küçük harf
    return _ud.normalize("NFC", (s or "").strip()).casefold()

def _parse_date(s: str):
    try:
        s = (s or "").strip()
        if not s:
            return None
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        return None

def _to_float(s: str):
    try:
        return float(s) if s not in (None, "") else None
    except Exception:
        return None

def _to_int(s: str):
    try:
        return int(s) if s not in (None, "") else None
    except Exception:
        return None
# -------------------------------------------------
# AI çıktı temizleyiciler (tekrar/eko önleme)
# -------------------------------------------------
def _strip_ai_artifacts(txt: str) -> str:
    """
    Modelin eklediği gereksiz tekrarları/prompt ekolarını temizler.
    - '--- Soru:' ,'Soru:', 'MEVCUT ÖNLEMLER:' gibi satırları atar
    - 'BENZER ÖNERİLER:' bloğunu en fazla 1 kez bırakır
    - 3+ boş satırı 1 boş satıra indirir
    """
    if not txt:
        return ""
    seen_benzer = False
    out_lines = []
    for raw in txt.splitlines():
        line = raw.strip()

        # prompt/eko/teknik satırlar
        if line.startswith(("--- Soru:", "Soru:", "MEVCUT ÖNLEMLER:", "AI ek not:", "AI Önerisi ile oluşturuldu")):
            continue
        if line.startswith("Not: Bu çıktı"):
            continue

        # "BENZER ÖNERİLER" sadece 1 kez
        if line.startswith("BENZER ÖNERİLER"):
            if seen_benzer:
                continue
            seen_benzer = True

        out_lines.append(raw)

    out = "\n".join(out_lines).strip()
    out = _re.sub(r"\n{3,}", "\n\n", out)

    return out


def _strip_ai_in_mitigation(mit: str | None) -> str | None:
    """
    Mitigation içindeki önceki AI çıktısını ayıklar (feedback loop'u kırar).
    '🤖', '---', 'Soru:' gibi işaretçilerden sonrası atılır.
    """
    if not mit:
        return None
    keep = []
    for raw in mit.splitlines():
        s = raw.strip()
        if s.startswith(("🤖", "---", "Soru:")) or "AI Önerisi" in s:
            break
        keep.append(raw)
    clean = "\n".join(keep).strip()
    return clean or None


def _guess_wkhtmltopdf_path() -> str | None:
    """Windows'ta yaygın wkhtmltopdf yollarını dener, yoksa PATH'e güvenir."""
    candidates = [
        r"C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe",
        r"C:\Program Files (x86)\wkhtmltopdf\bin\wkhtmltopdf.exe",
    ]
    env_path = os.getenv("WKHTMLTOPDF_PATH")
    if env_path:
        candidates.insert(0, env_path)

    for p in candidates:
        if Path(p).exists():
            return p
    return "wkhtmltopdf"  # PATH'te bulunabiliyorsa çalışır


# -------------------------------------------------
# Şema güvence: eksik kolonlar varsa ekle (SQLite)
# -------------------------------------------------
def ensure_schema():
    """SQLite üzerinde basit ALTER kontrolleri (geriye dönük uyum)."""

    def has_col(table, col):
        res = db.session.execute(text(f"PRAGMA table_info({table})")).fetchall()
        return any(r[1] == col for r in res)

    changed = False

    # --- risks tablosu için yeni alanlar ---
    for col in ["risk_type", "responsible", "mitigation", "duration", "start_month", "end_month"]:
        if not has_col("risks", col):
            db.session.execute(text(f"ALTER TABLE risks ADD COLUMN {col} TEXT"))
            changed = True

    # risks.project_id
    if not has_col("risks", "project_id"):
        db.session.execute(text("ALTER TABLE risks ADD COLUMN project_id INTEGER"))
        changed = True

    # ✅ risks.ref_code (Ref No — admin atar, benzersiz)
    if not has_col("risks", "ref_code"):
        db.session.execute(text("ALTER TABLE risks ADD COLUMN ref_code TEXT"))
        changed = True

    # --- accounts.role ---
    if not has_col("accounts", "role"):
        db.session.execute(text("ALTER TABLE accounts ADD COLUMN role TEXT DEFAULT 'uzman'"))
        changed = True

    # accounts.ref_code (kayıtta kullanılan referans)
    if not has_col("accounts", "ref_code"):
        db.session.execute(text("ALTER TABLE accounts ADD COLUMN ref_code TEXT"))
        changed = True

    # accounts.status (pending/active/disabled)
    if not has_col("accounts", "status"):
        db.session.execute(text(
            "ALTER TABLE accounts ADD COLUMN status TEXT DEFAULT 'pending'"
        ))
        db.session.execute(text(
            "UPDATE accounts SET status='pending' WHERE status IS NULL"
        ))
        changed = True

    # İndeksleri her koşulda dene (IF NOT EXISTS güvenli)
    db.session.execute(text(
        "CREATE INDEX IF NOT EXISTS ix_accounts_status ON accounts(status)"
    ))
    db.session.execute(text(
        "CREATE INDEX IF NOT EXISTS ix_accounts_ref_code ON accounts(ref_code)"
    ))

    # evaluations.detection (eski RPN alanı için geriye uyum)
    if not has_col("evaluations", "detection"):
        db.session.execute(text("ALTER TABLE evaluations ADD COLUMN detection INTEGER"))
        changed = True

    # project_info.project_duration
    if not has_col("project_info", "project_duration"):
        db.session.execute(text("ALTER TABLE project_info ADD COLUMN project_duration TEXT"))
        changed = True

    # suggestions ek kolonlar
    if not has_col("suggestions", "risk_code"):
        db.session.execute(text("ALTER TABLE suggestions ADD COLUMN risk_code TEXT"))
        changed = True
    if not has_col("suggestions", "default_prob"):
        db.session.execute(text("ALTER TABLE suggestions ADD COLUMN default_prob INTEGER"))
        changed = True
    if not has_col("suggestions", "default_sev"):
        db.session.execute(text("ALTER TABLE suggestions ADD COLUMN default_sev INTEGER"))
        changed = True

    # ✅ YENİ: Excel'den gelecek açıklama ve önlem alanları
    if not has_col("suggestions", "risk_desc"):
        db.session.execute(text("ALTER TABLE suggestions ADD COLUMN risk_desc TEXT"))
        changed = True
    if not has_col("suggestions", "mitigation_hint"):
        db.session.execute(text("ALTER TABLE suggestions ADD COLUMN mitigation_hint TEXT"))
        changed = True

    # suggestions.created_at / updated_at (backfill)
    if not has_col("suggestions", "created_at"):
        db.session.execute(text("ALTER TABLE suggestions ADD COLUMN created_at DATETIME"))
        db.session.execute(text("UPDATE suggestions SET created_at = CURRENT_TIMESTAMP WHERE created_at IS NULL"))
        changed = True
    if not has_col("suggestions", "updated_at"):
        db.session.execute(text("ALTER TABLE suggestions ADD COLUMN updated_at DATETIME"))
        db.session.execute(text("UPDATE suggestions SET updated_at = CURRENT_TIMESTAMP WHERE updated_at IS NULL"))
        changed = True

    if changed:
        db.session.commit()

    # referral_codes tablosu
    db.session.execute(text("""
        CREATE TABLE IF NOT EXISTS referral_codes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT UNIQUE NOT NULL,
            assigned_email TEXT,
            is_used INTEGER DEFAULT 0,
            created_by INTEGER,
            expires_at TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """))
    try:
        db.session.execute(text("CREATE UNIQUE INDEX IF NOT EXISTS ux_refcodes_code ON referral_codes(code)"))
    except Exception:
        pass

    if changed:
        db.session.commit()


def _gen_ref_code(prefix="PRJ", year=None, digits=6):
    y = year or datetime.now().year
    while True:
        seq = "".join(choices(string.digits, k=digits))
        code = f"{prefix}-{y}-{seq}"
        exists = Account.query.filter(Account.ref_code == code).first()
        if not exists:
            return code



# -------------------------------------------------
#  CSV / XLSX / XLS dosyadan satır okuma helper'ı
# -------------------------------------------------
def _read_rows_from_upload(file_storage):
    """
    CSV, XLSX, XLS dosyasını satır listesi (list[list[str]]) olarak döndürür.
    Header satırını dahil eder; ayıracı otomatik algılar.

    🆕 Excel için: başlık satırı ilk 10 satır içinde otomatik bulunur
    (ör: A3:G3). “Risk Kodları”, “Risk Faktörü”, “Kategoriler” gibi
    başlıklar normalize edilerek aranır.
    """
    filename = secure_filename(file_storage.filename or "")
    ext = (os.path.splitext(filename)[1] or "").lower()

    # --- EXCEL (.xlsx/.xls) ---
    if ext in (".xlsx", ".xls"):
        if not _pd:
            raise RuntimeError("Excel içe aktarmak için 'pandas' + 'openpyxl/xlrd' kurulu olmalı.")

        # normalize helper
        _TRMAP = str.maketrans({
            "ç":"c","ğ":"g","ı":"i","ö":"o","ş":"s","ü":"u",
            "Ç":"c","Ğ":"g","İ":"i","Ö":"o","Ş":"s","Ü":"u"
        })
        def _norm(s):
            s = str(s or "").replace("\n"," ").replace("\r"," ").strip().translate(_TRMAP).lower()
            return " ".join(s.split())

        # Tüm sayfaları başlıksız oku
        try:
            engine = "openpyxl" if ext == ".xlsx" else "xlrd"
            sheets = _pd.read_excel(file_storage, engine=engine, sheet_name=None, header=None)
        except Exception as e:
            raise RuntimeError(f"Excel okuma hatası: {e}")

        # Hedef başlık kümeleri
        must_keys = {"risk faktoru", "risk faktörü"}
        bonus_keys = {"risk kodlari", "risk kodları", "kategoriler", "kategori"}

        for sheet_name, df in sheets.items():
            if df is None or df.empty:
                continue

            # Başlık satırını ilk 10 satırda ara
            header_row = None
            look_rows = min(10, len(df))
            for i in range(look_rows):
                cols = [_norm(c) for c in list(df.iloc[i, :])]
                setcols = set(cols)
                if must_keys.issubset(setcols) or (("risk faktoru" in setcols or "risk faktörü" in setcols) and (setcols & bonus_keys)):
                    header_row = i
                    break
            if header_row is None:
                continue  # başka sayfaya bak

            # header bulundu → gövdeyi çıkar
            body = df.iloc[header_row+1:].copy()
            header_vals = list(df.iloc[header_row, :])
            # NaN'leri boş string yap
            body = body.fillna("")
            # kolon adları
            body.columns = header_vals

            # rows = [header] + data
            header_row_out = [str(c).replace("\n"," ").replace("\r"," ").strip() for c in header_vals]
            data_rows_out = body.astype(str).values.tolist()
            return [header_row_out] + data_rows_out

        # hiçbir sayfada başlık bulunamadı
        raise RuntimeError("Excel’de başlık satırı bulunamadı. İlk 10 satırda 'Risk Faktörü' bekleniyor.")

    # --- CSV ---
    raw = file_storage.read()

    text_data = None
    for enc in ("utf-8-sig", "utf-8", "cp1254", "iso-8859-9", "latin-1"):
        try:
            text_data = raw.decode(enc)
            break
        except Exception:
            continue
    if text_data is None:
        raise RuntimeError("Dosya kodlaması çözülemedi. CSV'yi 'UTF-8 (virgülle ayrılmış)' kaydedin.")

    sample = text_data[:4096]
    try:
        dialect = _csv.Sniffer().sniff(sample, delimiters=[",",";","\t","|"])
    except Exception:
        # basit fallback: ';' çoksa ';' kabul et, yoksa ','
        if sample.count(";") > sample.count(","):
            class _D: delimiter=";"
            dialect = _D()
        else:
            class _D: delimiter=","
            dialect = _D()

    reader = _csv.reader(io.StringIO(text_data), dialect)
    return [row for row in reader]


# ============================
# Yardımcılar (AI + RACI + KPI)
# ============================

def _smart_due(days: int = 30) -> str:
    return (date.today() + timedelta(days=days)).isoformat()

def _normalize(s: str) -> str:
    """Türkçe karakterleri sadeleştir + lower."""
    if not s:
        return ""
    tr_map = str.maketrans({
        "ç":"c","Ç":"c","ğ":"g","Ğ":"g","ı":"i","İ":"i",
        "ö":"o","Ö":"o","ş":"s","Ş":"s","ü":"u","Ü":"u"
    })
    return s.translate(tr_map).lower()

def _any_in(text: str, keywords) -> bool:
    t = _normalize(text)
    return any(k in t for k in keywords)

def _unique(seq):
    seen = set()
    out = []
    for x in seq:
        key = (x.get("action"), x.get("due"))
        if key not in seen:
            seen.add(key)
            out.append(x)
    return out


# Kategori anahtar kümeleri (normalize edilmiş aramayla eşleşir)
KEYSETS = {
    "insaat": [
        "beton","kalip","donati","dokum","vibrator","santiye","saha",
        "betonarme","formwork","rebar","pour","scaffold"
    ],
    "satinalma": [
        "satinalma","tedarik","malzeme","lojistik","irsaliye","siparis",
        "po","rfq","tedarikci","nakliye","sevkiyat","warehouse","supply"
    ],
    "sozlesme": [
        "sozlesme","legal","hukuk","onay","izin","reg","regulasyon",
        "idari sartname","teknik sartname","claim","variation","vo"
    ],
    "isg_cevre": [
        "isg","is guvenligi","kaza","ramak kala","cevre","emisyon","atik",
        "toz","gurultu","ppe","acil durum","ced","emission","waste","noise","spill"
    ],
    "geoteknik": [
        "zemin","geoteknik","kazi","iksa","zayif zemin","oturma","sev","sev stabilitesi",
        "cpt","spt","sonder","forekazik","ankraj"
    ],
    "kalite": [
        "kalite","denetim","tetkik","audit","muayene","itp","tutanak","numune",
        "slump","ndt","wps","pqr","kalibrasyon","inspection","hold point"
    ],
    "pmo": [
        "politik","organizasyonel","paydas","stakeholder","iletisim plani",
        "raporlama","kpi","koordinasyon","komite"
    ],
    "planlama": [
        "planlama","program","zaman cizelgesi","kritik yol","cpm",
        "ms project","primavera","p6","gant","delay","erteleme",
        "hava","ruzgar","yagis","sicaklik","weather","wind","rain","temperature","storm"
    ],
    "mep_elektrik": [
        "elektrik","og","ag","trafo","scada","pano","kablo","tray","aydinlatma",
        "topraklama","kesici","jenerator","ups","megger","loop test","komisyoning","commissioning"
    ],
    "mep_mekanik": [
        "mekanik","hvac","chiller","kazan","pompa","yangin","sprinkler","tesisat",
        "borulama","pnid","basinc testi","hidrostatik","duct","valf","esanjör","esanjör"
    ],
    "marine": [
        "deniz","marine","rihtim","iskele","kazik","celik kazik","dolfen","samandira",
        "batimetri","akinti","dalga","romorkor","barge","vinc barge","mendirek","dalgakiran","kran"
    ],
    "tasarim": [
        "tasarim","cizim","revizyon","ifc","shop drawing","shopdrawing","statik",
        "mimari","clash","detay","kesit","rfi"
    ],
    "teknik_ofis": [
        "teknik ofis","metraj","hakedis","atasman","boq","kesif","birim fiyat",
        "poz","revize kesif","maliyet analizi","progress"
    ],
    "finans": [
        "finans","butce","nakit akisi","cash flow","fatura","tahsilat","teminat",
        "kesinti","avans","kur riski","maliyet","capex","opex"
    ],
    "makine_bakim": [
        "ekipman","makine","bakim","ariza","yedek parca","operator","vinc",
        "excavator","loader","forklift","servis","periyodik kontrol","rigging","lifting plan","winch"
    ],
    "bim_bt": [
        "bim","model","revit","navisworks","ifc dosyasi","clash detection",
        "veri tabani","sunucu","yedekleme","network","cad","gis"
    ],
    "izin_ruhsat": [
        "ruhsat","belediye","imar","fenni mesul","tutanak","resmi yazi","dilekce",
        "trafik kesme izni","enkaz izin","izin sureci"
    ],
    "laboratuvar": [
        "laboratuvar","numune","slump","karot","cekme testi","basinc testi",
        "agrega","granulometri","ndt","ultrasonik test"
    ],
    "depo": [
        "depo","ambar","stok","stok sayim","emniyet stogu","raf",
        "malzeme teslim","giris cikis","stok devir","ambar fisi"
    ],
}

# Kategori -> aksiyon şablonları (metin, due_gun)
ACTION_TEMPLATES = {
    "insaat": [
        ("Dokum oncesi Kalip & Donati Checklist %100 tamamlansin", 7),
        ("ITP ve Muayene-Kabul plani revize edilip saha ekibine brief verilsin", 10),
        ("TS EN 206’a gore numune alma-kur plani ve tedarikci denetimi yapilsin", 14),
        ("Ustalara beton yerlestirme & vibrasyon toolbox talk (egitim)", 5),
    ],
    "satinalma": [
        ("Kritik malzemeler icin ikincil tedarikci onayi (dual sourcing)", 14),
        ("Satinalma sozlesmelerine gecikme cezasi & SLA maddeleri eklensin", 10),
        ("Lojistikte emniyet stok seviyesi ve takip KPI’lari tanimlansin", 7),
    ],
    "sozlesme": [
        ("Kritik izin/onaylar icin izleme matrisi ve sorumlu atamasi", 5),
        ("Sozlesme risk maddeleri (ceza/force majeure) gozden gecirme", 10),
        ("Isveren/danisman iletisim plani ve haftalik durum raporu", 7),
    ],
    "isg_cevre": [
        ("Cevresel Etki Plani guncelleme (toz, gurultu, atik yonetimi)", 7),
        ("Izleme ekipmani (gurultu/toz) kalibrasyon ve kayit duzeni", 10),
        ("Yerel otoriteye raporlama periyotlari ve sorumlular netlesin", 14),
    ],
    "geoteknik": [
        ("Zemin parametreleri guncellenip tasarim emniyet katsayilari teyit", 10),
        ("Iksa/sev stabilitesi gunluk izleme ve tetik degerleri", 5),
        ("Beklenmeyen zemin kosul proseduru (claim/KEsIF) hazir", 14),
    ],
    "kalite": [
        ("Kritik sureclere ic tetkik (haftalik) ve NCR/CCR takibi", 7),
        ("ITP’lerde muayene tutanaklari dijital arsive islesin", 10),
    ],
    "pmo": [
        ("Paydas haritasi ve iletisim frekansi (RACI ile hizali) guncellensin", 7),
        ("Aylik proje performans raporu (KPI/Trend) standardize edilsin", 10),
    ],
    "planlama": [
        ("Kritik yol (CPM) ve kaynak yukleri yeniden hesaplanip yayimlansin", 7),
        ("Hava/deniz kosullari icin program tamponlari (float) revize edilsin", 5),
        ("Gecikme nedenleri analizi ve toparlama plani (recovery) paylasilsin", 10),
    ],
    "mep_elektrik": [
        ("Test & Devreye Alma (T&C) planlari ve checklist’leri yayinlansin", 7),
        ("Topraklama/izolasyon (megger) testleri takvime baglansin", 10),
        ("Kritik ekipman icin yedek parca/stok plani olussun", 14),
    ],
    "mep_mekanik": [
        ("Hidrostatik/basinç test programi ve kabul kriterleri netlestsin", 7),
        ("Komisyoning sirasi (HVAC balancing vb.) planla ve ekip ata", 10),
        ("Yangin hatlari icin devreye alma proseduru ve tatbikat", 14),
    ],
    "marine": [
        ("Deniz calismalari icin metocean pencereleri ve izinler teyit", 5),
        ("Barge/vinc rigging planlari ve emniyet brifingi", 7),
        ("Batimetri/posizyonlama kayitlari gunluk arsivlensin", 10),
    ],
    "tasarim": [
        ("RFI/Shop drawing akisi ve onay SLA’lari netlestsin", 7),
        ("Clash detection (Navis) raporu ve cozum takip listesi", 10),
    ],
    "teknik_ofis": [
        ("Metraj-BOQ eslestirme ve fark analizi (variance) yayinlansin", 7),
        ("Hak edis dokumantasyonu (atasman/foto) standardize edilsin", 10),
    ],
    "finans": [
        ("Aylik nakit akis projeksiyonu ve sapma analizi (EV/MS) paylas", 7),
        ("Teminat/avans/kesinti takvimleri risk matrisi ile hizalansin", 10),
    ],
    "makine_bakim": [
        ("Periyodik bakim planlari (OEM) CMMS’e islenip hatirlatici ac", 7),
        ("Kritik ekipman icin ariza MTBF/MTTR KPI’lari takip edilsin", 10),
    ],
    "bim_bt": [
        ("Model versiyonlama ve yedekleme politikalari uygulanir olsun", 7),
        ("IFC cikti standartlari ve clash threshold degerleri sabitlensin", 10),
    ],
    "izin_ruhsat": [
        ("Ruhsat/izin takip matrisi ve sorumlu listesi guncellensin", 5),
        ("Resmi yazisma sablonlari ve dosyalama agaci standardize edilsin", 10),
    ],
    "laboratuvar": [
        ("Numune alma/kur/raporlama zinciri (traceability) garanti altina alınsın", 7),
        ("Cihaz kalibrasyon planlari ve sertifika arsivi kontrol edilsin", 10),
    ],
    "depo": [
        ("Stok sayim ve emniyet stogu esik degerleri (min/max) tanimlansin", 7),
        ("Giris-cikis ve lot/seri takibi icin barkod/etiket duzeni kurulsun", 10),
    ],
}

def _match_keys(text: str):
    """Metni KEYSETS'e gore tarar, eslesen anahtar listesi dondurur."""
    hits = []
    for key, kw in KEYSETS.items():
        if _any_in(text, kw):
            hits.append(key)
    return hits

def _dept_raci_defaults(cat_lower: str):
    """
    Kategori ipuçlarına göre ilgili departmanları ve tipik RACI rollerini öner.
    R: Responsible, A: Accountable, C: Consulted, I: Informed
    """
    rules = [
        (["beton","kalıp","donatı","döküm","vibratör","şantiye","saha","imalat","betoniyer","fore kazık","tünel","kalıp iskelesi",
          "betonarme","yapı","uygulama","derz","kür","scaffold","formwork","rebar","pour","site"],
         {"dept":"İnşaat/Şantiye","R":"Saha Şefi","A":"Proje Müdürü","C":["Kalite Müh.","Planlama"],"I":["İSG","Satınalma"]}),
        (["satınalma","tedarik","malzeme","lojistik","irsaliye","sipariş","po","rfq","tür","tyr","tedarikçi","nakliye","kargo","sevkiyat",
          "logistics","procurement","purchase","supply","warehouse"],
         {"dept":"Satınalma/Lojistik","R":"Satınalma Uzmanı","A":"Satınalma Müdürü","C":["İnşaat","Kalite"],"I":["Finans","Depo"]}),
        (["sözleşme","legal","hukuk","onay","izin","reg","regülasyon","yasal","idari şartname","teknik şartname","claim","hak talebi","itiraz",
          "contract","subcontract","variation","vo","ek protokol"],
         {"dept":"Sözleşme/Hukuk","R":"Sözleşme Uzmanı","A":"Hukuk Müdürü","C":["Proje Müdürü","Satınalma"],"I":["İşveren","Paydaşlar"]}),
        (["isg","iş güvenliği","kaza","ramak kala","çevre","emisyon","atık","toz","gürültü","ppé","ppe","risk analizi","acil durum",
          "çed","cevre","emission","waste","noise","spill"],
         {"dept":"İSG/Çevre","R":"İSG/Çevre Müh.","A":"İSG Müdürü","C":["Şantiye","Kalite"],"I":["İşveren","Yerel Otorite"]}),
        (["zemin","geoteknik","kazı","iksa","zayıf zemin","oturma","şev","şev stabilitesi","cpt","spt","sonder","forekazık","ankraj"],
         {"dept":"Geoteknik","R":"Geoteknik Müh.","A":"Teknik Ofis Müd.","C":["Şantiye","Kalite"],"I":["Danışman"]}),
        (["kalite","denetim","tetkik","audit","muayene","itp","mür","mür onayı","test planı","karot","numune","slump","ndt",
          "wps","pqr","welder","kalibrasyon","inspection","hold point","surveillance"],
         {"dept":"Kalite (QA/QC)","R":"Kalite Müh.","A":"Kalite Müdürü","C":["Şantiye","Sözleşme"],"I":["İşveren","Danışman"]}),
        (["politik","organizasyonel","paydaş","stakeholder","iletişim planı","raporlama","kpi","yönetim kurulu","koordinasyon","komite"],
         {"dept":"PMO/Paydaş Yönetimi","R":"PMO Uzmanı","A":"Proje Müdürü","C":["Hukuk","İletişim"],"I":["İşveren","Yerel Yönetim"]}),
        (["planlama","program","zaman çizelgesi","kritik yol","cpm","ms project","primavera","p6","gant","hava","rüzgar","yağış","sıcaklık",
          "hava durumu","weather","wind","delay","erteleme"],
         {"dept":"Planlama","R":"Planlama Uzmanı","A":"Proje Müdürü","C":["Şantiye","İSG"],"I":["İşveren"]}),
        (["elektrik","og","ag","trafo","kumanda","scada","pano","kablo","trays","aydınlatma","topraklama","kesici","jenerator","ups",
          "elektrifikasyon","test devreye alma","energize","megger","loop test"],
         {"dept":"MEP/Elektrik","R":"Elektrik Şefi","A":"MEP Müdürü","C":["Kalite","Planlama"],"I":["Satınalma","İşveren"]}),
        (["mekanik","hvac","chiller","kazan","pompa","yangın","sprinkler","tesisat","borulama","pnid","basınç testi","hidrostatik","commissioning",
          "duct","blower","valf","kolektör","eşanjör"],
         {"dept":"MEP/Mekanik","R":"Mekanik Şefi","A":"MEP Müdürü","C":["Kalite","Planlama"],"I":["Satınalma","İşveren"]}),
        (["deniz","marine","rıhtım","iskele","kazık","çelik kazık","dolfen","şamandıra","batimetri","akıntı","dalga","römorkör","barge","vinç barge",
          "fener","şamandıra","mendirek","dalgakıran","rıhtım kreni"],
         {"dept":"Deniz/Marine İşleri","R":"Marine Şantiye Şefi","A":"Deniz Yapıları Müdürü","C":["Geoteknik","Kalite"],"I":["Liman Başkanlığı","Kıyı Emniyeti"]}),
        (["tasarım","çizim","revizyon","ifc","shop drawing","shopdrawing","statik","mimari","koordine","clash","detay","kesit","proje onayı","rfı","rfi"],
         {"dept":"Tasarım/Statik-Mimari","R":"Tasarım Koordinatörü","A":"Teknik Ofis Müd.","C":["MEP","Kalite"],"I":["Danışman","İşveren"]}),
        (["teknik ofis","metraj","hakediş","ataşman","boq","keşif","birim fiyat","poz","revize keşif","progress","maliyet analizi","yıllık plan"],
         {"dept":"Teknik Ofis","R":"Teknik Ofis Müh.","A":"Teknik Ofis Müd.","C":["Planlama","Sözleşme"],"I":["Finans","Şantiye"]}),
        (["finans","bütçe","nakit akışı","cash flow","fatura","tahsilat","teminat","kesinti","avans","kur riski","maliyet","capex","opex"],
         {"dept":"Finans/Bütçe","R":"Finans Uzmanı","A":"Finans Müdürü","C":["Teknik Ofis","Satınalma"],"I":["Proje Müdürü"]}),
        (["ekipman","makine","bakım","arıza","yedek parça","operatör","vinç","excavator","loader","forklift","servis","kalibrasyon","periyodik kontrol",
          "lifting plan","rigging","winch"],
         {"dept":"Makine-Bakım","R":"Bakım Şefi","A":"Makine/Ekipman Müdürü","C":["İSG","Şantiye"],"I":["Satınalma","Depo"]}),
        (["bim","model","revit","navisworks","ifc dosyası","clash detection","veri tabanı","sunucu","yedekleme","network","cad","gis"],
         {"dept":"BIM/BT","R":"BIM Uzmanı","A":"BIM/BT Müdürü","C":["Tasarım","Planlama"],"I":["Tüm Birimler"]}),
        (["ruhsat","izin","belediye","imar","fenni mesul","asgari şantiye","tutanak","tutanak altı","resmi yazı","dilekçe","enkaz izin","trafik kesme izni"],
         {"dept":"İzin/Ruhsat","R":"Resmi İşler Sorumlusu","A":"Proje Müdürü","C":["Hukuk","PMO"],"I":["Yerel Otorite","İşveren"]}),
        (["laboratuvar","numune","slump","karot","çekme testi","basınç testi","yol çekici","agrega","granülometri","çelik çekme","ndt","ultrasonik test"],
         {"dept":"Laboratuvar/Test","R":"Lab Teknisyeni","A":"Kalite Müdürü","C":["Şantiye","Geoteknik"],"I":["Danışman","İşveren"]}),
        (["depo","ambar","stok","stok sayım","emniyet stoğu","raf","malzeme teslim","giriş çıkış","irsaliye kontrol","stok devir","ambar fişi"],
         {"dept":"Depo/Ambar","R":"Depo Sorumlusu","A":"Lojistik/Depo Müdürü","C":["Satınalma","Kalite"],"I":["Finans","Şantiye"]}),
        (["hava durumu","hava","rüzgar","yağış","sıcaklık","fırtına","dalga","akıntı","visibility","sis","weather","wind","rain","temperature","storm"],
         {"dept":"Planlama","R":"Planlama Uzmanı","A":"Proje Müdürü","C":["Şantiye","İSG","Deniz/Marine İşleri"],"I":["İşveren"]}),
    ]

    for keys, cfg in rules:
        if any(k in cat_lower for k in keys):
            return cfg
    # genel varsayılan
    return {"dept":"Proje Yönetimi", "R":"Risk Sahibi", "A":"Proje Müdürü", "C":["Kalite","Planlama"], "I":["İSG","Satınalma"]}

def _propose_actions(risk: "Risk"):
    """
    Her aksiyon: {dept, R, A, C, I, action, due}
    base RACI: _dept_raci_defaults(cat)
    """
    cat_raw = (risk.category or "")
    base = _dept_raci_defaults(cat_raw)

    matched = _match_keys(cat_raw)
    actions = []

    # Eslesme yoksa genel set
    if not matched:
        actions += [
            {**base, "action": "Risk icin ayrintili metod beyanı ve kontrol listesi hazirlanmasi", "due": _smart_due(7)},
            {**base, "action": "Haftalik izleme formu ac; trend/KPI takibi baslasin",               "due": _smart_due(7)},
        ]
        return actions

    # Eslesmelerin aksiyonlarini topla (en fazla 8 aksiyon, tekrar sil)
    MAX_ACTIONS = 8
    for key in matched:
        for text, days in ACTION_TEMPLATES.get(key, []):
            actions.append({**base, "action": text, "due": _smart_due(days)})
            if len(actions) >= MAX_ACTIONS:
                break
        if len(actions) >= MAX_ACTIONS:
            break

    return _unique(actions)

def _kpis_default(cat_lower: str):
    cat_lower = _normalize(cat_lower)

    common = [
        "Uygunsuzluk (NCR) sayisi = 0 / ay",
        "Rework saatleri ≤ toplam isçilik saatinin %2’si",
    ]

    if "beton" in cat_lower or "kalip" in cat_lower or "donati" in cat_lower or _any_in(cat_lower, KEYSETS["insaat"]):
        return common + [
            "Beton basinç testi basarisizlik orani ≤ %1",
            "Slump/sicaklik tolerans disi orani ≤ %2",
        ]
    if _any_in(cat_lower, KEYSETS["satinalma"]):
        return common + [
            "OTD (On-Time Delivery) ≥ %95",
            "Emniyet stogu altina dusus olay sayisi = 0 / ay",
        ]
    if _any_in(cat_lower, KEYSETS["sozlesme"]):
        return common + [
            "Kritik izin/onay gecikmesi = 0",
            "Sozlesme ihlal/NCR sayisi = 0",
        ]
    if _any_in(cat_lower, KEYSETS["isg_cevre"]):
        return common + [
            "Toz/gurultu limit asimlari = 0",
            "Atik bertaraf uygunsuzlugu = 0",
        ]
    if _any_in(cat_lower, KEYSETS["geoteknik"]):
        return common + [
            "Sev stabilitesi ihlal (trigger asimi) = 0",
            "Zemin parametre guncelleme gecikmesi = 0",
        ]
    if _any_in(cat_lower, KEYSETS["kalite"]):
        return common + [
            "NCR kapama ort. suresi ≤ 10 gun",
            "ITP adim uyum orani ≥ %98",
        ]
    if _any_in(cat_lower, KEYSETS["planlama"]):
        return common + [
            "Kritik faaliyet gecikme orani ≤ %3",
            "Gantt/P6 haftalik guncelleme tamamlama orani = %100",
        ]
    if _any_in(cat_lower, KEYSETS["mep_elektrik"]):
        return common + [
            "Izolasyon (megger) test basari orani ≥ %99",
            "T&C (elektrik) punch sayisi ≤ 5 / alan",
        ]
    if _any_in(cat_lower, KEYSETS["mep_mekanik"]):
        return common + [
            "Hidrostatik/basinç test basari orani ≥ %99",
            "HVAC balancing sapma ≤ %5",
        ]
    if _any_in(cat_lower, KEYSETS["marine"]):
        return common + [
            "Metocean pencere disi calisma olayi = 0",
            "Barge/rigging plan uygunsuzlugu = 0",
        ]
    if _any_in(cat_lower, KEYSETS["tasarim"]):
        return common + [
            "RFI ort. kapanma suresi ≤ 7 gun",
            "Shop drawing onay zamaninda tamamlama ≥ %95",
        ]
    if _any_in(cat_lower, KEYSETS["teknik_ofis"]):
        return common + [
            "Metraj–BOQ fark orani ≤ %1",
            "Hak edis teslim gecikmesi = 0",
        ]
    if _any_in(cat_lower, KEYSETS["finans"]):
        return common + [
            "Nakit akis sapma (plan vs gercek) ≤ %5",
            "Fatura gecikme orani ≤ %2",
        ]
    if _any_in(cat_lower, KEYSETS["makine_bakim"]):
        return common + [
            "MTBF artisi (aylik) ≥ %5",
            "Planli bakim gerceklesme orani ≥ %95",
        ]
    if _any_in(cat_lower, KEYSETS["bim_bt"]):
        return common + [
            "Clash sayisi (kritik) ≤ X/hafta (hedef belirlenmeli)",
            "Model versiyonlari yedekleme uyumu = %100",
        ]
    if _any_in(cat_lower, KEYSETS["izin_ruhsat"]):
        return common + [
            "Kritik izin gecikmesi = 0",
            "Resmi yazisma SLA uyum orani ≥ %95",
        ]
    if _any_in(cat_lower, KEYSETS["laboratuvar"]):
        return common + [
            "Numune izlenebilirlik (traceability) hatasi = 0",
            "Kalibrasyon gecikmesi = 0",
        ]
    if _any_in(cat_lower, KEYSETS["depo"]):
        return common + [
            "Stok sayim uyumsuzluk orani ≤ %1",
            "Lot/seri izlenebilirlik hatasi = 0",
        ]

    return common


    

def send_email(to_email: str, subject: str, body: str):
    """
    Güvenli ve UTF-8 uyumlu SMTP mail gönderimi.
    ENV:
      SMTP_HOST, SMTP_PORT, SMTP_USER, SMTP_PASS, SMTP_FROM
      SMTP_TLS=1 -> STARTTLS (genelde 587)
      SMTP_TLS=0 -> SSL/TLS   (genelde 465)
      SMTP_DEBUG=1 -> SMTP diyaloğunu logla
      SMTP_FALLBACK=1 -> bir mod başarısızsa diğerini dene
    """
    import os, smtplib
    from email.message import EmailMessage

    host = os.getenv("SMTP_HOST")
    # Port belirtilmemişse moda göre mantıklı varsayılan ver
    use_tls = os.getenv("SMTP_TLS", "").lower() in ("1", "true", "yes")
    port = int(os.getenv("SMTP_PORT", "0") or "0")
    if port == 0:
        port = 587 if use_tls else 465

    user = os.getenv("SMTP_USER")
    pwd  = os.getenv("SMTP_PASS")
    from_addr = os.getenv("SMTP_FROM", user or "no-reply@example.com")
    debug_on = os.getenv("SMTP_DEBUG", "").lower() in ("1","true","yes")
    do_fallback = os.getenv("SMTP_FALLBACK", "").lower() in ("1","true","yes")

    if not host or not port:
        msg = f"[MAIL-ERROR] SMTP config eksik (SMTP_HOST/SMTP_PORT). To={to_email} Subject={subject}"
        print(msg)
        return False, msg

    # Mesajı UTF-8 olarak hazırla (Türkçe karakterler sorunsuz)
    msg = EmailMessage()
    msg["From"] = from_addr
    msg["To"] = to_email
    msg["Subject"] = subject
    msg.set_content(body)  # text/plain; charset="utf-8"

    def _send_starttls():
        with smtplib.SMTP(host, port, timeout=25) as s:
            if debug_on: s.set_debuglevel(1)
            s.ehlo()
            s.starttls()
            s.ehlo()
            if user and pwd:
                s.login(user, pwd)
            s.send_message(msg)

    def _send_ssl():
        with smtplib.SMTP_SSL(host, port, timeout=25) as s:
            if debug_on: s.set_debuglevel(1)
            if user and pwd:
                s.login(user, pwd)
            s.send_message(msg)

    try:
        if use_tls:
            _send_starttls()
        else:
            _send_ssl()
        print(f"[MAIL] sent to {to_email} subj={subject}")
        return True, ""
    except Exception as e1:
        print(f"[MAIL-ERROR] primary send failed -> {e1}")
        if do_fallback:
            try:
                if use_tls:
                    _send_ssl()      # STARTTLS başarısızsa SSL'e düş
                else:
                    _send_starttls() # SSL başarısızsa STARTTLS'e düş
                print(f"[MAIL] sent (fallback) to {to_email} subj={subject}")
                return True, ""
            except Exception as e2:
                msg = f"[MAIL-ERROR] fallback failed -> {e2}"
                print(msg)
                return False, f"{e1} | FALLBACK: {e2}"
        return False, str(e1)


# -------------------------------------------------
#  Flask uygulaması oluştur
# -------------------------------------------------
def create_app():
    app = Flask(__name__)
    app.config["SECRET_KEY"] = os.getenv("SECRET_KEY", "dev-secret-change-me")

    # 1) DB URI önceliği
    default_sqlite_uri = "sqlite:////tmp/riskapp.db"
    db_uri = (os.getenv("DATABASE_URI") or os.getenv("DATABASE_URL") or default_sqlite_uri).strip()

    # Render bazen postgres:// döndürür; SQLAlchemy postgresql+psycopg2:// ister
    if db_uri.startswith("postgres://"):
        db_uri = db_uri.replace("postgres://", "postgresql+psycopg2://", 1)

    app.config["SQLALCHEMY_DATABASE_URI"] = db_uri
    app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
    app.config["CONSENSUS_THRESHOLD"] = 30

    # 2) SQLite ise: thread ayarı + dosya/klasör garantisi
    if db_uri.startswith("sqlite:"):
        engine_opts = app.config.setdefault("SQLALCHEMY_ENGINE_OPTIONS", {})
        conn_args = engine_opts.setdefault("connect_args", {})
        conn_args.setdefault("check_same_thread", False)

        # URI'den path çıkar (sqlite:////tmp/x.db -> //tmp/x.db gibi gelebilir)
        raw_path = urlparse(db_uri).path or "/tmp/riskapp.db"
        db_path = os.path.normpath(raw_path)

        # Güvensiz / yazılamayan yerlere düşerse /tmp'ye kaç
        unsafe_dirs = {"", "/", "/data", "//data"}
        dir_path = os.path.dirname(db_path)

        def _fallback_to_tmp():
            return "/tmp/riskapp.db", "/tmp"

        # root veya saçma dizinler
        if (not dir_path) or (dir_path in unsafe_dirs):
            db_path, dir_path = _fallback_to_tmp()

        # klasörü oluşturmayı dene + dosyayı yoksa yarat
        try:
            os.makedirs(dir_path, exist_ok=True)

            # klasör var ama yazılamıyorsa fallback
            if not os.access(dir_path, os.W_OK):
                db_path, dir_path = _fallback_to_tmp()
                os.makedirs(dir_path, exist_ok=True)

            with open(db_path, "a", encoding="utf-8"):
                pass
        except Exception:
            db_path, dir_path = _fallback_to_tmp()
            os.makedirs(dir_path, exist_ok=True)
            with open(db_path, "a", encoding="utf-8"):
                pass

        # SQLAlchemy URI'sini normalize edip geri yaz
        app.config["SQLALCHEMY_DATABASE_URI"] = f"sqlite:///{db_path}"

    # 3) DB init
    db.init_app(app)

    # 4) Şema / seed / indexler (tek noktadan, stabil sırayla)
    def bootstrap_db():
        uri = app.config.get("SQLALCHEMY_DATABASE_URI", "") or ""

        # Tablolar
        db.create_all()

        # SQLite için geriye dönük schema fixleri
        if uri.startswith("sqlite:"):
            ensure_schema()

        # Seed (istersen env ile kapat)
        if os.environ.get("SKIP_SEED") != "1":
            try:
                seed_if_empty()
            except OperationalError as e:
                app.logger.warning("Seed atlandı (DB şeması hazır değil): %s", e)

        # Performans indeksleri (idempotent)
        try:
            db.session.execute(text("CREATE INDEX IF NOT EXISTS ix_risks_project ON risks(project_id)"))
            db.session.execute(text("CREATE INDEX IF NOT EXISTS ix_risks_start   ON risks(start_month)"))
            db.session.execute(text("CREATE INDEX IF NOT EXISTS ix_risks_end     ON risks(end_month)"))

            # Ref No benzersizliği (kolon varsa iş görür)
            db.session.execute(text("CREATE UNIQUE INDEX IF NOT EXISTS ux_risks_ref_code ON risks(ref_code)"))

            db.session.commit()
        except Exception as e:
            db.session.rollback()
            app.logger.warning("Index create atlandı: %s", e)

    with app.app_context():
        bootstrap_db()

    def _sync_mitigations(risk: "Risk") -> None:
        """
        Formdan gelen mitigasyon/önlem satırlarını al,
        eski kayıtları sil, yenilerini ekle.
        """

        # Eski mitigasyonları sil
        Mitigation.query.filter_by(risk_id=risk.id).delete()

        # Formdan listeleri çek (hem []'li hem []'siz isimleri kabul et)
        texts   = request.form.getlist("mit_text[]")   or request.form.getlist("mit_text")
        owners  = request.form.getlist("mit_owner[]")  or request.form.getlist("mit_owner")
        dues    = request.form.getlist("mit_due[]")    or request.form.getlist("mit_due")
        status_ = request.form.getlist("mit_status[]") or request.form.getlist("mit_status")

        if not any([texts, owners, dues, status_]):
            return

        n = max(len(texts), len(owners), len(dues), len(status_))

        def _safe(lst, i, default=""):
            return lst[i] if i < len(lst) else default

        for i in range(n):
            text_val = (_safe(texts, i) or "").strip()
            if not text_val:
                continue

            owner_val  = (_safe(owners, i) or "").strip()
            status_val = (_safe(status_, i) or "").strip()
            due_raw    = (_safe(dues, i) or "").strip()

            due_date = None
            if due_raw:
                try:
                    due_date = datetime.strptime(due_raw, "%Y-%m-%d").date()
                except ValueError:
                    due_date = None

            m = Mitigation(
                risk_id=risk.id,
                text=text_val,
                owner=owner_val,      # model alan adın farklıysa burayı değiştir
                status=status_val,    # model alan adın farklıysa burayı değiştir
                due_date=due_date,    # model alan adın farklıysa burayı değiştir
            )
            db.session.add(m)

    # 🔼🔼🔼 BURADA BİTİYOR, SONRA ROUTE’LAR BAŞLIYOR 🔼🔼🔼
    

    

   
    def _build_suggestions_by_category(category_rows):
        """
        RiskCategory satırlarından -> { "cat_id": [ {text, risk_code, default_prob, default_sev}, ... ] }
        döner. Suggestion.category alanı kategori ADI tuttuğu için adı id’ye map’liyoruz.
        """
        id_to_name = {str(c.id): c.name for c in category_rows}
        name_to_id = {c.name: str(c.id) for c in category_rows}

        try:
            q = (Suggestion.query
                .filter(Suggestion.is_active.is_(True))
                .order_by(Suggestion.category.asc()))
            sug_rows = q.all()
        except Exception:
            sug_rows = []

        out = {}
        for s in sug_rows:
            cat_name = (s.category or "").strip()
            cat_id = name_to_id.get(cat_name)
            if not cat_id:
                continue
            out.setdefault(cat_id, []).append({
                "text": s.text,
                "risk_code": getattr(s, "risk_code", None),
                "default_prob": getattr(s, "default_prob", None),
                "default_sev": getattr(s, "default_sev", None),
            })
        return out
    @app.get("/api/suggestions")
    def api_suggestions():
        """
        ?cat_ids=1,3,7 -> { "1":[{text,...}], "3":[...], ... }

        Notlar:
        - İsim eşleşmesini case-insensitive ve boşluk/ayraç toleranslı yapar.
        - Yalnızca is_active=True olan Suggestion’lar döner.
        """
        cat_ids_param = (request.args.get("cat_ids") or "").strip()
        if not cat_ids_param:
            return jsonify({})

        req_ids = [s for s in cat_ids_param.split(",") if s.strip()]

        # İstenen kategori satırlarını çek
        cats = (RiskCategory.query
                .filter(RiskCategory.is_active.is_(True), RiskCategory.id.in_(req_ids))
                .all())
        if not cats:
            return jsonify({})

        # --- Normalizasyon yardımcıları ---
        import unicodedata as _ud
        def _norm_name(s: str) -> str:
            # Unicode NFC -> casefold -> iç boşlukları tek boşluğa indir
            s = _ud.normalize("NFC", (s or "").strip())
            s = s.casefold()
            s = " ".join(s.split())
            # " / " ve "/" varyasyonlarını aynılaştır
            s = s.replace(" / ", "/").replace(" /", "/").replace("/ ", "/")
            return s

        # id <-> name haritaları
        id_to_name = {str(c.id): (c.name or "").strip() for c in cats}
        id_to_norm = {cid: _norm_name(nm) for cid, nm in id_to_name.items()}
        norm_to_id = {v: k for k, v in id_to_norm.items()}

        # Aranacak isim seti (lower/casefold)
        from sqlalchemy import func
        target_norms = list(id_to_norm.values())
        # LOWER karşılaştırması için “orijinal” varyasyonları da ekleyelim
        target_lowers = [t.lower() for t in target_norms]

        # Veritabanından sadece ilgili kategorileri çek (case-insensitive)
        # Not: LOWER(category) IN (:lower1, :lower2, ...)
        q = (Suggestion.query
            .filter(Suggestion.is_active.is_(True))
            .filter(func.lower(Suggestion.category).in_(target_lowers))
            .order_by(Suggestion.category.asc(), Suggestion.id.desc()))
        rows = q.all()

        out = {str(cid): [] for cid in id_to_name.keys()}
        for s in rows:
            key = _norm_name(s.category)
            cid = norm_to_id.get(key)
            if not cid:
                # Çok nadir: "SÖZLEŞME / ONAY SÜREÇLERİ" vs "sözleşme/onay süreçleri"
                # yine de yakalayamadıysak bir “yakın eşleşme” deneriz:
                for k_norm, k_id in norm_to_id.items():
                    if key.replace(" ", "") == k_norm.replace(" ", ""):
                        cid = k_id
                        break
            if cid:
                out.setdefault(cid, []).append({
                    "text": s.text,
                    "risk_code": getattr(s, "risk_code", None),
                    "default_prob": getattr(s, "default_prob", None),
                    "default_sev": getattr(s, "default_sev", None),
                })

        return jsonify(out)


    # -------------------------------------------------
    #  Yetki kontrol dekoratörü
    # -------------------------------------------------
    def role_required(role):
        def decorator(fn):
            @wraps(fn)
            def wrapper(*args, **kwargs):
                if "username" not in session:
                    return redirect(url_for("login"))
                if session.get("role") != role:
                    flash("Bu işlemi yapmak için yetkiniz yok.", "danger")
                    return redirect(url_for("dashboard"))
                return fn(*args, **kwargs)
            return wrapper
        return decorator
    
    @app.before_request
    def require_login():
        # Giriş gerektirmeyen endpoint'ler (endpoint adları)
        allowed = {"static", "login", "setup_step1", "forgot_password", "health"}
        ep = (request.endpoint or "")

        # (Opsiyonel) Herkese açık bırakmak istediğin API endpoint'leri (endpoint adları)
        public_api = {
            # "api_category_names",  # örnek: /api/category-names herkese açık olsun istiyorsan yorumdan çıkar
        }

        # --- API çağrıları: /api/... veya api_* endpoint'leri için 401 JSON döndür ---
        if request.path.startswith("/api/") or ep.startswith("api_"):
            if ep in public_api:
                return  # public API -> oturum şartı yok
            if "username" not in session:
                return jsonify({"error": "unauthorized"}), 401
            return  # oturum varsa devam

        # --- Web sayfaları için klasik redirect ---
        if "username" not in session and (ep not in allowed):
            return redirect(url_for("login"))
        
    


    # -------------------------------------------------
    #  Şifre Sıfırlama
    # -------------------------------------------------
    @app.route("/forgot", methods=["GET", "POST"], endpoint="forgot_password")
    def forgot_password():
        if request.method == "POST":
            email = request.form.get("email", "").strip()
            new_pw = request.form.get("new_password", "")
            new_pw2 = request.form.get("new_password2", "")

            if not email or not new_pw or not new_pw2:
                flash("Lütfen tüm alanları doldurun.", "danger")
                return render_template("forgot.html", email=email)
            if new_pw != new_pw2:
                flash("Yeni şifreler eşleşmiyor.", "danger")
                return render_template("forgot.html", email=email)

            acc = Account.query.filter_by(email=email).first()
            if not acc:
                flash("Bu e-posta ile kayıt bulunamadı.", "danger")
                return render_template("forgot.html", email=email)

            acc.password_hash = generate_password_hash(new_pw)
            db.session.commit()
            flash("Şifre güncellendi. Şimdi giriş yapabilirsiniz.", "success")
            return redirect(url_for("login"))

        return render_template("forgot.html")

    # -------------------------------------------------
    #  Aktif proje yardımcıları
    # -------------------------------------------------
    def _get_active_project_id():
        """Oturumdaki aktif proje yoksa kullanıcının son projesini ata."""
        pid = session.get("project_id")
        acc_id = session.get("account_id")
        if pid:
            return pid
        if not acc_id:
            return None
        proj = ProjectInfo.query.filter_by(account_id=acc_id).order_by(ProjectInfo.created_at.desc()).first()
        if proj:
            session["project_id"] = proj.id
            return proj.id
        return None

    @app.context_processor
    def inject_models_and_active_project():
        # base.html'de mini proje seçici için
        return {
            "ProjectInfo": ProjectInfo,
            "active_project_id": session.get("project_id")
        }

    # Küçük yardımcı: categories -> identify geri dönüş
    def _should_go_identify():
        if (request.args.get("next") or "").lower() == "identify":
            return True
        if (request.form.get("next") or "").lower() == "identify":
            return True
        ref = request.referrer or ""
        return "next=identify" in ref

    # -------------------------------------------------
    #  Onboarding / Landing
    # -------------------------------------------------
    @app.route("/")
    def index():
        return redirect(url_for("welcome"))

    @app.route("/welcome")
    def welcome():
        return render_template("welcome.html")

    # -------------------------------------------------
    #  Giriş — e-posta + şifre
    # -------------------------------------------------
    @app.route("/login", methods=["GET","POST"])
    def login():
        if Account.query.count() == 0:
            return redirect(url_for("setup_step1"))

        if request.method == "POST":
            email = (request.form.get("email") or "").strip()
            password = request.form.get("password") or ""
            ref_code_input = (request.form.get("ref_code") or "").strip().upper()

            acc = Account.query.filter_by(email=email).first()
            if not acc or not check_password_hash(acc.password_hash, password):
                flash("E-posta veya şifre hatalı.", "danger")
                return render_template("login.html", email=email)

            # Admin kullanıcılar ref kodu girmeden oturum açabilir.
            is_admin = (acc.role or "uzman") == "admin"

            if not is_admin:
                # Uzman/diğer roller için ref kodu ve aktiflik kontrolleri devam
                if (acc.status or "pending") != "active":
                    flash("Hesabınız henüz aktif değil. Admin onayı bekleniyor.", "warning")
                    return render_template("login.html", email=email)
                if not acc.ref_code:
                    flash("Referans kodu atanmadı. Lütfen admin ile iletişime geçin.", "warning")
                    return render_template("login.html", email=email)
                if not ref_code_input:
                    flash("Referans kodu zorunludur.", "danger")
                    return render_template("login.html", email=email)
                if acc.ref_code.strip().upper() != ref_code_input:
                    flash("Referans kodu geçersiz.", "danger")
                    return render_template("login.html", email=email)

            # Buraya geldiysen giriş başarıldı (admin/uzman fark etmez)
            session["account_id"] = acc.id
            session["username"] = acc.contact_name
            session["role"] = acc.role or "uzman"
            flash(f"Hoş geldin, {acc.contact_name}!", "success")
            return redirect(url_for("dashboard"))

        return render_template("login.html")


    @app.route("/logout")
    def logout():
        session.clear()
        return redirect(url_for("welcome"))

    # -------------------------------------------------
    #  Dashboard
    # -------------------------------------------------
    # -------------------------------------------------
#  Dashboard
# -------------------------------------------------
    @app.route("/dashboard")
    def dashboard():
        pid = _get_active_project_id()
        query = Risk.query
        if pid:
            query = query.filter(Risk.project_id == pid)

        risks = query.order_by(Risk.updated_at.desc()).all()

        # --- 5x5 matris (olasılık × şiddet) ---
        # Eski: ortalama P/S kullanıyordu, o yüzden hücreler kayıyordu.
        # Yeni: HER RİSK İÇİN SON Evaluation (en büyük id) alınır,
        #       P ve S direkt o kayıttan okunur, key = "P-S".
        matrix = defaultdict(int)

        for r in risks:
            evals = sorted(r.evaluations or [], key=lambda e: e.id)
            if not evals:
                continue

            last = evals[-1]
            p = last.probability or 0
            s = last.severity or 0

            # P veya S yoksa matrise sokma
            if not p or not s:
                continue

            # 1..5 aralığında bırak (yine de emniyet)
            p = max(1, min(5, int(p)))
            s = max(1, min(5, int(s)))

            key = f"{p}-{s}"
            matrix[key] += 1

        # Jinja'ya sade dict gitsin
        matrix = dict(matrix)

        # --- Kategori bazlı dağılım ---
        # Eşikler UI ile uyumlu:
        # 1–4 Düşük, 5–10 Orta, 11–15 Yüksek, 16–25 Çok Yüksek
        def _score_bucket(sc):
            if sc is None:
                return None
            try:
                sc = float(sc)
            except Exception:
                return None

            if sc >= 16:
                return "vhigh"   # Çok Yüksek
            if sc >= 11:
                return "high"    # Yüksek
            if sc >= 5:
                return "mid"     # Orta
            if sc >= 1:
                return "low"     # Düşük
            return None

        by_cat = defaultdict(
            lambda: {"cat": "", "total": 0, "low": 0, "mid": 0, "high": 0, "vhigh": 0}
        )

        for r in risks:
            cat = (getattr(r, "category", None) or "Genel")

            # r.score() varsa ve sayısal ise onu kullan, yoksa P×S türet
            sc = None
            s_method = getattr(r, "score", None)
            if callable(s_method):
                try:
                    sc = s_method()
                    sc = float(sc) if sc is not None else None
                except Exception:
                    sc = None
            if sc is None:
                try:
                    p2, s2 = r.avg_prob(), r.avg_sev()
                    if p2 and s2:
                        sc = float(p2) * float(s2)
                except Exception:
                    sc = None

            b = _score_bucket(sc)
            row = by_cat[cat]
            row["cat"] = cat
            if b:
                row[b] += 1
                row["total"] += 1

        # Listeyi toplam sayıya göre azalan sırala, sonra ada göre
        category_stats = sorted(by_cat.values(), key=lambda x: (-x["total"], x["cat"]))

        # Toplam satırı ekle – şablonda en alta “Toplam Riskler”
        if category_stats:
            totals = {"cat": "Toplam Riskler", "total": 0, "low": 0, "mid": 0, "high": 0, "vhigh": 0}
            for row in category_stats:
                totals["total"] += row["total"]
                totals["low"]   += row["low"]
                totals["mid"]   += row["mid"]
                totals["high"]  += row["high"]
                totals["vhigh"] += row["vhigh"]
            category_stats.append(totals)

        return render_template(
            "dashboard.html",
            risks=risks,
            matrix=matrix,
            category_stats=category_stats,
        )

    
    @app.get("/admin/refcodes")
    @role_required("admin")
    def admin_refcodes_list():
        rows = db.session.execute(text("""
            SELECT id, code, assigned_email, is_used, created_by, expires_at, created_at
            FROM referral_codes
            ORDER BY is_used ASC, created_at DESC
        """)).fetchall()
        return render_template("admin_refcodes.html", rows=rows)

    @app.post("/admin/refcodes/create")
    @role_required("admin")
    def admin_refcodes_create():
        # Tek kod oluştur (prefix opsiyonel), istersen count ile çoğaltırız.
        prefix = (request.form.get("prefix") or "PRJ").strip().upper()
        expires = (request.form.get("expires_at") or "").strip() or None  # YYYY-MM-DD ya da boş
        code = _gen_ref_code(prefix=prefix)
        db.session.execute(text("""
            INSERT INTO referral_codes (code, assigned_email, is_used, created_by, expires_at)
            VALUES (:code, NULL, 0, :uid, :exp)
        """), {"code": code, "uid": session.get("account_id"), "exp": expires})
        db.session.commit()
        flash(f"Referans kodu üretildi: {code}", "success")
        return redirect(url_for("admin_refcodes_list"))

    @app.post("/admin/refcodes/<int:rid>/delete")
    @role_required("admin")
    def admin_refcodes_delete(rid):
        db.session.execute(text("DELETE FROM referral_codes WHERE id=:i"), {"i": rid})
        db.session.commit()
        flash("Kod silindi.", "success")
        return redirect(url_for("admin_refcodes_list"))

    @app.post("/admin/refcodes/<int:rid>/lock")
    @role_required("admin")
    def admin_refcodes_lock(rid):
        email = (request.form.get("email") or "").strip()
        db.session.execute(text("""
            UPDATE referral_codes SET assigned_email=:e WHERE id=:i
        """), {"e": email or None, "i": rid})
        db.session.commit()
        flash("Kod kilidi güncellendi.", "success")
        return redirect(url_for("admin_refcodes_list"))
    
    

    # -------------------------------------------------
    #  CSV Export – Riskler
    # -------------------------------------------------
    # === XLSX Risk Analizi (biçimli) ===
    # === XLSX Risk Analizi (biçimli) ===
    # === XLSX Risk Analizi (biçimli) ===
    @app.route("/risks/export.xlsx")
    def risks_export_xlsx():
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except Exception:
            flash("Excel dışa aktarmak için 'openpyxl' gerekli.", "danger")
            return redirect(url_for("risk_select"))

        # SQLAlchemy func lazımsa (çoğu projede zaten global import ediliyor ama garanti olsun)
        from sqlalchemy import func

        pid    = _get_active_project_id()
        q      = (request.args.get("q") or "").strip()
        status = (request.args.get("status") or "").strip()
        title  = (request.args.get("title") or "DENİZ YAPILARI İNŞAAT PROJESİ RİSK ANALİZİ").strip()

        query = Risk.query
        if pid:
            query = query.filter(Risk.project_id == pid)
        if q:
            like = f"%{q}%"
            query = query.filter(
                (Risk.title.ilike(like)) |
                (Risk.category.ilike(like)) |
                (Risk.description.ilike(like))
            )
        if status:
            query = query.filter(Risk.status == status)

        # kategori -> kayıtlar
        risks = query.order_by(Risk.category.asc().nullsfirst(), Risk.id.asc()).all()

        # ---------------------------------------------------------
        # ✅ MALİYET TOPLAMLARI (tek sorgu)
        # ---------------------------------------------------------
        risk_ids = [r.id for r in risks]
        cost_map = {}  # {risk_id: {"TRY": 123, "USD": 0, "EUR": 0}}

        if risk_ids:
            rows = (
                db.session.query(
                    CostItem.risk_id,
                    func.coalesce(CostItem.currency, "TRY").label("cur"),
                    func.coalesce(func.sum(CostItem.total), 0).label("sum_total"),
                )
                .filter(CostItem.risk_id.in_(risk_ids))
                .group_by(CostItem.risk_id, "cur")
                .all()
            )

            for rid, cur, total in rows:
                cur = (cur or "TRY").upper()
                cost_map.setdefault(rid, {})
                cost_map[rid][cur] = float(total or 0)

        # kategori bucket
        buckets: dict[str, list[Risk]] = {}
        for r in risks:
            buckets.setdefault((r.category or "GENEL RİSKLER").strip(), []).append(r)

        # ---- Excel ---
        wb = Workbook()
        ws = wb.active
        ws.title = "Risk Analizi"

        thin = Side(style="thin", color="808080")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        H    = Font(bold=True, size=12)
        HBIG = Font(bold=True, size=14)
        HCAT = Font(bold=True, size=11)
        AL   = Alignment(vertical="center", horizontal="left", wrap_text=True)
        AC   = Alignment(vertical="center", horizontal="center", wrap_text=True)

        FILL_LOW   = PatternFill("solid", fgColor="92D050")  # yeşil
        FILL_MED   = PatternFill("solid", fgColor="FFFF00")  # sarı
        FILL_HIGH  = PatternFill("solid", fgColor="FFC000")  # turuncu
        FILL_VHIGH = PatternFill("solid", fgColor="FF0000")  # kırmızı
        FILL_CAT   = PatternFill("solid", fgColor="E6E6E6")  # kategori satırı
        FILL_HEAD  = PatternFill("solid", fgColor="D9D9D9")  # tablo başlık

        def level_for_rpn(rpn: float | None):
            """1–4 Düşük, 5–10 Orta, 11–15 Yüksek, 16–25 Çok Yüksek."""
            if rpn is None:
                return "", None
            r = float(rpn)
            if r <= 4:
                return "Düşük", FILL_LOW
            if r <= 10:
                return "Orta", FILL_MED
            if r <= 15:
                return "Yüksek", FILL_HIGH
            return "Çok Yüksek", FILL_VHIGH

        # ✅ HEAD’e maliyet kolonlarını ekledik
        HEAD = [
            "No", "Risk Adı", "Risk Tanımlaması", "Risk Sahibi",
            "P", "S", "D", "Risk Seviyesi", "Karşı Önlemler",
            "Maliyet (TRY)", "Maliyet (USD)", "Maliyet (EUR)"
        ]

        # ✅ genişlikler güncellendi
        widths = [5, 22, 48, 18, 6, 6, 6, 16, 42, 14, 14, 14]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # maliyet kolon indexleri (Excel format için)
        COL_COST_TRY = HEAD.index("Maliyet (TRY)") + 1
        COL_COST_USD = HEAD.index("Maliyet (USD)") + 1
        COL_COST_EUR = HEAD.index("Maliyet (EUR)") + 1

        row = 1
        # büyük başlık
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(HEAD))
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = HBIG
        cell.alignment = AC
        row += 2

        # ===== Legend (sağ üst, yatay) =====
        base_col = len(HEAD) + 2
        ws.cell(row=1, column=base_col, value="Legend").font = H

        legend = [
            ("Çok Yüksek Risk", FILL_VHIGH),
            ("Yüksek Risk",     FILL_HIGH),
            ("Orta Risk",       FILL_MED),
            ("Düşük Risk",      FILL_LOW),
        ]

        row_legend = 2
        col = base_col
        for text, fill in legend:
            col += 1
            c = ws.cell(row=row_legend, column=col, value=text)
            c.alignment = AC
            c.fill = fill
            c.border = border
            ws.column_dimensions[get_column_letter(col)].width = max(len(text) + 4, 16)

        # her kategori için blok
        for cat, items in buckets.items():
            # kategori bandı
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(HEAD))
            kc = ws.cell(row=row, column=1, value=f"Risk Kategorisi : {cat}")
            kc.font = HCAT
            kc.fill = FILL_CAT
            kc.alignment = AL
            kc.border = border
            row += 1

            # tablo başlıkları
            for col_idx, head in enumerate(HEAD, 1):
                c = ws.cell(row=row, column=col_idx, value=head)
                c.font = H
                c.fill = FILL_HEAD
                c.alignment = AC
                c.border = border
            row += 1

            # satırlar
            for idx, r in enumerate(items, 1):
                # --- SON değerlendirme P/S ---
                last_eval = None
                if r.evaluations:
                    last_eval = sorted(r.evaluations, key=lambda e: e.id)[-1]

                if last_eval and last_eval.probability is not None and last_eval.severity is not None:
                    p_val = float(last_eval.probability)
                    s_val = float(last_eval.severity)
                else:
                    p_val = r.avg_prob()
                    s_val = r.avg_sev()

                # --- RPN: score() varsa onu kullan, yoksa P×S ---
                sc = None
                score_fn = getattr(r, "score", None)
                if callable(score_fn):
                    try:
                        sc = score_fn()
                        sc = float(sc) if sc is not None else None
                    except Exception:
                        sc = None
                if sc is None and p_val is not None and s_val is not None:
                    sc = float(p_val) * float(s_val)

                lvl_txt, lvl_fill = level_for_rpn(sc)

                # ✅ maliyetler
                cm = cost_map.get(r.id, {})
                c_try = float(cm.get("TRY", 0) or 0)
                c_usd = float(cm.get("USD", 0) or 0)
                c_eur = float(cm.get("EUR", 0) or 0)

                values = [
                    idx,
                    (r.title or ""),
                    (r.description or ""),
                    (r.responsible or ""),
                    (round(p_val, 2) if p_val is not None else ""),
                    (round(s_val, 2) if s_val is not None else ""),
                    "",  # D kullanılmıyor
                    lvl_txt,
                    (r.mitigation or ""),
                    c_try,
                    c_usd,
                    c_eur,
                ]

                for col_idx, val in enumerate(values, 1):
                    c = ws.cell(row=row, column=col_idx, value=val)
                    c.alignment = AL if col_idx in (2, 3, 9) else AC
                    c.border = border

                    if col_idx == 8 and lvl_fill:
                        c.fill = lvl_fill

                    # ✅ maliyet kolonları sayı formatı
                    if col_idx in (COL_COST_TRY, COL_COST_USD, COL_COST_EUR):
                        c.number_format = '#,##0.00'

                row += 1

            row += 1  # kategori sonrası boş satır

        import io
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        fname = f"risk_analizi_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return Response(
            bio.read(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'}
        )


    # -------------------------------------------------
    #  Risk Tanımlama (liste seç)
    # -------------------------------------------------
    @app.route("/identify", methods=["GET", "POST"])
    def risk_identify():
        # -----------------------------
        # 1) Filtre / arama / sayfalama
        # -----------------------------
        q        = (request.args.get("q") or "").strip()
        cat_param_present = ("cat" in request.args)
        cat = request.args.get("cat", "__all__")
        cat = (cat if cat is not None else "__all__").strip()
 
        page     = int(request.args.get("page", 1) or 1)
        per_page = 175  # ihtiyacına göre 25/100 yapabilirsin

        # Kategori dropdown'ı: aktif RiskCategory; yoksa Suggestion'lardan türet
        rcats = (RiskCategory.query
                .filter(RiskCategory.is_active == True)
                .order_by(RiskCategory.name.asc())
                .all())
        filter_cat_names = [ (r.name or "").strip() for r in rcats if (r.name or "").strip() ]

        if not filter_cat_names:
            raw = [x[0] for x in db.session.query(Suggestion.category).distinct().all()]
            filter_cat_names = sorted(
                [r.strip() for r in raw if r and r.strip()],
                key=lambda s: s.lower()
            )



        # -----------------------------
        # 2) Liste sorgusu (Suggestion)
        # -----------------------------
        base_q = Suggestion.query

        # Kategori filtresi
        # Kategori filtresi (Suggestion)
        if cat != "__all__" and cat_param_present:
            if cat == "":  # "Genel / Kategorisiz"
                base_q = base_q.filter((Suggestion.category.is_(None)) | (Suggestion.category == ""))
            else:
                base_q = base_q.filter(Suggestion.category == cat)


        # Arama filtresi
        if q:
            like = f"%{q}%"
            base_q = base_q.filter(or_(
                Suggestion.text.ilike(like),
                Suggestion.category.ilike(like),
                Suggestion.risk_code.ilike(like)
            ))

        base_q = base_q.order_by(Suggestion.category.asc(), Suggestion.id.desc())

        # Sayfalama
        pagination = base_q.paginate(page=page, per_page=per_page, error_out=False)
        items = pagination.items
        total = pagination.total
        pages = pagination.pages or 1

        # -----------------------------
        # 3) Görünüm için gruplama
        # -----------------------------
        def _disp_name(name: str) -> str:
            name = (name or "").strip()
            return name if name else "Genel / Kategorisiz"

        categories = {}

        # Bu sayfadaki kayıtları kategorilere dağıt
        for s in items:
            key = _disp_name(s.category)
            categories.setdefault(key, []).append(s)

        # Boş kategori kartları da gözüksün
        for rc in rcats:
            categories.setdefault(_disp_name(rc.name), [])

        if cat and cat != "__all__":
            categories.setdefault(_disp_name(cat), [])

        # Alfabetik sırala
        categories = dict(sorted(categories.items(), key=lambda kv: kv[0].lower()))

        # -----------------------------
        # Yardımcı: seçili id'leri topla
        # -----------------------------
        def _collect_selected_ids():
            ids = request.form.getlist("selected")
            if not ids:
                raw = request.form.get("selected_json", "[]")
                try:
                    data = json.loads(raw)
                    ids = [int(x) for x in data if str(x).isdigit()]
                except Exception:
                    ids = []
            return [int(sid) for sid in ids if str(sid).isdigit()]

        # -----------------------------
        # 4) POST: Seçilenlerden işlem
        # -----------------------------
        if request.method == "POST":
            action = (request.form.get("action") or "").strip()

            # A) Seçilen şablonlardan riskleri DOĞRUDAN oluştur
            if action == "add_selected":
                selected_ids = _collect_selected_ids()
                if not selected_ids:
                    flash("Lütfen en az bir risk seçin.", "danger")
                    return render_template(
                        "risk_identify.html",
                        categories=categories,
                        q=q, cat=cat, page=page, pages=pages, total=total,
                        per_page=per_page, filter_cat_names=filter_cat_names
                    )

                owner = session.get("username")
                pid = _get_active_project_id()
                cnt = 0

                for sid in selected_ids:
                    s = Suggestion.query.get(int(sid))
                    if not s:
                        continue
                    r = Risk(
                        title=s.text[:150],
                        category=s.category,
                        description=s.text,
                        owner=owner,
                        project_id=pid
                    )
                    db.session.add(r)
                    db.session.flush()
                    db.session.add(Comment(
                        risk_id=r.id,
                        text=f"Tanımlı risk seçildi: {datetime.utcnow().isoformat(timespec='seconds')} UTC",
                        is_system=True
                    ))
                    cnt += 1

                db.session.commit()
                flash(f"{cnt} risk eklendi.", "success")
                return redirect(url_for("dashboard"))

            # B) Seçilen şablonları risk_new formunda aç (from_suggestions ile)
            if action == "pick_for_new":
                selected_ids = _collect_selected_ids()
                if not selected_ids:
                    flash("Lütfen en az bir şablon seçin.", "danger")
                    return render_template(
                        "risk_identify.html",
                        categories=categories,
                        q=q, cat=cat, page=page, pages=pages, total=total,
                        per_page=per_page, filter_cat_names=filter_cat_names
                    )

                # Örn: [12, 14, 27] -> "12,14,27"
                id_str = ",".join(str(i) for i in selected_ids)

                flash(
                    f"{len(selected_ids)} şablon seçildi. Yeni risk formunda düzenleyip oluşturabilirsiniz.",
                    "success"
                )

                # /risk/new?from_suggestions=12,14,27
                return redirect(url_for("risk_new", from_suggestions=id_str))

            # ❌ Bilinmeyen/boş action: “geçersiz işlem” demeden GET görünümüne dön
            return redirect(url_for("risk_identify", q=q, cat=cat, page=page))

        # -----------------------------
        # GET: Sayfa render
        # -----------------------------
        return render_template(
            "risk_identify.html",
            categories=categories,
            q=q, cat=cat, page=page, pages=pages, total=total,
            per_page=per_page, filter_cat_names=filter_cat_names
        )


    # -------------------------------------------------
    #  Şablon (Suggestion) düzenleme / silme (ADMIN)
    # -------------------------------------------------
    @app.post("/admin/suggestions/<int:sid>/update")
    @role_required("admin")
    def admin_suggestion_update(sid):
        s = Suggestion.query.get_or_404(sid)

        new_text     = (request.form.get("text") or "").strip()
        new_category = (request.form.get("category") or s.category or "").strip()

        # ham risk kodu
        risk_code_raw = (request.form.get("risk_code") or "").strip()

        # edit formundaki alanlar
        new_risk_desc       = (request.form.get("risk_desc") or "").strip()
        new_mitigation_hint = (request.form.get("mitigation_hint") or "").strip()

        # ------- Kelime limitleri (create ile aynı) -------
        def _word_count(s_: str) -> int:
            return len(s_.split()) if s_ else 0

        MAX_MAIN_WORDS = 120      # Risk metni
        MAX_DESC_WORDS = 200      # Risk faktörü (açıklama)
        MAX_HINT_WORDS = 200      # Önerilen önlemler

        # Sadece dolu gelen alanları kontrol ediyoruz
        if new_text and _word_count(new_text) > MAX_MAIN_WORDS:
            flash(f"Risk metni çok uzun (en fazla {MAX_MAIN_WORDS} kelime).", "danger")
            return redirect(url_for("risk_identify"))

        if new_risk_desc and _word_count(new_risk_desc) > MAX_DESC_WORDS:
            flash(f"Risk faktörü açıklaması çok uzun (en fazla {MAX_DESC_WORDS} kelime).", "danger")
            return redirect(url_for("risk_identify"))

        if new_mitigation_hint and _word_count(new_mitigation_hint) > MAX_HINT_WORDS:
            flash(f"Önerilen önlemler çok uzun (en fazla {MAX_HINT_WORDS} kelime).", "danger")
            return redirect(url_for("risk_identify"))

        # ------- Risk kodu formatı: ABC12 -------
        new_code = risk_code_raw.upper() or None
        if new_code:
            # ilk 3 harf, son 2 rakam
            if not re.match(r"^[A-Z]{3}[0-9]{2}$", new_code):
                flash("Risk kodu formatı hatalı. Örnek: UYR02 (ilk 3 karakter harf, son 2 karakter rakam).", "danger")
                return redirect(url_for("risk_identify"))

        def _toi(x):
            if x in (None, ""):
                return None
            try:
                v = int(x)
                return max(1, min(5, v))
            except Exception:
                return None

        new_p = _toi(request.form.get("default_prob"))
        new_s = _toi(request.form.get("default_sev"))

        # ---- Asıl alanları güncelle ----
        if new_text:
            s.text = new_text

        if new_risk_desc:
            s.risk_desc = new_risk_desc
        elif not s.risk_desc:
            # tamamen boşsa en azından text’e yaslan
            s.risk_desc = s.text

        # boş stringe set etme, None yap
        s.mitigation_hint = new_mitigation_hint or None

        s.category = new_category or s.category
        s.risk_code = new_code  # format kontrolünden geçtiyse buraya geldi

        if new_p is not None:
            s.default_prob = new_p
        if new_s is not None:
            s.default_sev = new_s

        db.session.commit()
        flash("Şablon güncellendi.", "success")
        return redirect(url_for("risk_identify"))



    @app.post("/admin/suggestions/<int:sid>/delete")
    @role_required("admin")
    def admin_suggestion_delete(sid):
        s = Suggestion.query.get_or_404(sid)
        db.session.delete(s)
        db.session.commit()
        flash("Şablon silindi.", "success")
        return redirect(url_for("risk_identify"))
    
    @app.post("/admin/suggestions/create")
    @role_required("admin")
    def admin_suggestion_create():
        text     = (request.form.get("text") or "").strip()
        category = (request.form.get("category") or "").strip()

        # risk_code ham hali
        risk_code_raw = (request.form.get("risk_code") or "").strip()

        # Yeni alanlar: formdaki textarea isimleriyle birebir aynı
        risk_desc       = (request.form.get("risk_desc") or "").strip()
        mitigation_hint = (request.form.get("mitigation_hint") or "").strip()

        # ------- Kelime limitleri -------
        def _word_count(s: str) -> int:
            return len(s.split()) if s else 0

        MAX_MAIN_WORDS = 120      # Risk metni
        MAX_DESC_WORDS = 200      # Risk faktörü (açıklama)
        MAX_HINT_WORDS = 200      # Önerilen önlemler

        if _word_count(text) > MAX_MAIN_WORDS:
            flash(f"Risk metni çok uzun (en fazla {MAX_MAIN_WORDS} kelime).", "danger")
            return redirect(url_for("risk_identify"))

        if _word_count(risk_desc) > MAX_DESC_WORDS:
            flash(f"Risk faktörü açıklaması çok uzun (en fazla {MAX_DESC_WORDS} kelime).", "danger")
            return redirect(url_for("risk_identify"))

        if _word_count(mitigation_hint) > MAX_HINT_WORDS:
            flash(f"Önerilen önlemler çok uzun (en fazla {MAX_HINT_WORDS} kelime).", "danger")
            return redirect(url_for("risk_identify"))

        # ------- Risk kodu formatı: ABC12 -------
        risk_code = risk_code_raw.upper() or None
        if risk_code:
            # ilk 3 harf, son 2 rakam
            if not re.match(r"^[A-Z]{3}[0-9]{2}$", risk_code):
                flash("Risk kodu formatı hatalı. Örnek: UYR02 (ilk 3 karakter harf, son 2 karakter rakam).", "danger")
                return redirect(url_for("risk_identify"))

        def _toi(x):
            if x in (None, ""):
                return None
            try:
                v = int(x)
                return max(1, min(5, v))
            except Exception:
                return None

        default_prob = _toi(request.form.get("default_prob"))
        default_sev  = _toi(request.form.get("default_sev"))

        if not text:
            flash("Risk metni zorunludur.", "danger")
            return redirect(url_for("risk_identify"))

        if not category:
            category = "Genel"

        # Kategori tablosunda yoksa otomatik oluştur
        from sqlalchemy import func as _func
        rc = (
            RiskCategory.query
            .filter(_func.lower(RiskCategory.name) == _func.lower(category))
            .first()
        )
        if not rc:
            db.session.add(RiskCategory(name=category, is_active=True))

        # risk_desc boşsa text’i kopyalıyoruz, mitigation_hint boşsa None
        s = Suggestion(
            text=text,
            category=category,
            risk_code=risk_code,
            default_prob=default_prob,
            default_sev=default_sev,
            risk_desc=risk_desc or text,
            mitigation_hint=mitigation_hint or None,
        )

        db.session.add(s)
        db.session.commit()
        flash("Yeni şablon eklendi.", "success")
        return redirect(url_for("risk_identify") + f"#cat-{category.replace(' ', '-')}")




    # -------------------------------------------------
    #  Yeni Risk  (Kategori dropdown RiskCategory’den)
    # -------------------------------------------------
    
    @app.route("/risks/new", methods=["GET", "POST"])
    def risk_new():
        """
        Yeni riskler identify ekranında seçilen 'sepet' üzerinden oluşturulur.
        İki mod:
        - merge=1  -> tüm şablonlardan TEK risk oluştur (rapor mantığı)
        - merge=0  -> her şablondan ayrı risk (mevcut davranış)
        """

        # -----------------------------------------
        # 0) from_suggestions query paramı (yeni akış)
        #    /risks/new?from_suggestions=12,14,27 gibi
        #    varsa BUNU kullan; yoksa eski session sepetini kullan
        # -----------------------------------------
        from_str = (request.args.get("from_suggestions") or "").strip()

        picked_ids = []
        if from_str:
            try:
                picked_ids = [
                    int(part.strip())
                    for part in from_str.split(",")
                    if part.strip().isdigit()
                ]
            except Exception:
                picked_ids = []

            # Sepeti session'a da yaz (geri dönünce vs. işe yarar)
            session["picked_rows"] = picked_ids
        else:
            picked_ids = session.get("picked_rows") or []

        picked_suggestions = []
        if picked_ids:
            picked_suggestions = (
                Suggestion.query
                .filter(Suggestion.id.in_(picked_ids))
                .order_by(Suggestion.category.asc(), Suggestion.id.desc())
                .all()
            )

        # -----------------------------------------
        # GET + POST için PREFILL alanları hazırla
        # (Risk Tanımı / Risk Azaltıcı Önlemler Excel kolonlarından)
        # -----------------------------------------
        title_prefill = ""
        description_prefill = ""
        mitigation_prefill = ""

        if picked_suggestions:
            # Tek şablon seçiliyse: direkt o satırdan doldur
            if len(picked_suggestions) == 1:
                s0 = picked_suggestions[0]
                # Başlık: text'in ilk 150 karakteri
                title_prefill = (s0.text or "")[:150]

                # Açıklama: Risk Tanımı varsa onu, yoksa text
                description_prefill = (s0.risk_desc or s0.text or "") or ""

                # Önlemler: Risk Azaltıcı Önlemler
                mitigation_prefill = s0.mitigation_hint or ""
            else:
                # Birden fazla şablon: bullet list yapalım
                title_prefill = (picked_suggestions[0].text or "")[:150]

                desc_lines = []
                mit_lines = []
                for s in picked_suggestions:
                    code = (s.risk_code or "").strip()
                    label = f"[{code}] " if code else ""
                    base_text = (s.risk_desc or s.text or "").strip()
                    if base_text:
                        desc_lines.append(f"- {label}{base_text}")
                    if (s.mitigation_hint or "").strip():
                        mit_lines.append(f"- {label}{s.mitigation_hint.strip()}")

                description_prefill = "\n".join(desc_lines)
                mitigation_prefill = "\n".join(mit_lines)

        # -----------------------------------------
        # POST: Sepetten risk(ler) oluşturma
        # -----------------------------------------
        if request.method == "POST":
            action = (request.form.get("action") or "").strip()
            if action == "create_from_picked":
                # 1) Sepet ID’leri
                raw = (request.form.get("picked_ids") or "").strip()
                if raw:
                    try:
                        sel_ids = [int(x) for x in raw.split(",") if str(x).strip().isdigit()]
                    except Exception:
                        sel_ids = []
                else:
                    sel_ids = list(picked_ids)

                if not sel_ids:
                    flash("Şablon seçimi boş görünüyor.", "warning")
                    return render_template(
                        "risk_new.html",
                        picked_suggestions=picked_suggestions,
                        title_prefill=title_prefill,
                        description_prefill=description_prefill,
                        mitigation_prefill=mitigation_prefill,
                    )

                # 2) Ortak alanlar
                title_common       = (request.form.get("title") or "").strip() or None
                description_common = (request.form.get("description") or "").strip() or None
                mitigation_common  = (request.form.get("mitigation") or "").strip() or None

                # 🔴 mitigation_common boşsa, prefill'den al:
                mitigation_effective = mitigation_common or mitigation_prefill or None

                responsible = (request.form.get("responsible") or "").strip() or None
                duration    = (request.form.get("duration") or "").strip() or None

                # YYYY-MM (JS doldurur)
                start_month = (request.form.get("start_month") or "").strip() or None
                end_month   = (request.form.get("end_month")   or "").strip() or None

                # Tek kayıtta birleştir?
                merge_mode = (request.form.get("merge") == "1")

                owner = session.get("username")
                pid   = _get_active_project_id()

                def _toi(v):
                    try:
                        vv = int(v)
                        return max(1, min(5, vv))
                    except Exception:
                        return None

                # ==== A) TEK KAYIT (merge) ====
                if merge_mode:
                    sug_rows = (
                        Suggestion.query
                        .filter(Suggestion.id.in_(sel_ids))
                        .order_by(Suggestion.category.asc(), Suggestion.id.desc())
                        .all()
                    )
                    if not sug_rows:
                        flash("Şablonlar yüklenemedi.", "danger")
                        return render_template(
                            "risk_new.html",
                            picked_suggestions=picked_suggestions,
                            title_prefill=title_prefill,
                            description_prefill=description_prefill,
                            mitigation_prefill=mitigation_prefill,
                        )

                    # Kategori: ilk dolu kategori (yoksa Genel)
                    cat = None
                    for s in sug_rows:
                        if (s.category or "").strip():
                            cat = s.category.strip()
                            break
                    cat = cat or "Genel"

                    # Açıklama: kullanıcı açıklaması + şablon listesi
                    bullets = []
                    for s in sug_rows:
                        code = (s.risk_code or "").strip()
                        bullets.append(f"- {s.text}" + (f"  ({code})" if code else ""))
                    bullets_text = "\n".join(bullets)

                    final_desc = (description_common or "").strip()
                    if final_desc:
                        final_desc += "\n\n**Birleştirilen Şablonlar:**\n" + bullets_text
                    else:
                        final_desc = "**Birleştirilen Şablonlar:**\n" + bullets_text

                    # ---------- P/S + RPN ORTALAMASI ve tek tek kayıtlar ----------
                    p_vals, s_vals, rpn_vals = [], [], []
                    eval_items = []  # (p0, s0, s_row)

                    for s in sug_rows:
                        p0 = _toi(getattr(s, "default_prob", None))
                        s0 = _toi(getattr(s, "default_sev", None))

                        if p0 is not None:
                            p_vals.append(p0)
                        if s0 is not None:
                            s_vals.append(s0)

                        if p0 and s0:
                            rpn_vals.append(p0 * s0)
                            eval_items.append((p0, s0, s))

                    p_init = s_init = avg_rpn = None
                    if p_vals:
                        p_init = round(sum(p_vals) / len(p_vals))
                    if s_vals:
                        s_init = round(sum(s_vals) / len(s_vals))
                    if rpn_vals:
                        avg_rpn = sum(rpn_vals) / len(rpn_vals)

                    # Risk kaydını oluştur
                    r = Risk(
                        title=(title_common or (sug_rows[0].text or "")[:150]),
                        category=cat,
                        description=final_desc,
                        mitigation=mitigation_effective,   # 🔴 burası
                        responsible=responsible,
                        duration=duration,
                        start_month=start_month,
                        end_month=end_month,
                        owner=owner,
                        project_id=pid,
                    )
                    db.session.add(r)
                    db.session.flush()

                    # 1) Her şablon için ayrı Evaluation satırı
                    for p0, s0, s in eval_items:
                        code = (s.risk_code or "").strip()
                        label = code or f"#{s.id}"
                        db.session.add(Evaluation(
                            risk_id=r.id,
                            evaluator=owner or "System",
                            probability=int(p0),
                            severity=int(s0),
                            detection=None,
                            comment=f"{label} şablon varsayılan P/S (P={p0}, S={s0}, RPN={p0 * s0})"
                        ))

                    # 2) En son: ORTALAMA Evaluation satırı
                    if p_init is not None and s_init is not None:
                        if avg_rpn is not None:
                            comment = f"Birleştirilmiş şablonların ortalaması (RPN ort: {avg_rpn:.2f})"
                        else:
                            comment = "Birleştirilmiş şablonların ortalaması"

                        db.session.add(Evaluation(
                            risk_id=r.id,
                            evaluator=owner or "System",
                            probability=int(p_init),
                            severity=int(s_init),
                            detection=None,
                            comment=comment
                        ))

                    db.session.add(Comment(
                        risk_id=r.id,
                        text=(
                            "Toplu oluşturma (tek kayıt): "
                            + ", ".join([f"#{s.id}" for s in sug_rows])
                            + f" — {datetime.utcnow().isoformat(timespec='seconds')} UTC"
                        ),
                        is_system=True
                    ))
                    db.session.commit()
                    session.pop("picked_rows", None)
                    flash("Seçilen şablonlardan **tek bir risk** oluşturuldu.", "success")
                    return redirect(url_for("risk_detail", risk_id=r.id))

                # ==== B) AYRI AYRI ====
                created_ids = []
                for sid in sel_ids:
                    s = Suggestion.query.get(int(sid))
                    if not s:
                        continue
                    r = Risk(
                        title=(title_common or (s.text or "")[:150]),
                        category=(s.category or None),
                        description=(description_common or (s.text or None)),
                        mitigation=mitigation_effective,   # 🔴 burada da aynı metni kullan
                        responsible=responsible,
                        duration=duration,
                        start_month=start_month,
                        end_month=end_month,
                        owner=owner,
                        project_id=pid,
                    )

                    db.session.add(r)
                    db.session.flush()   # id hemen gelsin
                    created_ids.append(r.id)

                    p0 = _toi(getattr(s, "default_prob", None))
                    s0 = _toi(getattr(s, "default_sev", None))
                    if p0 and s0:
                        db.session.add(Evaluation(
                            risk_id=r.id,
                            evaluator=owner or "System",
                            probability=p0,
                            severity=s0,
                            detection=None,
                            comment="Şablon varsayılan değerlerinden"
                        ))

                    db.session.add(Comment(
                        risk_id=r.id,
                        text=f"Tanımlı şablondan oluşturuldu: {datetime.utcnow().isoformat(timespec='seconds')} UTC",
                        is_system=True
                    ))

                db.session.commit()
                session.pop("picked_rows", None)

                created = len(created_ids)
                flash(f"{created} risk oluşturuldu.", "success")

                # 🔴 BURADAN İTİBAREN: DASHBOARD YERİNE BULK DETAY
                if not created_ids:
                    return redirect(url_for("risk_new"))

                main_id = created_ids[0]

                if len(created_ids) == 1:
                    # Tek risk => normal detay sayfası
                    return redirect(url_for("risk_detail", risk_id=main_id))

                # Birden fazla risk => bulk parametresiyle P/S paneli
                bulk_param = ",".join(str(x) for x in created_ids)
                return redirect(url_for("risk_detail", risk_id=main_id, bulk=bulk_param))

        # -----------------------------------------
        # GET: Formu render et
        # -----------------------------------------
        return render_template(
            "risk_new.html",
            picked_suggestions=picked_suggestions,
            title_prefill=title_prefill,
            description_prefill=description_prefill,
            mitigation_prefill=mitigation_prefill,
        )










       # -------------------------------------------------
    #  Risk Listesi / Arama
    # -------------------------------------------------
   # -------------------------------------------------
#  Risk Listesi / Arama
# -------------------------------------------------
    @app.route("/risks")
    def risk_select():
        pid = _get_active_project_id()
        q = request.args.get("q", "").strip()

        # Matristen gelen hücre filtresi
        p = request.args.get("p", type=int)
        s = request.args.get("s", type=int)

        query = Risk.query
        if pid:
            query = query.filter(Risk.project_id == pid)

        if q:
            like = f"%{q}%"
            query = query.filter(
                (Risk.title.ilike(like)) |
                (Risk.category.ilike(like)) |
                (Risk.description.ilike(like))
            )

        # Hücreye tıklama filtresi: SON değerlendirmedeki P/S
        if p and s:
            latest_eval_sub = (
                db.session.query(
                    Evaluation.risk_id,
                    func.max(Evaluation.id).label("max_id")
                )
                .group_by(Evaluation.risk_id)
                .subquery()
            )

            query = (
                query
                .join(latest_eval_sub, latest_eval_sub.c.risk_id == Risk.id)
                .join(Evaluation, Evaluation.id == latest_eval_sub.c.max_id)
                .filter(
                    Evaluation.probability == p,
                    Evaluation.severity == s,
                )
            )

        risks = query.order_by(Risk.updated_at.desc()).all()

        # ==========================================================
        # ✅ YENİ: Risk başına TRY maliyet toplamı
        # ==========================================================
        risk_ids = [r.id for r in risks]
        cost_map = {}

        if risk_ids:
            rows = (
                db.session.query(
                    CostItem.risk_id,
                    func.coalesce(func.sum(CostItem.total), 0).label("total_try")
                )
                .filter(CostItem.risk_id.in_(risk_ids))
                .filter(CostItem.currency == "TRY")          # listede TRY gösteriyoruz
                .filter(CostItem.risk_id.isnot(None))
            )

            if pid:
                rows = rows.filter(CostItem.project_id == pid)

            rows = rows.group_by(CostItem.risk_id).all()

            # Decimal -> float (template'te rahat formatlamak için)
            cost_map = {rid: float(total) for (rid, total) in rows}

        return render_template("risk_select.html", risks=risks, q=q, cost_map=cost_map)

    # -------------------------------------------------
    #  Risk Sil (Admin)
    # -------------------------------------------------
    @app.route("/risks/<int:risk_id>/delete", methods=["POST"])
    @role_required("admin")
    def risk_delete(risk_id):
        r = Risk.query.get_or_404(risk_id)
        title = r.title
        db.session.delete(r)
        db.session.commit()
        flash(f"Risk silindi: {title}", "success")
        return redirect(url_for("risk_select"))


    @app.route("/risks/delete_all", methods=["POST"])
    @role_required("admin")
    def risks_delete_all():
        """
        Tüm riskleri toplu siler.
        Sadece admin rolü kullanabilir.
        """
        risks = Risk.query.all()
        deleted = len(risks)

        if deleted == 0:
            flash("Silinecek risk bulunamadı.", "info")
            return redirect(url_for("risk_select"))

        for r in risks:
            db.session.delete(r)

        db.session.commit()
        flash(f"Tüm riskler silindi. ({deleted} kayıt)", "success")
        return redirect(url_for("risk_select"))


    # -------------------------------------------------
    #  Risk Detay + Konsensüs + Öneri
    # -------------------------------------------------
    # -------------------------------------------------
#  Risk Detay + Konsensüs + Öneri
# -------------------------------------------------
    @app.route("/risks/<int:risk_id>", methods=["GET", "POST"])
    def risk_detail(risk_id):
        # ✅ Aktif proje zorunlu
        project_id = _active_project_id()
        if not project_id:
            flash("Aktif proje yok. Önce proje seç.", "warning")
            return redirect(url_for("dashboard"))

        # ✅ Risk mutlaka bu projeye ait olmalı
        r = Risk.query.filter_by(id=risk_id, project_id=project_id).first()
        if not r:
            flash("Risk bulunamadı (ya da bu projeye ait değil).", "warning")
            return redirect(url_for("index"))

        # Formda göstermek için aktif kategori adları (liste)
        cats = [
            x.name for x in RiskCategory.query
            .filter(RiskCategory.is_active.is_(True))
            .order_by(RiskCategory.name.asc())
            .all()
        ]

        # ========= TOPLU DEĞERLENDİRME (bulk) =========
        # /risks/3?bulk=3,4,5 gibi bir URL'den geliyorsa
        bulk_risks = None
        bulk_raw = (request.args.get("bulk", "") or "").strip()
        if bulk_raw:
            try:
                id_list = sorted({int(x) for x in bulk_raw.split(",") if x.strip()})
            except ValueError:
                id_list = []

            if id_list:
                # ✅ bulk riskler de aynı projeden olmalı
                bulk_risks = (
                    Risk.query
                    .filter(Risk.project_id == project_id, Risk.id.in_(id_list))
                    .order_by(Risk.id.asc())
                    .all()
                )

        # ========= POST: RİSK FORMUNU KAYDET =========
        if request.method == "POST":
            # ----- Diğer alanlar -----
            r.title        = request.form.get("title", r.title)
            r.description  = request.form.get("description", r.description)
            r.status       = request.form.get("status", r.status)
            r.risk_type    = (request.form.get("risk_type") or None)
            r.responsible  = (request.form.get("responsible") or None)
            r.mitigation   = (request.form.get("mitigation") or None)
            r.duration     = (request.form.get("duration") or None)
            r.start_month  = (request.form.get("start_month") or None)  # YYYY-MM (hidden)
            r.end_month    = (request.form.get("end_month") or None)    # YYYY-MM (hidden)

            # ===== KATEGORİLER (ÇOKLU) =====
            selected = request.form.getlist("categories")  # <select multiple name="categories">

            # Özel kategori alanı: "A, B, C" gibi virgüllü
            custom_raw = request.form.get("category_custom", "")
            custom = [x.strip() for x in (custom_raw or "").split(",") if x.strip()]

            # Listede "__custom__" sentineli seçilmişse onu at; custom listesini ekle
            cats_final = [c for c in selected if c != "__custom__"] + custom

            # Risk objesine set et (ilkini geri uyumluluk için r.category'ye de yazar)
            r.set_categories(cats_final)

            # ✅ Mitigation satırlarını senkronize et
            # (_sync_mitigations fonksiyonun create_app içinde tanımlı olmalı)
            _sync_mitigations(r)

            # Sistem notu (aynı transaction içinde)
            db.session.add(Comment(
                risk_id=r.id,
                text=f"Risk düzenlendi: {datetime.utcnow().isoformat(timespec='seconds')} UTC",
                is_system=True
            ))

            db.session.commit()

            flash("Değişiklikler kaydedildi.", "success")
            return redirect(url_for("risk_detail", risk_id=r.id))

        # ========= GET: ÖNERİLER =========
        cats_sel = r.categories_list or ([r.category] if r.category else [])
        if cats_sel:
            sugg = Suggestion.query.filter(Suggestion.category.in_(cats_sel)).all()
        else:
            sugg = []

        # ========= Konsensüs =========
        threshold = int(current_app.config.get("CONSENSUS_THRESHOLD", 30))
        pair_counts = {}
        for e in (r.evaluations or []):
            pair = (e.probability, e.severity)
            pair_counts[pair] = pair_counts.get(pair, 0) + 1

        consensus = None
        if pair_counts:
            (p_val, s_val), cnt = max(pair_counts.items(), key=lambda kv: kv[1])
            if cnt >= threshold:
                consensus = {"p": p_val, "s": s_val, "count": cnt}

        # ========= Geçmiş değerlendirmeler / ortalama =========
        eval_history = sorted(
            list(r.evaluations or []),
            key=lambda ev: ev.created_at
        )

        avg_p = avg_s = None
        last_p = last_s = None
        use_avg = False

        if eval_history:
            last = eval_history[-1]
            last_p = last.probability
            last_s = last.severity

            if len(eval_history) >= 2:
                probs = [ev.probability for ev in eval_history if ev.probability is not None]
                sevs  = [ev.severity for ev in eval_history if ev.severity is not None]
                if probs:
                    avg_p = sum(probs) / len(probs)
                if sevs:
                    avg_s = sum(sevs) / len(sevs)
                if avg_p is not None or avg_s is not None:
                    use_avg = True

        # ========= Sistemin önerdiği P/S =========
        ps_reco = None
        if cats_sel:
            rows = (
                db.session.query(Evaluation.probability, Evaluation.severity)
                .join(Risk, Risk.id == Evaluation.risk_id)
                .outerjoin(RiskCategoryRef, RiskCategoryRef.risk_id == Risk.id)
                .filter(
                    Risk.project_id == project_id,  # ✅ projeye kilitle
                    or_(
                        RiskCategoryRef.name.in_(cats_sel),
                        Risk.category.in_(cats_sel)
                    )
                )
                .all()
            )
            probs = [p for (p, s) in rows if p is not None]
            sevs  = [s for (p, s) in rows if s is not None]
            if probs or sevs:
                p_mode = Counter(probs).most_common(1)
                s_mode = Counter(sevs).most_common(1)
                ps_reco = {
                    "p": p_mode[0][0] if p_mode else None,
                    "s": s_mode[0][0] if s_mode else None
                }

        # ========= ✅ Bu riske bağlı maliyetler =========
        risk_costs = (
            CostItem.query
            .filter_by(project_id=project_id, risk_id=r.id)
            .order_by(CostItem.id.desc())
            .all()
        )

        # (opsiyonel) para birimine göre toplam
        cost_totals = {}
        for c in (risk_costs or []):
            cur = (c.currency or "TRY").upper()
            val = c.total if c.total is not None else Decimal("0")

            # güvenli Decimal dönüşümü
            try:
                val_dec = val if isinstance(val, Decimal) else Decimal(str(val))
            except Exception:
                val_dec = Decimal("0")

            prev = cost_totals.get(cur, Decimal("0"))
            cost_totals[cur] = prev + val_dec

        return render_template(
            "risk_detail.html",
            r=r,
            suggestions=sugg,
            consensus=consensus,
            threshold=threshold,
            ps_reco=ps_reco,
            categories=cats,
            eval_history=eval_history,
            avg_p=avg_p,
            avg_s=avg_s,
            last_p=last_p,
            last_s=last_s,
            use_avg=use_avg,
            bulk_risks=bulk_risks,

            # ✅ maliyet blokları (risk detail’de göstermek için)
            risk_costs=risk_costs,
            cost_totals=cost_totals,
        )


    # -------------------------------------------------
    #  Yorum / Değerlendirme
    # -------------------------------------------------
    @app.route("/risk/<int:risk_id>/comment", methods=["POST"])
    def add_comment(risk_id):
        r = Risk.query.get_or_404(risk_id)

        # admin alttaki formdan "normal yorum" girdiyse
        text = (request.form.get("text") or "").strip()

        # Zengin AI Yorum butonu, BOŞ text ile POST atıyor
        if not text:
            # burada senin gönderdiğin make_ai_risk_comment devreye giriyor
            text = make_ai_risk_comment(risk_id)
            is_system = True
        else:
            is_system = False

        c = Comment(
            risk_id=r.id,
            text=text,
            is_system=is_system,
        )
        db.session.add(c)
        db.session.commit()

        flash("Yorum eklendi.", "success")
        return redirect(url_for("risk_detail", risk_id=r.id))

    @app.route("/risks/<int:risk_id>/evaluation", methods=["POST"])
    def add_eval(risk_id):
        r = Risk.query.get_or_404(risk_id)
        evaluator = request.form.get("evaluator") or session.get("username")

        # P ve S zorunlu; 1..5’e sıkıştır
        p = int(request.form.get("probability", "3"))
        s = int(request.form.get("severity", "3"))
        p = min(max(p, 1), 5)
        s = min(max(s, 1), 5)

        c = request.form.get("comment", "")

        db.session.add(Evaluation(
            risk_id=r.id,
            evaluator=evaluator,
            probability=p,
            severity=s,
            detection=None,  # D kullanılmıyor
            comment=c
        ))
        r.status = "Assessed"
        db.session.commit()
        flash("Değerlendirme eklendi.", "success")
        return redirect(url_for("risk_detail", risk_id=r.id))

    @app.get("/health")
    def health():
        return {"ok": True}, 200



    # -------------------------------------------------
    #  Raporlar
    # -------------------------------------------------
    @app.route("/reports")
    def reports():
        pid = _get_active_project_id()

        query = Risk.query
        if pid:
            query = query.filter(Risk.project_id == pid)

        risks = query.order_by(Risk.updated_at.desc()).all()

        # ✅ reports listesinde göstermek için risklerin maliyet toplamları (risk_id + currency bazında)
        cost_map = {}
        if risks:
            risk_ids = [r.id for r in risks]

            rows = (
                db.session.query(
                    CostItem.risk_id,
                    func.coalesce(CostItem.currency, "TRY").label("currency"),
                    func.coalesce(func.sum(CostItem.total), 0).label("total"),
                )
                .filter(CostItem.risk_id.in_(risk_ids))
                .group_by(CostItem.risk_id, "currency")
                .all()
            )

            for rid, cur, total in rows:
                cost_map.setdefault(rid, []).append((cur, float(total or 0)))

            # (opsiyonel) her riskte para birimlerini sabit sıraya sokalım
            order = {"TRY": 0, "USD": 1, "EUR": 2}
            for rid in cost_map:
                cost_map[rid].sort(key=lambda x: order.get(x[0], 99))

        return render_template("reports.html", risks=risks, cost_map=cost_map)


    @app.route("/reports/<int:risk_id>")
    def report_view(risk_id):
        r = Risk.query.get_or_404(risk_id)

        # ✅ Bu risk’e bağlı maliyet kalemleri
        cost_items = (
            CostItem.query
            .filter(CostItem.risk_id == r.id)
            .order_by(CostItem.id.desc())
            .all()
        )

        # ✅ Para birimine göre toplamlar (TRY/USD/EUR ayrı ayrı)
        cost_totals = (
            db.session.query(
                CostItem.currency,
                func.coalesce(func.sum(CostItem.total), 0)
            )
            .filter(CostItem.risk_id == r.id)
            .group_by(CostItem.currency)
            .all()
        )

        # mevcut suggestions aynı kalsın
        suggestions = Suggestion.query.filter(Suggestion.category == (r.category or "")).all()

        return render_template(
            "report_view.html",
            r=r,
            suggestions=suggestions,
            cost_items=cost_items,
            cost_totals=cost_totals,
        )
        
# -------------------------------------------------
#  Ortak context: Zaman Çizelgesi verisi
# -------------------------------------------------
    def build_schedule_context():
        pid = _get_active_project_id()
        query = Risk.query
        if pid:
            query = query.filter(Risk.project_id == pid)

        # --- Basit filtreler (opsiyonel) ---
        q = (request.args.get("q") or "").strip()
        if q:
            like = f"%{q}%"
            query = query.filter(
                (Risk.title.ilike(like)) |
                (Risk.category.ilike(like)) |
                (Risk.description.ilike(like))
            )

        cat = (request.args.get("category") or "").strip()
        if cat:
            query = query.filter(Risk.category == cat)

        owner = (request.args.get("owner") or "").strip()
        if owner:
            query = query.filter(Risk.responsible == owner)

        status = (request.args.get("status") or "").strip()
        if status:
            query = query.filter(Risk.status == status)

        risks = (
            query
            .order_by(
                Risk.start_month.is_(None),
                Risk.start_month.asc(),
                Risk.updated_at.desc(),
                Risk.title.asc(),
            )
            .all()
        )

        # --- Yardımcı: yyyy-mm aralığını normalize et ---
        def _norm_range(sm, em):
            s = _parse_ym(sm)
            e = _parse_ym(em)
            if s and not e:
                e = s
            if e and not s:
                s = e
            if s and e and s > e:
                s, e = e, s
            return s, e

        # --- Ay penceresi (min..max) ---
        min_ym, max_ym = None, None
        for r in risks:
            s, e = _norm_range(r.start_month, r.end_month)
            if s and e:
                if (min_ym is None) or (s < min_ym):
                    min_ym = s
                if (max_ym is None) or (e > max_ym):
                    max_ym = e

        # Varsayılan: bugün + 5 ay (toplam 6 ay)
        if not min_ym or not max_ym:
            today = date.today()
            min_ym = (today.year, today.month)
            y, m = today.year, today.month
            for _ in range(5):
                y, m = _next_ym(y, m)
            max_ym = (y, m)

        # --- Sütun ayları ---
        months = []
        y, m = min_ym
        while True:
            months.append(_ym_to_str(y, m))
            if (y, m) == max_ym:
                break
            y, m = _next_ym(y, m)

        # --- Grade map (UI sınıfları için) ---
        _gmap = {
            "high": "critical",
            "medium": "moderate",
            "low": "low",
            "none": "acceptable",
        }

        # --- Satırlar ---
        rows = []
        for r in risks:
            s, e = _norm_range(r.start_month, r.end_month)
            active = set()
            if s and e:
                yy, mm = s
                while True:
                    active.add(_ym_to_str(yy, mm))
                    if (yy, mm) == e:
                        break
                    yy, mm = _next_ym(yy, mm)

            g = _gmap.get((r.grade() or "none").lower(), "acceptable")
            rows.append({
                "risk": r,
                "active": active,               # tabloda bar çizdirme
                "grade": g,                     # gx-... sınıfı için
                "startYM": r.start_month or "", # takvim (YYYY-MM)
                "endYM":   r.end_month or "",
            })

        # --- Filtre dropdown verileri ---
        categories = sorted({
            (r.category or "").strip()
            for r in risks
            if (r.category or "").strip()
        })
        owners = sorted({
            (r.responsible or "").strip()
            for r in risks
            if (r.responsible or "").strip()
        })
        statuses = sorted({
            (r.status or "").strip()
            for r in risks
            if (r.status or "").strip()
        })

        # Kullanıcının seçtiği ay/yıl (calendar + PDF için)
        today = date.today()
        cur_m = int(request.args.get("month") or today.month)
        cur_y = int(request.args.get("year") or today.year)

        return dict(
            months=months,
            rows=rows,
            categories=categories,
            owners=owners,
            statuses=statuses,
            current_month=cur_m,
            current_year=cur_y,
        )


    # -------------------------------------------------
    #  Zaman Çizelgesi — HTML
    # -------------------------------------------------
    @app.route("/schedule")
    def schedule():
        ctx = build_schedule_context()
        return render_template("schedule.html", **ctx)


    # -------------------------------------------------
    #  Zaman Çizelgesi — PDF
    # -------------------------------------------------
    @app.route("/schedule/pdf")
    def schedule_pdf():
        ctx = build_schedule_context()

        # Oluşturma zamanı (PDF footer’ında kullanacağız)
        ctx["generated_at"] = datetime.now().strftime("%d.%m.%Y %H:%M")

        # PDF için özel tasarım şablonu
        html = render_template("schedule_pdf.html", **ctx)

        pdf_bytes = HTML(
            string=html,
            base_url=request.host_url,
        ).write_pdf()

        buf = BytesIO(pdf_bytes)

        m = ctx.get("current_month") or date.today().month
        y = ctx.get("current_year") or date.today().year
        filename = f"risk_schedule_{y}_{str(m).zfill(2)}.pdf"

        return send_file(
            buf,
            as_attachment=True,
            download_name=filename,
            mimetype="application/pdf",
        )


    # -------------------------------------------------
    #  Kayıt — Tek Adım (Hesap + Proje)
    # -------------------------------------------------
    @app.route("/setup/1", methods=["GET", "POST"])
    def setup_step1():
        if request.method == "POST":
            # Form alanları
            lang = request.form.get("language") or "Türkçe"
            name = request.form.get("contact_name", "").strip()
            title = request.form.get("contact_title", "").strip()
            email = request.form.get("email", "").strip()
            password = request.form.get("password", "")
            workplace_name = request.form.get("workplace_name", "").strip()
            workplace_address = request.form.get("workplace_address", "").strip()
            project_duration = request.form.get("project_duration", "").strip()
            # NOT: ref_code artık kayıt ekranında alınmıyor

            # Zorunlu alan kontrolü
            if not all([name, email, password, workplace_name, workplace_address]):
                flash("Lütfen zorunlu alanları doldurun.", "danger")
                return render_template("setup_step1.html", form=request.form, hide_nav=True)

            # E-posta tekillik kontrolü
            if Account.query.filter_by(email=email).first():
                flash("Bu e-posta adresi zaten kayıtlı, lütfen giriş yapın.", "danger")
                return render_template("setup_step1.html", form=request.form, hide_nav=True)

            # İlk kullanıcı admin + active (bootstrap), diğerleri uzman + pending
            first_user = (Account.query.count() == 0)
            role   = "admin"  if first_user else "uzman"
            status = "active" if first_user else "pending"

            # Hesap oluştur
            acc = Account(
                language=lang,
                contact_name=name,
                contact_title=title,
                email=email,
                password_hash=generate_password_hash(password),
                role=role,
                status=status,
                # ref_code: yönetici atayana dek None
            )
            db.session.add(acc)
            db.session.flush()  # acc.id için

            # Proje oluştur
            proj = ProjectInfo(
                account_id=acc.id,
                workplace_name=workplace_name,
                workplace_address=workplace_address,
                project_duration=project_duration or None
            )
            db.session.add(proj)
            db.session.commit()

            if first_user:
                # İlk admin otomatik giriş
                flash("İlk admin hesabı oluşturuldu.", "success")
                session["account_id"] = acc.id
                session["username"]   = acc.contact_name
                session["role"]       = acc.role
                session["project_id"] = proj.id
                return redirect(url_for("dashboard"))
            else:
                # Başvuru alındı — admin onayı sonrası ref kodu mail edilecek
                send_email(
                    to_email=email,
                    subject="Kayıt alındı — admin onayı bekleniyor",
                    body=(
                        f"Merhaba {name},\n\n"
                        "Kayıt talebiniz alındı. Admin onayı sonrasında size Referans Kodunuz e-posta ile iletilecek. "
                        "Giriş için e-posta + şifre + referans kodu gereklidir.\n\n"
                        "Teşekkürler."
                    )
                )
                flash("Kayıt alındı. Admin onayı sonrası referans kodu e-posta ile gönderilecek.", "info")
                return redirect(url_for("login"))

        # GET
        return render_template("setup_step1.html", hide_nav=True)

    # -------------------------------------------------
    #  AYARLAR — Hesap ve Proje
    # -------------------------------------------------
    @app.route("/settings/account", methods=["GET", "POST"])
    def settings_account():
        acc = Account.query.get(session.get("account_id"))
        if not acc:
            return redirect(url_for("logout"))

        if request.method == "POST":
            acc.contact_name = request.form.get("contact_name", acc.contact_name).strip()
            acc.contact_title = request.form.get("contact_title", acc.contact_title).strip()
            acc.language = request.form.get("language", acc.language).strip()

            current_pw = request.form.get("current_password", "")
            new_pw = request.form.get("new_password", "")
            new_pw2 = request.form.get("new_password2", "")
            if new_pw or new_pw2:
                if not current_pw or not check_password_hash(acc.password_hash, current_pw):
                    flash("Mevcut şifre hatalı.", "danger")
                    return render_template("settings_account.html", acc=acc)
                if new_pw != new_pw2:
                    flash("Yeni şifreler eşleşmiyor.", "danger")
                    return render_template("settings_account.html", acc=acc)
                acc.password_hash = generate_password_hash(new_pw)

            db.session.commit()
            session["username"] = acc.contact_name
            flash("Hesap bilgileri güncellendi.", "success")
            return redirect(url_for("settings_account"))

        return render_template("settings_account.html", acc=acc)

    @app.route("/settings/project", methods=["GET", "POST"])
    def settings_project():
        acc_id = session.get("account_id")
        proj = ProjectInfo.query.filter_by(account_id=acc_id).order_by(ProjectInfo.created_at.desc()).first()
        if request.method == "POST":
            name = request.form.get("workplace_name", "").strip()
            addr = request.form.get("workplace_address", "").strip()
            duration = request.form.get("project_duration", "").strip()

            if not name or not addr:
                flash("İş yeri unvanı ve adres zorunludur.", "danger")
                return render_template("settings_project.html", proj=proj)

            if proj:
                proj.workplace_name = name
                proj.workplace_address = addr
                proj.project_duration = duration or None
            else:
                proj = ProjectInfo(
                    account_id=acc_id,
                    workplace_name=name,
                    workplace_address=addr,
                    project_duration=duration or None
                )
                db.session.add(proj)

            db.session.commit()
            flash("Proje bilgileri güncellendi.", "success")
            return redirect(url_for("settings_project"))

        return render_template("settings_project.html", proj=proj)

    # -------------------------------------------------
    #  SORUMLU ÖZETİ (liste)
    # -------------------------------------------------
    @app.route("/responsibles")
    def responsibles():
        pid = _get_active_project_id()
        query = Risk.query
        if pid:
            query = query.filter(Risk.project_id == pid)

        # Sorumlusu olan riskler
        risks = (
            query
            .filter(Risk.responsible.isnot(None))
            .filter(Risk.responsible != "")
            .all()
        )

        from collections import defaultdict
        buckets = defaultdict(lambda: {
            "responsible": "",
            "count": 0,
            "_sum_rpn": 0.0,
            "_n_rpn": 0,
            "critical": None,   # en yüksek skorlu risk
        })

        # Uygulamanın geri kalanıyla tutarlı skor hesabı
        def _risk_score(r):
            sc = None

            # 1) score() metodu varsa onu kullan
            s_method = getattr(r, "score", None)
            if callable(s_method):
                try:
                    sc = s_method()
                except Exception:
                    sc = None

            # 2) Yoksa avg_rpn()
            if sc is None:
                try:
                    sc = r.avg_rpn()
                except Exception:
                    sc = None

            try:
                return float(sc) if sc is not None else None
            except Exception:
                return None

        for r in risks:
            name = (r.responsible or "").strip()
            if not name:
                continue

            score = _risk_score(r)
            row = buckets[name]
            row["responsible"] = name
            row["count"] += 1

            if score is not None:
                row["_sum_rpn"] += score
                row["_n_rpn"] += 1

                # en kritik risk: en yüksek skorlu olan
                cur_crit = row["critical"]
                if cur_crit is None:
                    row["critical"] = r
                else:
                    cur_score = _risk_score(cur_crit)
                    if cur_score is None or score > cur_score:
                        row["critical"] = r

        # sözlükleri listeye çevir + ortalama RPN
        rows = []
        for data in buckets.values():
            total = data.pop("_sum_rpn")
            n     = data.pop("_n_rpn")
            data["avg_rpn"] = (total / n) if n else None
            rows.append(data)

        # Ortalama RPN'e göre azalan sırala
        rows.sort(
            key=lambda x: (
                x["avg_rpn"] is None,
                -(x["avg_rpn"] or 0.0),
                x["responsible"].lower(),
            )
        )

        return render_template("responsible_summary.html", rows=rows)

    # -------------------------------------------------
    #  SORUMLU ÖZETİ CSV
    # -------------------------------------------------
    @app.route("/responsibles/export.csv")
    def responsibles_export_csv():
        pid = _get_active_project_id()
        query = Risk.query
        if pid:
            query = query.filter(Risk.project_id == pid)
        risks = query.order_by(Risk.responsible.asc(), Risk.updated_at.desc()).all()

        buckets = {}
        for r in risks:
            name = (r.responsible or "").strip()
            if not name:
                continue
            buckets.setdefault(name, []).append(r)

        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(["Sorumlu", "Risk Sayısı", "Ortalama RPN", "En Kritik Risk", "En Kritik RPN"])

        for name, items in buckets.items():
            rpns = [float(ri.avg_rpn()) for ri in items if ri.avg_rpn() is not None]
            avg_rpn = (sum(rpns) / len(rpns)) if rpns else None

            critical = None
            best = -1.0
            for ri in items:
                val = ri.avg_rpn()
                if val is not None and float(val) > best:
                    best = float(val)
                    critical = ri

            writer.writerow([
                name,
                len(items),
                f"{avg_rpn:.2f}" if avg_rpn is not None else "",
                (critical.title if critical else ""),
                (f"{critical.avg_rpn():.2f}" if critical and critical.avg_rpn() is not None else "")
            ])

        resp = Response(output.getvalue(), mimetype="text/csv; charset=utf-8")
        resp.headers["Content-Disposition"] = "attachment; filename=responsible_summary.csv"
        return resp

    # -------------------------------------------------
    #  Kütüphane İçe Aktar (CSV/XLSX/XLS) — Sadece admin
    # -------------------------------------------------
    @app.route("/admin/import/suggestions", methods=["GET", "POST"])
    @role_required("admin")
    def import_suggestions():
        """
        CSV/XLSX içe aktarma:

        
        """
        if request.method == "POST":
            f = request.files.get("file")
            if not f or f.filename == "":
                flash("Bir CSV/XLSX/XLS dosyası seçin.", "danger")
                return render_template("import_suggestions.html")

            # 1) Dosyayı oku (header + satırlar)
            try:
                rows = _read_rows_from_upload(f)
            except RuntimeError as e:
                flash(str(e), "danger")
                return render_template("import_suggestions.html")
            except Exception as e:
                flash(f"Dosya okunamadı: {e}", "danger")
                return render_template("import_suggestions.html")

            if not rows:
                flash("Boş dosya.", "warning")
                return render_template("import_suggestions.html")

            # 2) Header analizi (normalize)
            raw_header = rows[0]
            _TRMAP = str.maketrans({
                "ç": "c", "ğ": "g", "ı": "i", "ö": "o", "ş": "s", "ü": "u",
                "Ç": "c", "Ğ": "g", "İ": "i", "Ö": "o", "Ş": "s", "Ü": "u"
            })

            def _norm(s: str) -> str:
                s = str(s or "").replace("\n", " ").replace("\r", " ").strip().translate(_TRMAP).lower()
                return " ".join(s.split())

            # Header boş ise uyar
            if not raw_header or all(str(c or "").strip() == "" for c in raw_header):
                flash("Başlık satırı boş görünüyor.", "danger")
                return render_template("import_suggestions.html")

            header = [_norm(c) for c in raw_header]

            def find_exact(keys):
                """Başlıkları birebir (normalize edilmiş) eşleştir."""
                keys = [k.strip().lower() for k in keys]
                for i, h in enumerate(header):
                    for k in keys:
                        if h == k:
                            return i
                return None

            # ZORUNLU kolonlar
            text_col = find_exact(["risk faktoru", "risk faktörü"])   # B sütunu: Risk Faktörü (kısa ad)
            # Kategori sütunu opsiyonel (yoksa tahmin edeceğiz)
            cat_col = find_exact(["kategoriler", "kategori"])

            # YENİ: Kısa risk adı ("Risk", "Risk Adı" vb.) — Excel'de ayrı kolonsa
            risk_title_col = find_exact([
                "risk",
                "risk adi",
                "risk adı",
                "riskler",
            ])

            # YENİ: Risk Tanımı / Risk Azaltıcı Önlemler kolonları
            risk_desc_col = find_exact(["risk tanimi", "risk tanımı"])  # C sütunu
            mitigation_col = find_exact([
                "risk azaltici onlemler",
                "risk azaltıcı önlemler",
                "risk azaltici onlem",
                "risk azaltıcı önlem",
            ])  # D sütunu

            # OPSİYONEL kolonlar (Kod, P, S)
            code_col = find_exact([
                "risk kodlari", "risk kodları",
                "risk kodu", "risk kod", "kod", "code",
            ])
            prob_col = find_exact([
                "ortalama risk olasiligi",
                "olasilik", "olasılık",
                "probability", "p (1-5)",
            ])
            sev_col = find_exact([
                "ortalama risk etkisi",
                "siddet", "şiddet",
                "etki", "severity", "s (1-5)",
            ])

            # Zorunlu başlık kontrolleri
            if text_col is None:
                flash("Başlık bulunamadı: 'Risk Faktörü' kolonu yok.", "danger")
                return render_template("import_suggestions.html")

            # Aynı kolona çarpma guard'ı
            if cat_col is not None and text_col == cat_col:
                flash("‘Risk Faktörü’ ve ‘Kategori’ aynı sütuna işaret ediyor. Dosya başlıklarını kontrol edin.", "danger")
                return render_template("import_suggestions.html")

            # Kategori bulunamadıysa: son sütunu kategori varsay (text ile çakışmasın)
            n_cols = len(header)
            if cat_col is None and n_cols > 1:
                candidate = n_cols - 1
                if candidate != text_col:
                    cat_col = candidate

            # -------------------------------------------------
            # P/Ş kolonları bulunamadıysa: kalan kolonlarda 1..5
            # aralığında yoğunluk arayıp otomatik tahmin et
            # -------------------------------------------------
            def _looks_like_score(col_idx):
                hits = 0
                for row in rows[1: min(len(rows), 25)]:
                    if col_idx >= len(row):
                        continue
                    try:
                        v = str(row[col_idx]).replace(",", ".").strip()
                        if v == "":
                            continue
                        f = float(v)
                        if 1.0 <= f <= 5.0:
                            hits += 1
                    except Exception:
                        pass
                return hits

            if prob_col is None or sev_col is None:
                candidates = []
                protected = {text_col}
                if cat_col is not None:
                    protected.add(cat_col)
                if code_col is not None:
                    protected.add(code_col)
                for i in range(len(header)):
                    if i in protected:
                        continue
                    candidates.append((_looks_like_score(i), i))
                candidates.sort(reverse=True)
                if prob_col is None and len(candidates) >= 1 and candidates[0][0] > 0:
                    prob_col = candidates[0][1]
                if sev_col is None and len(candidates) >= 2 and candidates[1][0] > 0:
                    sev_col = candidates[1][1]

            # 3) Yardımcılar
            def _clean(x):
                return str(x or "").strip()

            def _toi(x):
                try:
                    v = int(round(float(str(x).replace(",", ".").strip())))
                    return max(1, min(5, v))
                except Exception:
                    return None

            def _is_category_title(row):
                get = lambda idx: (row[idx] if idx is not None and idx < len(row) else "")
                code_val = _clean(get(code_col))
                text_val = _clean(get(text_col))
                prob_val = _clean(get(prob_col))
                sev_val  = _clean(get(sev_col))
                cat_val  = _clean(get(cat_col)) if cat_col is not None else ""
                only_text = (text_val != "" and code_val == "" and prob_val == "" and sev_val == "" and cat_val == "")
                looks_like = (
                    (text_val.isupper() and len(text_val.split()) <= 10)
                    or ("RİSKLER" in text_val.upper())
                    or text_val.endswith(":")
                )
                return only_text and looks_like

            PREFIX_TO_CATEGORY = {
                "YÖR": "YÖNETSEL RİSKLER",
                "SOR": "SÖZLEŞME / ONAY SÜREÇLERİ",
                "UYR": "UYGULAMA / YAPIM RİSKLERİ",
                "GER": "ZEMİN KOŞULLARI / GEOTEKNİK",
                "ÇER": "ÇEVRESEL RİSKLER",
                "CER": "ÇEVRESEL RİSKLER",
                "DTR": "DENETİM / TETKİK / RAPOR",
                "POR": "POLİTİK / ORGANİZASYONEL",
                "TYR": "TEDARİK / MALZEME",
            }

            def guess_category_from_code(code):
                if not code:
                    return None
                code = str(code).strip().upper()
                letters = "".join([c for c in code if c.isalpha()])
                return PREFIX_TO_CATEGORY.get(letters[:3])

            def _looks_like_sentence(x: str) -> bool:
                x = (x or "").strip()
                if not x:
                    return False
                words = x.split()
                return (len(words) >= 7) and (not x.isupper())

            created, skipped, updated_cnt = 0, 0, 0
            current_category = None

            # 4) Satırları işle
            for row in rows[1:]:
                if not row or all((_clean(c) == "") for c in row):
                    continue

                # Kategori başlığı satırı mı?
                if _is_category_title(row):
                    current_category = _clean(row[text_col]).rstrip(":")
                    if current_category:
                        rc = (
                            RiskCategory.query
                            .filter(func.lower(RiskCategory.name) == func.lower(current_category))
                            .first()
                        )
                        if not rc:
                            db.session.add(RiskCategory(name=current_category, is_active=True))
                    continue

                # Normal risk satırı
                r = list(row)
                idxs = [i for i in [code_col, text_col, cat_col, prob_col, sev_col, risk_desc_col, mitigation_col, risk_title_col] if i is not None]
                need_len = (max(idxs) if idxs else -1)
                while len(r) <= need_len:
                    r.append("")

                code     = _clean(r[code_col]) if code_col is not None else ""
                text     = _clean(r[text_col]) if text_col is not None else ""
                cat_cell = _clean(r[cat_col])  if cat_col  is not None else ""

                # YENİ: Kısa risk adı
                if risk_title_col is not None and risk_title_col < len(r):
                    risk_title_raw = _clean(r[risk_title_col])
                else:
                    risk_title_raw = ""
                risk_title = risk_title_raw or None

                # YENİ: Risk Tanımı + Azaltıcı Önlemler
                if risk_desc_col is not None and risk_desc_col < len(r):
                    risk_desc_raw = _clean(r[risk_desc_col])
                else:
                    risk_desc_raw = ""
                if mitigation_col is not None and mitigation_col < len(r):
                    mitigation_hint_raw = _clean(r[mitigation_col])
                else:
                    mitigation_hint_raw = ""

                risk_desc       = risk_desc_raw or None
                mitigation_hint = mitigation_hint_raw or None

                # text boşsa önce risk_title'dan, o da yoksa risk_desc'ten türet
                if not text:
                    if risk_title:
                        text = risk_title[:255]
                    elif risk_desc:
                        text = risk_desc[:255]

                # Hem text hem risk_desc hem risk_title boşsa satırı atla
                if not text and not risk_desc and not risk_title:
                    continue

                # Kategori önceliği: hücre > current_category > kod prefix > Genel
                if cat_cell:
                    category = cat_cell
                elif current_category:
                    category = current_category
                else:
                    category = guess_category_from_code(code) or "Genel"

                # Kategori sağlamlaştırma
                if category:
                    if category.strip() == text.strip():
                        category = current_category or guess_category_from_code(code) or "Genel"
                    elif _looks_like_sentence(category) and ("RİSKLER" not in category.upper()):
                        category = current_category or guess_category_from_code(code) or "Genel"

                # RiskCategory tablosuna da yaz
                if category:
                    rc = (
                        RiskCategory.query
                        .filter(func.lower(RiskCategory.name) == func.lower(category))
                        .first()
                    )
                    if not rc:
                        db.session.add(RiskCategory(name=category, is_active=True))

                # P/S değerleri
                p_val = _toi(r[prob_col]) if (prob_col is not None and prob_col < len(r)) else None
                s_val = _toi(r[sev_col])  if (sev_col  is not None and sev_col  < len(r)) else None

                # Tekillik: kategori + text kombinasyonu
                existing = Suggestion.query.filter(
                    Suggestion.category == (category or ""),
                    Suggestion.text == text
                ).first()

                if existing:
                    changed = False
                    if p_val and not existing.default_prob:
                        existing.default_prob = p_val
                        changed = True
                    if s_val and not existing.default_sev:
                        existing.default_sev = s_val
                        changed = True
                    if code and not existing.risk_code:
                        existing.risk_code = code
                        changed = True
                    # YENİ: risk_title / risk_desc / mitigation_hint güncelle
                    if risk_title is not None and (existing.risk_title or "") != risk_title:
                        existing.risk_title = risk_title
                        changed = True
                    if risk_desc is not None and (existing.risk_desc or "") != risk_desc:
                        existing.risk_desc = risk_desc
                        changed = True
                    if mitigation_hint is not None and (existing.mitigation_hint or "") != mitigation_hint:
                        existing.mitigation_hint = mitigation_hint
                        changed = True

                    if changed:
                        db.session.add(existing)
                        updated_cnt += 1
                    else:
                        skipped += 1
                    continue

                # Yeni kayıt
                db.session.add(Suggestion(
                    category        = category or "",
                    text            = text,               # Risk Faktörü (kısa ifade)
                    risk_code       = code or None,
                    default_prob    = p_val,
                    default_sev     = s_val,
                    risk_title      = risk_title,         # "Risk" / "Risk Adı"
                    risk_desc       = risk_desc,          # Risk Tanımı
                    mitigation_hint = mitigation_hint,    # Risk Azaltıcı Önlemler
                ))
                created += 1

            db.session.commit()
            flash(
                f"İçe aktarma tamamlandı. Eklenen: {created}, güncellenen: {updated_cnt}, atlanan: {skipped}.",
                "success",
            )
            return redirect(url_for("risk_identify"))

        # GET → basit upload formu
        return render_template("import_suggestions.html")



    # -------------------------------------------------
    #  Kütüphane Dışa Aktar (CSV / XLSX) — Sadece admin
    # -------------------------------------------------
    @app.route("/admin/export/suggestions.csv")
    @role_required("admin")
    def export_suggestions_csv():
        rows = Suggestion.query.order_by(Suggestion.category.asc(), Suggestion.text.asc()).all()
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(["Risk Kodu", "Kategori", "Öneri Metni", "Vars. P", "Vars. Ş", "Oluşturma", "Güncelleme"])
        for s in rows:
            writer.writerow([
                s.risk_code or "",
                s.category or "",
                s.text or "",
                s.default_prob or "",
                s.default_sev or "",
                s.created_at.strftime("%Y-%m-%d %H:%M") if getattr(s, "created_at", None) else "",
                s.updated_at.strftime("%Y-%m-%d %H:%M") if getattr(s, "updated_at", None) else "",
            ])
        resp = Response(output.getvalue(), mimetype="text/csv; charset=utf-8")
        resp.headers["Content-Disposition"] = "attachment; filename=suggestions_export.csv"
        return resp

    @app.route("/admin/export/suggestions.xlsx")
    @role_required("admin")
    def export_suggestions_xlsx():
        if not _pd:
            flash("Excel dışa aktarmak için 'pandas' + 'openpyxl' gerekli.", "danger")
            return redirect(url_for("risk_identify"))
        rows = Suggestion.query.order_by(Suggestion.category.asc(), Suggestion.text.asc()).all()
        data = []
        for s in rows:
            data.append({
                "Risk Kodu": s.risk_code or "",
                "Kategori": s.category or "",
                "Öneri Metni": s.text or "",
                "Vars. P": s.default_prob or "",
                "Vars. Ş": s.default_sev or "",
                "Oluşturma": s.created_at.strftime("%Y-%m-%d %H:%M") if getattr(s, "created_at", None) else "",
                "Güncelleme": s.updated_at.strftime("%Y-%m-%d %H:%M") if getattr(s, "updated_at", None) else "",
            })
        df = _pd.DataFrame(data)
        bio = io.BytesIO()
        with _pd.ExcelWriter(bio, engine="openpyxl") as wr:
            df.to_excel(wr, index=False, sheet_name="Suggestions")
        bio.seek(0)
        resp = Response(bio.read(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp.headers["Content-Disposition"] = "attachment; filename=suggestions_export.xlsx"
        return resp
    
        # -------------------------------------------------
    #  ADMIN — Kullanıcı Yönetimi
    # -------------------------------------------------
    @app.post("/admin/users/<int:uid>/assign-ref")
    @role_required("admin")
    def admin_assign_ref(uid):
        acc = Account.query.get_or_404(uid)

        # 1) Girdi: form ya da JSON
        raw = (
            request.form.get("ref_code")
            or (request.get_json(silent=True) or {}).get("ref_code")
            or ""
        ).strip().upper()

        # 2) Format doğrulaması
        PATTERN = r"^PRJ-[A-Z0-9]{6}$"
        if raw and not re.fullmatch(PATTERN, raw):
            flash("Geçersiz referans kodu formatı (örn. PRJ-ABC123).", "danger")
            return redirect(url_for("admin_users"))

        # 3) Kod üretimi (boş bırakılmışsa otomatik üret)
        code = raw
        MAX_TRIES = 8
        tries = 0
        while not code:
            tries += 1
            candidate = _gen_ref_code(prefix="PRJ")
            exists = Account.query.filter(Account.ref_code == candidate).first()
            if not exists:
                code = candidate
                break
            if tries >= MAX_TRIES:
                flash("Referans kodu üretilemedi, lütfen tekrar deneyin.", "danger")
                return redirect(url_for("admin_users"))

        # 4) Başka bir kullanıcıda var mı?
        clash = Account.query.filter(
            Account.ref_code == code,
            Account.id != acc.id
        ).first()
        if clash:
            flash("Bu referans kodu başka bir kullanıcıda mevcut.", "danger")
            return redirect(url_for("admin_users"))

        # 5) Atama ve commit
        acc.ref_code = code
        acc.status = "active"

        try:
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            flash("Veritabanı hatası: referans kodu atanamadı (unique kısıtı).", "danger")
            return redirect(url_for("admin_users"))

        # 6) Opsiyonel bilgilendirme maili
        try:
            if acc.email:
                ok, err = send_email(
                    to_email=acc.email,
                    subject="RiskApp — Referans Kodunuz",
                    body=(
                        f"Merhaba {acc.contact_name},\n\n"
                        f"Sistem üzerinde giriş yaparken kullanacağınız referans kodunuz:\n\n"
                        f"    {code}\n\n"
                        "Giriş için e-posta + şifre + referans kodu gereklidir.\n\n"
                        "Teşekkürler."
                    )
                )
                if not ok:
                    flash(f"Ref. kodu atandı fakat e-posta gönderilemedi: {err}", "warning")
                else:
                    flash(f"Referans kodu atandı ve e-posta gönderildi: {code}", "success")
            else:
                flash(f"Referans kodu atandı: {code}", "success")
        except Exception as e:
            flash(f"Referans kodu atandı fakat e-posta gönderilemedi: {e}", "warning")

        return redirect(url_for("admin_users"))

 

    @app.route("/admin/users", methods=["GET", "POST"])
    @role_required("admin")
    def admin_users():
        """
        Kullanıcı yönetimi:
        - GET: liste + istatistik
        - POST: formdaki action'a göre rol / durum / ref kod işlemleri
        """

        # ----- POST: butonlara basılınca -----
        if request.method == "POST":
            action = (request.form.get("action") or "").strip()
            uid_raw = request.form.get("user_id") or ""
            if not uid_raw.isdigit():
                flash("Geçersiz kullanıcı bilgisi.", "danger")
                return redirect(url_for("admin_users"))

            uid = int(uid_raw)
            acc = Account.query.get(uid)
            if not acc:
                flash("Kullanıcı bulunamadı.", "danger")
                return redirect(url_for("admin_users"))

            is_self = (acc.id == session.get("account_id"))

            # 1) Rol güncelle
            if action == "set_role":
                new_role = (request.form.get("new_role") or "").strip()
                if new_role not in ("admin", "uzman"):
                    flash("Geçersiz rol seçimi.", "danger")
                else:
                    acc.role = new_role
                    db.session.commit()
                    flash("Kullanıcı rolü güncellendi.", "success")

            # 2) Durum güncelle
            elif action == "set_status":
                new_status = (request.form.get("new_status") or "").strip()
                if new_status not in ("pending", "active", "disabled"):
                    flash("Geçersiz durum seçimi.", "danger")
                elif is_self and new_status != acc.status:
                    flash("Kendi hesabınızın durumunu değiştiremezsiniz.", "warning")
                else:
                    acc.status = new_status
                    db.session.commit()
                    flash("Kullanıcı durumu güncellendi.", "success")

            # 3) Ref kod ata / güncelle (boş ise otomatik üret)
            elif action == "assign_ref":
                # Formdan kod al (boş ise otomatik üretilecek)
                raw = (request.form.get("ref_code") or "").strip().upper()

                PATTERN = r"^PRJ-[A-Z0-9]{6}$"
                if raw and not re.fullmatch(PATTERN, raw):
                    flash("Geçersiz referans kodu formatı (örn. PRJ-ABC123).", "danger")
                    return redirect(url_for("admin_users"))

                # Kod üretimi (boş bırakıldıysa otomatik)
                code = raw
                MAX_TRIES = 8
                tries = 0
                while not code:
                    tries += 1
                    candidate = _gen_ref_code(prefix="PRJ")
                    exists = Account.query.filter(Account.ref_code == candidate).first()
                    if not exists:
                        code = candidate
                        break
                    if tries >= MAX_TRIES:
                        flash("Referans kodu üretilemedi, lütfen tekrar deneyin.", "danger")
                        return redirect(url_for("admin_users"))

                # Başka kullanıcıda var mı?
                clash = Account.query.filter(
                    Account.ref_code == code,
                    Account.id != acc.id
                ).first()
                if clash:
                    flash("Bu referans kodu başka bir kullanıcıda mevcut.", "danger")
                    return redirect(url_for("admin_users"))

                # Atama + active yap
                acc.ref_code = code
                acc.status = "active"

                try:
                    db.session.commit()
                except IntegrityError:
                    db.session.rollback()
                    flash("Veritabanı hatası: referans kodu atanamadı (unique kısıtı).", "danger")
                    return redirect(url_for("admin_users"))

                # Checkbox'a göre mail gönder
                notify = request.form.get("notify_email") == "1"
                if notify and acc.email:
                    try:
                        ok, err = send_email(
                            to_email=acc.email,
                            subject="RiskApp — Referans Kodunuz",
                            body=(
                                f"Merhaba {acc.contact_name},\n\n"
                                f"Sistem üzerinde giriş yaparken kullanacağınız referans kodunuz:\n\n"
                                f"    {code}\n\n"
                                "Giriş için e-posta + şifre + referans kodu gereklidir.\n\n"
                                "Teşekkürler."
                            )
                        )
                        if not ok:
                            flash(f"Ref. kodu atandı fakat e-posta gönderilemedi: {err}", "warning")
                        else:
                            flash(f"Referans kodu atandı ve e-posta gönderildi: {code}", "success")
                    except Exception as e:
                        flash(f"Ref. kodu atandı fakat e-posta gönderilemedi: {e}", "warning")
                else:
                    flash(f"Referans kodu atandı: {code}", "success")

            # 4) Ref kod temizle
            elif action == "clear_ref":
                acc.ref_code = None
                db.session.commit()
                flash("Referans kodu silindi.", "success")

            else:
                flash("Bilinmeyen işlem.", "danger")

            return redirect(url_for("admin_users"))

        # ----- GET: sayfayı listele -----
        users = Account.query.order_by(Account.created_at.desc()).all()

        totals = {
            "all": len(users),
            "active": sum(1 for u in users if (u.status or "pending") == "active"),
            "pending": sum(1 for u in users if (u.status or "pending") == "pending"),
            "disabled": sum(1 for u in users if (u.status or "pending") == "disabled"),
            "admins": sum(1 for u in users if (u.role or "uzman") == "admin"),
        }

        return render_template(
            "admin_users.html",
            users=users,
            totals=totals,
        )


    @app.post("/admin/users/<int:uid>/set-status")
    @role_required("admin")
    def admin_users_set_status(uid):
        """
        Form: status=pending|active|disabled
        """
        acc = Account.query.get_or_404(uid)
        status = (request.form.get("status") or "").strip()
        if status not in ("pending", "active", "disabled"):
            flash("Geçersiz durum seçimi.", "danger")
            return redirect(url_for("admin_users"))

        acc.status = status
        db.session.commit()
        flash("Kullanıcı durumu güncellendi.", "success")
        return redirect(url_for("admin_users"))

    @app.post("/admin/users/<int:uid>/delete")
    @role_required("admin")
    def admin_users_delete(uid):
        """
        Kullanıcı silme (kendi hesabını silemez).
        """
        acc = Account.query.get_or_404(uid)

        if acc.id == session.get("account_id"):
            flash("Kendi hesabınızı silemezsiniz.", "danger")
            return redirect(url_for("admin_users"))

        db.session.delete(acc)
        db.session.commit()
        flash("Kullanıcı silindi.", "success")
        return redirect(url_for("admin_users"))

    




    @app.post("/admin/risks/<int:rid>/set-ref")
    @role_required("admin")
    def admin_set_risk_ref(rid):
        r = Risk.query.get_or_404(rid)
        code = (request.form.get("ref_code") or "").strip().upper()

        if not code:
            flash("Ref No boş olamaz.", "danger")
            return redirect(url_for("risk_detail", risk_id=r.id))

        if not _REF_PATTERN.match(code):
            flash("Ref No formatı hatalı. Örn: R-PRJ12-2025-0034", "danger")
            return redirect(url_for("risk_detail", risk_id=r.id))

        exists = db.session.execute(
            text("SELECT id FROM risks WHERE ref_code = :c AND id != :id LIMIT 1"),
            {"c": code, "id": r.id}
        ).fetchone()
        if exists:
            flash("Bu Ref No başka bir kayıtta kullanılıyor.", "danger")
            return redirect(url_for("risk_detail", risk_id=r.id))

        r.ref_code = code
        db.session.commit()
        flash("Ref No güncellendi.", "success")
        return redirect(url_for("risk_detail", risk_id=r.id))


    @app.get("/admin/users/<int:uid>/compose-ref")
    @role_required("admin")
    def admin_compose_ref(uid):
        acc = Account.query.get_or_404(uid)
        if not acc.ref_code:
            flash("Bu kullanıcıya henüz referans kodu atanmadı.", "warning")
            return redirect(url_for("admin_users"))

        subject = "Referans Kodunuz"
        body = (
            f"Merhaba {acc.contact_name},\n\n"
            f"Referans Kodunuz: {acc.ref_code}\n"
            "Girişte e-posta + şifre + referans kodu gereklidir.\n\n"
            "İyi çalışmalar."
        )

        gmail_url = (
            "https://mail.google.com/mail/?view=cm&fs=1"
            f"&to={quote(acc.email)}"
            f"&su={quote(subject)}"
            f"&body={quote(body)}"
        )
        return redirect(gmail_url)



    # -------------------------------------------------
    #  Proje değiştir
    # -------------------------------------------------
    @app.route("/projects/switch", methods=["POST"])
    def switch_project():
        pid = request.form.get("project_id")
        acc_id = session.get("account_id")
        if not pid or not acc_id:
            return redirect(url_for("dashboard"))
        p = ProjectInfo.query.filter_by(id=int(pid), account_id=acc_id).first()
        if p:
            session["project_id"] = p.id
            flash(f"Aktif proje: {p.workplace_name}", "success")
        else:
            flash("Bu projeye erişiminiz yok.", "danger")
        return redirect(request.referrer or url_for("dashboard"))
    
           # --- AI Nasıl Çalışır (animasyonlu anlatım) ---
    @app.route("/ai/how-it-works")
    def ai_how_it_works():
        return render_template("ai_how_it_works.html")
    
    @app.route("/mitigations")
    def mitigations_list():
        # 🔐 Login kontrolü
        if "account_id" not in session:
            return redirect(url_for("login", next=request.path))

        account_id = session["account_id"]
        project_id = request.args.get("project_id", type=int)

        # Mitigation + Risk join
        q = Mitigation.query.join(Risk, Mitigation.risk_id == Risk.id)

        # Proje filtreleme (isteğe bağlı)
        if project_id:
            q = q.filter(Risk.project_id == project_id)

        mitigations = q.order_by(Mitigation.id.desc()).all()

        # Hesaba bağlı projeleri çek (dropdown için)
        projects = (
            ProjectInfo.query
            .filter(ProjectInfo.account_id == account_id)
            .order_by(ProjectInfo.workplace_name)
            .all()
        )

        return render_template(
            "mitigations.html",          # ✅ yeni dosyamız
            mitigations=mitigations,
            projects=projects,
            selected_project_id=project_id,
            )

        




    # -------------------------------------------------
    #  AI — RAG tabanlı aksiyon/mitigation önerisi (TEMİZLENMİŞ)
    # -------------------------------------------------
    @app.route("/ai/suggest/<int:risk_id>", methods=["POST"])
    def ai_suggest(risk_id):
        r = Risk.query.get_or_404(risk_id)

        # 0) Mitigation'daki eski AI metnini ayıkla (feedback loop fix)
        clean_mit = _strip_ai_in_mitigation(r.mitigation)
        base_mit = (clean_mit or (r.mitigation or "")).strip()

        # 1) Bağlam: aynı kategorideki öneriler
        ctx_suggestions = (
            Suggestion.query
            .filter(Suggestion.category == (r.category or ""))
            .order_by(Suggestion.id.desc())
            .limit(50)
            .all()
        )
        ctx_text = "\n".join(
            f"- {s.text} (P:{s.default_prob or '-'}, S:{s.default_sev or '-'})"
            for s in ctx_suggestions
        ) or "- (bağlam bulunamadı)"

        # 2) P/S tahmini (sayısal bağlam) — hata verirse app çökmemesi için try/except
        hint = None
        rpn_ai = None
        numeric_line = ""
        try:
            ps = PSEstimator(alpha=5.0)
            ps.fit(db.session)
            hint = ps.suggest(r.category or None)
        except Exception as e:
            current_app.logger.exception("PSEstimator hata verdi: %s", e)
            hint = None

        if hint and hint.get("p") and hint.get("s"):
            try:
                rpn_ai = int(hint["p"]) * int(hint["s"])
                numeric_line = (
                    f"Tahmini Olasılık **P={hint['p']}**, "
                    f"Şiddet **S={hint['s']}**, "
                    f"Tahmini RPN ≈ **{rpn_ai}**."
                )
            except Exception:
                numeric_line = (
                    f"Tahmini Olasılık **P={hint.get('p', '-')}**, "
                    f"Şiddet **S={hint.get('s', '-')}**."
                )
        else:
            numeric_line = "Tahmini P/S değeri üretilemedi (yetersiz veri ya da model hatası)."

        # 3) Prompt'ı hazırla
        title = r.title or "(başlık yok)"
        desc  = r.description or "(açıklama yok)"
        cat   = r.category or "(kategori yok)"

        mit_block = base_mit if base_mit else "- (tanımlı mevcut önlem yok)"

        prompt = f"""
    Sen bir inşaat/altyapı projeleri için çalışan uzman bir risk yönetimi danışmanısın.

    Aşağıdaki risk için, uygulanabilir ve sahada yapılabilir nitelikte 3–7 arası aksiyon/mitigation maddesi üret:

    - Kısa, net, madde madde yaz.
    - Her madde tek bir aksiyonu anlatsın.
    - Gereksiz uzun girişler, tekrarlar ve “bu sadece bir öneridir” gibi ifadeler kullanma.
    - Aynı şeyi farklı cümlelerle tekrar etme.
    - ISO 31000, FMEA ve inşaat sahası pratikleriyle uyumlu olsun.

    RİSK BİLGİSİ
    ------------
    - Başlık: {title}
    - Kategori: {cat}
    - Açıklama: {desc}

    MEVCUT ÖNLEMLER
    ----------------
    {mit_block}

    SAYISAL ÖZET
    ------------
    {numeric_line}

    BENZER ŞABLONLARDAN NOTLAR
    --------------------------
    {ctx_text}

    Lütfen sadece doğrudan kullanılabilir aksiyon/mitigation maddelerini üret.
    "BENZER ÖNERİLER" gibi başlıklar ekleme, giriş/sonuç paragrafı yazma.
    """

        # 4) OpenAI / local LLM çağrısı (+ fallback: _propose_actions)
        cleaned = ""
        try:
            raw = ai_complete(prompt)
            cleaned = _strip_ai_artifacts(raw or "").strip()
        except Exception as e:
            current_app.logger.exception("AI önerisi alınırken hata: %s", e)
            cleaned = ""

        # Eğer AI hiçbir şey veremediyse → _propose_actions fallback
        if not cleaned:
            try:
                actions = _propose_actions(r)
            except Exception as e2:
                current_app.logger.exception("_propose_actions hata verdi: %s", e2)
                actions = []

            if actions:
                cleaned_lines = [
                    f"- {a['action']} (Termin: {a['due']})"
                    for a in actions
                ]
                cleaned = "\n".join(cleaned_lines)
            else:
                flash("Ne AI ne de hazır aksiyon seti anlamlı bir öneri üretemedi.", "warning")
                return redirect(url_for("risk_detail", risk_id=r.id))

        # 5) Mitigation alanına ekle (mevcut metni bozmadan altına AI bloğu koy)
        ts = datetime.utcnow().isoformat(timespec="seconds")
        header = f"---\n🤖 AI Önerisi ({ts} UTC):\n"

        if base_mit:
            new_mit = f"{base_mit}\n\n{header}{cleaned}"
        else:
            new_mit = f"{header}{cleaned}"

        r.mitigation = new_mit

        # 6) Sistem yorumu düş
        db.session.add(Comment(
            risk_id=r.id,
            text=f"AI mitigation önerisi oluşturuldu: {ts} UTC",
            is_system=True,
        ))
        db.session.commit()

        flash("AI önerisi mitigation alanına eklendi.", "success")
        return redirect(url_for("risk_detail", risk_id=r.id))





    @app.route("/categories", methods=["GET", "POST"])
    def categories_index():
        q = (request.args.get("q") or "").strip()
        query = RiskCategory.query
        if q:
            like = f"%{q}%"
            query = query.filter(or_(
                RiskCategory.name.ilike(like),
                RiskCategory.code.ilike(like),
                RiskCategory.description.ilike(like)
            ))
        categories = query.order_by(RiskCategory.is_active.desc(), RiskCategory.name.asc()).all()

        if request.method == "POST":
            name = (request.form.get("name") or "").strip()
            if not name:
                flash("Kategori adı zorunludur.", "danger")
                return redirect(url_for("categories_index", next=request.args.get("next")))

            code = (request.form.get("code") or "").strip() or None
            color = (request.form.get("color") or "").strip() or None
            description = (request.form.get("description") or "").strip() or None

            # İstersen name unique değilse bunu kaldırabilirsin; senin mevcut davranışın aynı kalsın diye bıraktım:
            if RiskCategory.query.filter_by(name=name).first():
                flash("Bu isimde kategori zaten var.", "danger")
                return redirect(url_for("categories_index", next=request.args.get("next")))

            cat = RiskCategory(name=name, code=code, color=color, description=description, is_active=True)
            db.session.add(cat)

            try:
                db.session.commit()
            except IntegrityError:
                db.session.rollback()
                flash("Kaydedilemedi. Kod benzersiz olmalı veya veri kısıtı var.", "danger")
                return redirect(url_for("categories_index", next=request.args.get("next")))

            flash("Kategori eklendi.", "success")

            if _should_go_identify():
                return redirect(url_for("risk_identify"))

            return redirect(url_for("categories_index"))

        return render_template("categories.html", categories=categories, q=q)


    @app.route("/categories/<int:cid>/edit", methods=["POST"])
    def categories_edit(cid):
        cat = RiskCategory.query.get_or_404(cid)

        # ✅ Eski adı yakala (Suggestion.category güncellemesi için)
        old_name = (cat.name or "").strip()

        name = (request.form.get("name") or cat.name).strip()
        if not name:
            flash("Kategori adı zorunludur.", "danger")
            return redirect(url_for("categories_index", next=request.args.get("next")))

        # form alanlarını güncelle
        cat.name = name
        cat.code = (request.form.get("code") or "").strip() or None
        cat.color = (request.form.get("color") or "").strip() or None
        cat.description = (request.form.get("description") or "").strip() or None
        cat.is_active = _truthy(request.form.get("is_active"))

        try:
            # ✅ kategori adı değiştiyse Suggestion.category string’lerini de taşı
            new_name = (cat.name or "").strip()
            if old_name and new_name and old_name != new_name:
                Suggestion.query.filter(Suggestion.category == old_name).update(
                    {Suggestion.category: new_name},
                    synchronize_session=False
                )

            db.session.commit()

        except IntegrityError:
            db.session.rollback()
            flash("Güncellenemedi. Kod benzersiz olmalı veya veri kısıtı var.", "danger")
            return redirect(url_for("categories_index", next=request.args.get("next")))

        flash("Kategori güncellendi.", "success")

        if _should_go_identify():
            return redirect(url_for("risk_identify"))
        return redirect(url_for("categories_index"))



    @app.route("/categories/<int:cid>/delete", methods=["POST"])
    def categories_delete(cid):
        # 1) kategoriyi al
        cat = RiskCategory.query.get_or_404(cid)

        try:
            # 2) hard delete dene
            db.session.delete(cat)
            db.session.commit()
            flash("Kategori silindi.", "success")

        except IntegrityError:
            # 3) FK vb. yüzden silinemedi -> rollback
            db.session.rollback()

            # ✅ KRİTİK: objeyi yeniden yükle (rollback sonrası state karışmasın)
            cat2 = RiskCategory.query.get(cid)
            if cat2:
                cat2.is_active = False
                db.session.commit()

            flash("Kategori kullanımda olduğu için silinemedi, pasif yapıldı.", "warning")

        # 4) nereye dönecek?
        if _should_go_identify():
            return redirect(url_for("risk_identify"))
        return redirect(url_for("categories_index"))

    # -------------------------
    # API (JSON) endpointler
    # -------------------------
    @app.get("/api/categories")
    def api_categories_list():
        q = (request.args.get("q") or "").strip()
        query = RiskCategory.query
        if q:
            like = f"%{q}%"
            query = query.filter(or_(
                RiskCategory.name.ilike(like),
                RiskCategory.code.ilike(like),
                RiskCategory.description.ilike(like)
            ))
        rows = query.order_by(RiskCategory.is_active.desc(), RiskCategory.name.asc()).all()

        return jsonify([
            {
                "id": r.id,
                "name": r.name,
                "code": r.code,
                "color": r.color,
                "description": r.description,
                "is_active": bool(r.is_active),
            }
            for r in rows
        ])


    @app.post("/api/categories")
    def api_categories_create():
        name = (request.form.get("name") or "").strip()
        if not name:
            return jsonify({"error": "name required"}), 400

        code = (request.form.get("code") or "").strip() or None
        color = (request.form.get("color") or "").strip() or None
        description = (request.form.get("description") or "").strip() or None

        # Senin eski davranışın: name duplicate ise 409
        if RiskCategory.query.filter_by(name=name).first():
            return jsonify({"error": "duplicate name"}), 409

        cat = RiskCategory(name=name, code=code, color=color, description=description, is_active=True)
        db.session.add(cat)

        try:
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            return jsonify({"error": "duplicate code or constraint error"}), 409

        return jsonify({"ok": True, "id": cat.id})


    @app.patch("/api/categories/<int:cid>")
    def api_categories_update(cid):
        cat = RiskCategory.query.get_or_404(cid)
        data = request.form

        def norm(v): return (v or "").strip()

        # Frontend her zaman name gönderiyor; name boş olursa 400
        nm = norm(data.get("name"))
        if not nm:
            return jsonify({"error": "name required"}), 400

        cat.name = nm
        cat.code = norm(data.get("code")) or None
        cat.color = norm(data.get("color")) or None
        cat.description = norm(data.get("description")) or None

        # checkbox unchecked ise JS "" gönderiyor -> False
        if "is_active" in data:
            cat.is_active = _truthy(data.get("is_active"))

        try:
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            return jsonify({"error": "duplicate code or constraint error"}), 409

        return jsonify({"ok": True})


    @app.delete("/api/categories/<int:cid>")
    def api_categories_delete(cid):
        cat = RiskCategory.query.get_or_404(cid)

        try:
            db.session.delete(cat)
            db.session.commit()
            return jsonify({"ok": True, "deleted": True})
        except IntegrityError:
            # kullanımda ise: soft delete (pasif)
            db.session.rollback()
            cat.is_active = False
            db.session.commit()
            return jsonify({
                "ok": True,
                "deleted": False,
                "message": "Kullanımda olduğu için silinmedi; pasif yapıldı."
            }), 200


    # --- Kategori yardımcıları (aktif adlar) ---
    def active_category_names():
        rows = (RiskCategory.query
                .filter(RiskCategory.is_active == True)
                .order_by(RiskCategory.name.asc())
                .all())
        return [r.name for r in rows]


    @app.get("/api/category-names")
    def api_category_names():
        return jsonify(active_category_names())
    @app.post("/api/categories/<int:cid>/delete")
    def api_categories_delete_post(cid):
            cat = RiskCategory.query.get_or_404(cid)

            try:
                db.session.delete(cat)
                db.session.commit()
                return jsonify({"ok": True, "deleted": True})
            except IntegrityError:
                # kullanımda ise: soft delete (pasif)
                db.session.rollback()
                cat.is_active = False
                db.session.commit()
                return jsonify({
                    "ok": True,
                    "deleted": False,
                    "message": "Kullanımda olduğu için silinmedi; pasif yapıldı."
                }), 200

    
    

    

    

    @app.post("/api/categories/<int:cid>/update")
    def api_categories_update_post(cid):
        cat = RiskCategory.query.get_or_404(cid)

        def norm(v): 
            return (v or "").strip()

        nm = norm(request.form.get("name"))
        if not nm:
            return jsonify({"error": "name required"}), 400

        cat.name = nm
        cat.code = norm(request.form.get("code")) or None
        cat.color = norm(request.form.get("color")) or None
        cat.description = norm(request.form.get("description")) or None

        # checkbox gelmezse False say
        cat.is_active = _truthy(request.form.get("is_active"))

        try:
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            return jsonify({"error": "duplicate code or constraint error"}), 409

        return jsonify({"ok": True})


        
    
    # -------------------------------------------------
    #  ADMIN — Tek seferlik prefix'e göre kategori düzeltme (opsiyonel)
    # -------------------------------------------------
    @app.post("/admin/tools/fix-suggestion-cats")
    @role_required("admin")
    def admin_fix_suggestion_cats():
        mapping = {
            "YÖR": "YÖNETSEL RİSKLER",
            "SOR": "SÖZLEŞME / ONAY SÜREÇLERİ",
            "UYR": "UYGULAMA / YAPIM RİSKLERİ",
            "GER": "ZEMİN KOŞULLARI / GEOTEKNİK",
            "ÇER": "ÇEVRESEL RİSKLER",
            "CER": "ÇEVRESEL RİSKLER",
            "DTR": "DENETİM / TETKİK / RAPOR",
            "POR": "POLİTİK / ORGANİZASYONEL",
            "TYR": "TEDARİK / MALZEME",
        }
        fixed = 0
        for s in Suggestion.query.all():
            code = (s.risk_code or "").upper().strip()
            if not code:
                continue
            letters = "".join([c for c in code if c.isalpha()])[:3]
            new_cat = mapping.get(letters)
            if not new_cat:
                continue
            # mevcut kategori boşsa veya barizce yanlışsa düzelt
            if not s.category or s.category.upper() in {code, ""}:
                s.category = new_cat
                fixed += 1
        db.session.commit()
        flash(f"Kategori düzeltme tamam: {fixed} kayıt güncellendi.", "success")
        return redirect(url_for("risk_identify"))

    # -------------------------------------------------
    #  ADMIN — Mevcut AI yorumlarını temizle
    # -------------------------------------------------
    @app.post("/admin/tools/clean-ai-comments")
    @role_required("admin")
    def admin_clean_ai_comments():
        patt_ai_head = re.compile(r"^\s*🤖\s*AI Önerisi:\s*", re.I)
        changed, skipped = 0, 0
        rows = Comment.query.filter(Comment.is_system == True).all()
        for c in rows:
            raw = c.text or ""
            # Sadece AI başlıklı olanları hedefleyelim
            if "AI Önerisi" not in raw:
                skipped += 1
                continue
            # Başlığı ayıkla, gövdeyi temizle
            body = patt_ai_head.sub("", raw, count=1)
            body = _strip_ai_artifacts(body)
            body = body.strip()
            if not body:
                # Boş kaldıysa yorumu sil
                db.session.delete(c)
                changed += 1
                continue
            # Tek, temiz başlık yeniden ekle
            c.text = "🤖 AI Önerisi:\n" + body
            changed += 1
        db.session.commit()
        flash(f"AI yorum temizliği tamamlandı. Güncellenen/silinen: {changed}, atlanan: {skipped}.", "success")
        return redirect(url_for("risk_select"))
    

    # ======= Takvim API'ları (JSON feed + tarih güncelle + ICS export) =======
    api = Blueprint("api_v1", __name__)

    def _require_login_or_abort():
        if "username" not in session:
            abort(401)

    @api.get("/schedule/events")
    def api_schedule_events():
        """
        Takvim/FullCalendar beslemesi.
        İsteğe bağlı filtreler: q, category, owner, status
        """
        _require_login_or_abort()

        q      = (request.args.get("q") or "").strip()
        cat    = (request.args.get("category") or "").strip()
        owner  = (request.args.get("owner") or "").strip()
        status = (request.args.get("status") or "").strip()

        pid = _get_active_project_id()
        query = Risk.query
        if pid:
            query = query.filter(Risk.project_id == pid)

        if q:
            like = f"%{q}%"
            query = query.filter(
                (Risk.title.ilike(like)) |
                (Risk.category.ilike(like)) |
                (Risk.description.ilike(like)) |
                (Risk.responsible.ilike(like))
            )
        if cat:
            query = query.filter(Risk.category == cat)
        if owner:
            query = query.filter(Risk.responsible == owner)
        if status:
            query = query.filter(Risk.status == status)

        rows = query.order_by(Risk.updated_at.desc()).all()

        def first_day(ym: str | None) -> str | None:
            return f"{ym}-01" if ym else None

        def last_day(ym: str | None) -> str | None:
            if not ym:
                return None
            y, m = _parse_ym(ym) or (None, None)
            if not y:
                return None
            ny, nm = _next_ym(y, m)
            return (date(ny, nm, 1) - timedelta(days=1)).isoformat()

        events = []
        for r in rows:
            s, e = r.start_month, r.end_month
            if s and not e:
                e = s
            if e and not s:
                s = e
            if not s and not e:
                continue

            start_iso = first_day(s)
            end_incl  = last_day(e)
            end_excl  = (datetime.fromisoformat(end_incl) + timedelta(days=1)).date().isoformat() if end_incl else None

            # risk seviyesi → className
            _gmap = {"high": "critical", "medium": "moderate", "low": "low", "none": "acceptable"}
            gname = _gmap.get((r.grade() or "none").lower(), "acceptable")

            events.append({
                "id": r.id,
                "title": (r.title or "Risk"),
                "start": start_iso,
                "end": end_excl,           # FullCalendar end exclusive kullanır
                "allDay": True,
                "className": [f"gx-{gname}"],
                "extendedProps": {
                    "category": r.category,
                    "status": r.status,
                    "responsible": r.responsible,
                    "rpn": r.avg_rpn(),
                    "start_month": r.start_month,
                    "end_month": r.end_month,
                }
            })

        return jsonify(events)

    @api.patch("/risks/<int:risk_id>/dates")
    def api_risk_update_dates(risk_id: int):
        """
        Sürükle-bırak/yeniden boyutlandırma sonrası tarih güncellemesi.
        Body JSON: { "start": "YYYY-MM-DD", "end": "YYYY-MM-DD" }  # end exclusive
        """
        _require_login_or_abort()
        r = Risk.query.get_or_404(risk_id)
        data = request.get_json(force=True, silent=True) or {}

        def to_ym(d: str | None) -> str | None:
            return d[:7] if d else None

        start_d = data.get("start")
        end_d   = data.get("end")

        r.start_month = to_ym(start_d)

        if end_d:
            try:
                end_inc = datetime.fromisoformat(end_d[:10]) - timedelta(days=1)
                r.end_month = f"{end_inc.year:04d}-{end_inc.month:02d}"
            except Exception:
                return jsonify({"ok": False, "error": "invalid end date"}), 400
        else:
            r.end_month = r.start_month

        # tek commit sürümü:
        db.session.add(Comment(
            risk_id=r.id,
            text=f"Tarih güncellendi: {r.start_month or '—'} → {r.end_month or '—'}",
            is_system=True
        ))
        db.session.commit()

        return jsonify({"ok": True})

    @app.get("/admin/tools/test-mail")
    @role_required("admin")
    def admin_test_mail():
        acc = Account.query.get(session.get("account_id"))
        to = (request.args.get("to") or (acc.email if acc else None) or "").strip()
        if not to:
            flash("Alıcı e-posta bulunamadı. ?to=mail@ornek.com ile deneyin.", "warning")
            return redirect(url_for("admin_users"))

        ok = send_email(
            to_email=to,
            subject="SMTP Test — RiskApp",
            body="Bu bir test mesajıdır. SMTP ayarlarınız çalışıyor. 📬"
        )
        flash("Test e-postası gönderildi." if ok else "E-posta gönderimi başarısız. Log’a bakınız.",
            "success" if ok else "danger")
        return redirect(url_for("admin_users"))

    @api.get("/schedule/export/ics")
    def api_schedule_export_ics():
        """
        Aynı filtrelerle (.ics) takvim dışa aktarımı.
        Parametreler: q, category, owner, status
        """
        _require_login_or_abort()

        q      = (request.args.get("q") or "").strip()
        cat    = (request.args.get("category") or "").strip()
        owner  = (request.args.get("owner") or "").strip()
        status = (request.args.get("status") or "").strip()

        pid = _get_active_project_id()
        query = Risk.query
        if pid:
            query = query.filter(Risk.project_id == pid)
        if q:
            like = f"%{q}%"
            query = query.filter(
                (Risk.title.ilike(like)) |
                (Risk.category.ilike(like)) |
                (Risk.description.ilike(like)) |
                (Risk.responsible.ilike(like))
            )
        if cat:
            query = query.filter(Risk.category == cat)
        if owner:
            query = query.filter(Risk.responsible == owner)
        if status:
            query = query.filter(Risk.status == status)

        rows = query.order_by(Risk.updated_at.desc()).all()

        def first_day(ym: str | None) -> str | None:
            return f"{ym}-01" if ym else None

        def last_day(ym: str | None) -> str | None:
            if not ym:
                return None
            y, m = _parse_ym(ym) or (None, None)
            if not y:
                return None
            ny, nm = _next_ym(y, m)
            return (date(ny, nm, 1) - timedelta(days=1)).isoformat()

        lines = ["BEGIN:VCALENDAR", "VERSION:2.0", "PRODID:-//RiskApp//Schedule//TR"]
        for r in rows:
            s, e = r.start_month, r.end_month
            if s and not e: e = s
            if e and not s: s = e
            if not s and not e:
                continue

            dtstart = (first_day(s) or "")[:10].replace("-", "")
            last = last_day(e)
            dtend = (datetime.fromisoformat(last) + timedelta(days=1)).date().isoformat().replace("-", "") if last else ""

            title = (r.title or "").replace("\n", " ").replace("\r", " ")
            lines += [
                "BEGIN:VEVENT",
                f"UID:risk-{r.id}@riskapp",
                f"DTSTART;VALUE=DATE:{dtstart}",
                f"DTEND;VALUE=DATE:{dtend}",
                f"SUMMARY:{title}",
                "END:VEVENT"
            ]
        lines += ["END:VCALENDAR"]
        ics = "\r\n".join(lines)
        return Response(
            ics,
            mimetype="text/calendar; charset=utf-8",
            headers={"Content-Disposition": "attachment; filename=risk_schedule.ics"}
        )

    app.register_blueprint(api, url_prefix="/api")

    # performans için yardımcı indeksler (varsayılan SQLite'ta idempotent)
    with app.app_context():
        try:
            db.session.execute(text("CREATE INDEX IF NOT EXISTS ix_risks_project ON risks(project_id)"))
            db.session.execute(text("CREATE INDEX IF NOT EXISTS ix_risks_start   ON risks(start_month)"))
            db.session.execute(text("CREATE INDEX IF NOT EXISTS ix_risks_end     ON risks(end_month)"))
            db.session.commit()
        except Exception:
            pass

    @app.post("/api/risks/<int:rid>/set-months")
    def api_set_months(rid):
        r = Risk.query.get_or_404(rid)
        sm = (request.form.get("start_month") or "").strip() or None
        em = (request.form.get("end_month")  or "").strip() or None

        # YYYY-MM formatını çok basit doğrula
        def _ok(ym):
            if not ym: return True
            try:
                y,m = ym.split("-")
                y,m = int(y), int(m)
                return 1 <= m <= 12 and 1900 <= y <= 2100
            except Exception:
                return False

        if not _ok(sm) or not _ok(em):
            return jsonify({"ok": False, "error": "bad format"}), 400

        r.start_month = sm
        r.end_month   = em
        db.session.commit()
        return jsonify({"ok": True, "start_month": r.start_month, "end_month": r.end_month})   


    @app.route("/risks/<int:risk_id>/mitigations", methods=["GET", "POST"])
    def mitigations_list_create(risk_id):
        r = Risk.query.get_or_404(risk_id)

        if request.method == "POST":
            title = (request.form.get("title") or "").strip()
            if not title:
                flash("Başlık (title) zorunlu.", "error")
                return redirect(url_for("mitigations_list_create", risk_id=risk_id))

            m = Mitigation(
                risk_id=r.id,
                title=title,
                owner=(request.form.get("owner") or None),
                status=(request.form.get("status") or "planned"),
                due_date=_parse_date(request.form.get("due_date")),
                cost=_to_float(request.form.get("cost")),
                effectiveness=_to_int(request.form.get("effectiveness")),
                notes=(request.form.get("notes") or None),
            )
            db.session.add(m)
            db.session.commit()
            flash("Mitigation eklendi.", "success")
            return redirect(url_for("mitigations_list_create", risk_id=risk_id))

        return render_template("mitigations_list.html", r=r)

    # --- CRUD: Düzenleme ---
    @app.route("/mitigations/<int:mid>/edit", methods=["GET", "POST"])
    def mitigation_edit(mid):
        m = Mitigation.query.get_or_404(mid)
        r = m.risk

        if request.method == "POST":
            title = (request.form.get("title") or "").strip()
            if not title:
                flash("Başlık (title) zorunlu.", "error")
                return redirect(url_for("mitigation_edit", mid=mid))

            m.title = title
            m.owner = (request.form.get("owner") or None)
            m.status = (request.form.get("status") or "planned")
            m.due_date = _parse_date(request.form.get("due_date"))
            m.cost = _to_float(request.form.get("cost"))
            m.effectiveness = _to_int(request.form.get("effectiveness"))
            m.notes = (request.form.get("notes") or None)
            db.session.commit()
            flash("Mitigation güncellendi.", "success")
            return redirect(url_for("mitigations_list_create", risk_id=r.id))

        return render_template("mitigation_edit.html", r=r, m=m)

    # --- CRUD: Silme ---
    @app.route("/mitigations/<int:mid>/delete", methods=["POST"])
    def mitigation_delete(mid):
        m = Mitigation.query.get_or_404(mid)
        rid = m.risk_id
        db.session.delete(m)
        db.session.commit()
        flash("Mitigation silindi.", "success")
        return redirect(url_for("mitigations_list_create", risk_id=rid))

    # -------------------------------------------------
    #  PDF Rapor (WeasyPrint -> pdfkit fallback)
    # -------------------------------------------------
    @app.get("/risks/<int:risk_id>/report.pdf")
    def risk_report_pdf(risk_id: int):
        risk = Risk.query.get_or_404(risk_id)

        # ✅ Bu risk’e bağlı maliyet kalemleri
        cost_items = (
            CostItem.query
            .filter(CostItem.risk_id == risk.id)
            .order_by(CostItem.id.desc())
            .all()
        )

        # ✅ Para birimine göre toplamlar (TRY/USD/EUR ayrı ayrı)
        cost_totals = (
            db.session.query(
                CostItem.currency,
                func.coalesce(func.sum(CostItem.total), 0)
            )
            .filter(CostItem.risk_id == risk.id)
            .group_by(CostItem.currency)
            .all()
        )

        # (opsiyonel) suggestion’ları da aynı template kullanıyorsan ver
        suggestions = Suggestion.query.filter(Suggestion.category == (risk.category or "")).all()

        # ✅ Aynı HTML şablonunu kullanıyoruz (print-friendly CSS zaten içinde)
        html_str = render_template(
            "report_view.html",
            r=risk,
            cost_items=cost_items,
            cost_totals=cost_totals,
            suggestions=suggestions,
        )

        # --- 1) WeasyPrint dene (varsa ve çalışabiliyorsa) ---
        if HTML and CSS:
            try:
                pdf_bytes = HTML(string=html_str, base_url=request.url_root).write_pdf(
                    stylesheets=[CSS(string="""
                        @page { size: A4; margin: 12mm; }
                        * { -webkit-print-color-adjust: exact; print-color-adjust: exact; }
                        thead { display: table-header-group; }
                        body { background: #fff; color: #111; }
                    """)]
                )
                return Response(
                    pdf_bytes,
                    mimetype="application/pdf",
                    headers={"Content-Disposition": f'inline; filename="risk_{risk_id}.pdf"'}
                )
            except Exception as e:
                print("WeasyPrint çalışmadı, pdfkit'e geçiliyor:", e)

        # --- 2) pdfkit (wkhtmltopdf) fallback ---
        if pdfkit:
            try:
                wkhtml = _guess_wkhtmltopdf_path()
                config = pdfkit.configuration(wkhtmltopdf=wkhtml) if wkhtml else None
            except Exception as e:
                print("pdfkit configuration error:", e)
                config = None

            pdf_bytes = pdfkit.from_string(
                html_str,
                False,
                configuration=config,
                options={
                    "page-size": "A4",
                    "margin-top": "12mm",
                    "margin-right": "12mm",
                    "margin-bottom": "12mm",
                    "margin-left": "12mm",
                    "encoding": "UTF-8",
                    "enable-local-file-access": None,
                },
            )
            return Response(
                pdf_bytes,
                mimetype="application/pdf",
                headers={"Content-Disposition": f'inline; filename="risk_{risk_id}.pdf"'}
            )

        return Response(
            "PDF üretimi için uygun backend bulunamadı. WeasyPrint için GTK/Pango/Cairo, "
            "ya da wkhtmltopdf kurulumu gerekir.",
            status=500,
            mimetype="text/plain; charset=utf-8",
        )
    @app.context_processor
    def _endpoint_utils():
        def has_endpoint(name: str) -> bool:
            return name in current_app.view_functions
        return dict(has_endpoint=has_endpoint)
    
    @app.route("/debug/ai_comment/<int:risk_id>")
    def debug_ai_comment(risk_id):
        text = make_ai_risk_comment(risk_id)
        # Çok basic: plain text döndürelim
        return f"<pre>{text}</pre>"
    
    
    @app.route("/risks/export.csv")
    def risks_export_csv():
        pid    = _get_active_project_id()
        q      = (request.args.get("q") or "").strip()
        status = (request.args.get("status") or "").strip()

        query = Risk.query
        if pid:
            query = query.filter(Risk.project_id == pid)
        if q:
            like = f"%{q}%"
            query = query.filter(
                (Risk.title.ilike(like)) |
                (Risk.category.ilike(like)) |
                (Risk.description.ilike(like))
            )
        if status:
            query = query.filter(Risk.status == status)

        risks = query.order_by(Risk.category.asc().nullsfirst(), Risk.id.asc()).all()

        output = StringIO()
        writer = csv.writer(output)

        writer.writerow([
            "No",
            "Risk Adı",
            "Risk Tanımlaması",
            "Risk Sahibi",           # r.owner
            "P",
            "S",
            "Risk Seviyesi",
            "Karşı Önlemler",
            "Kategori",
            "Durum",
            "Sorumlu",              # r.responsible
            "Başlangıç(YYYY-MM)",
            "Bitiş(YYYY-MM)",
        ])

        def level_for_rpn(rpn: float | None) -> str:
            """1–4 Düşük, 5–10 Orta, 11–15 Yüksek, 16–25 Çok Yüksek."""
            if rpn is None:
                return ""
            r = float(rpn)
            if r <= 4:
                return "Düşük"
            if r <= 10:
                return "Orta"
            if r <= 15:
                return "Yüksek"
            return "Çok Yüksek"

        from collections import defaultdict
        counters = defaultdict(int)

        for r in risks:
            key = (r.category or "GENEL RİSKLER").strip()
            counters[key] += 1

            # SON değerlendirme P/S
            last_eval = None
            if r.evaluations:
                last_eval = sorted(r.evaluations, key=lambda e: e.id)[-1]

            if last_eval and last_eval.probability is not None and last_eval.severity is not None:
                p_val = float(last_eval.probability)
                s_val = float(last_eval.severity)
            else:
                p_val = r.avg_prob()
                s_val = r.avg_sev()

            # RPN: r.score()
            sc = None
            score_fn = getattr(r, "score", None)
            if callable(score_fn):
                try:
                    sc = score_fn()
                    sc = float(sc) if sc is not None else None
                except Exception:
                    sc = None
            if sc is None and p_val is not None and s_val is not None:
                sc = float(p_val) * float(s_val)

            writer.writerow([
                counters[key],                             # No
                r.title or "",                             # Risk Adı
                r.description or "",                       # Risk Tanımlaması
                getattr(r, "owner", "") or "",             # Risk Sahibi (oluşturan kişi)
                f"{p_val:.2f}" if p_val is not None else "",   # P (son değerlendirme)
                f"{s_val:.2f}" if s_val is not None else "",   # S
                level_for_rpn(sc),                         # Risk Seviyesi
                r.mitigation or "",                        # Karşı Önlemler
                r.category or "",                          # Kategori
                r.status or "",                            # Durum
                r.responsible or "",                       # Sorumlu
                r.start_month or "",                       # Başlangıç(YYYY-MM)
                r.end_month or "",                         # Bitiş(YYYY-MM)
            ])

        resp = Response(output.getvalue(), mimetype="text/csv; charset=utf-8")
        resp.headers["Content-Disposition"] = "attachment; filename=risks_export.csv"
        return resp

    
        # -------------------------------------------------
    #  Mevcut riskleri birleştirme (ADMIN)
    # -------------------------------------------------
    @app.post("/risks/merge")
    @role_required("admin")
    def risks_merge():
        """
        /risks ekranından seçilen riskleri tek bir raporda birleştirir.
        Beklenen form field:
          - risk_ids: "3,5,7" gibi virgüllü ID listesi
          - title (opsiyonel): yeni risk başlığı
        """
        raw_ids = (request.form.get("risk_ids") or "").strip()
        if not raw_ids:
            flash("Birleştirmek için en az bir risk seçmelisiniz.", "danger")
            return redirect(url_for("risk_select"))

        try:
            ids = sorted({
                int(x) for x in raw_ids.split(",")
                if x.strip().isdigit()
            })
        except ValueError:
            flash("Geçersiz risk ID listesi.", "danger")
            return redirect(url_for("risk_select"))

        if len(ids) < 2:
            flash("Birleştirme için en az 2 risk seçmelisiniz.", "warning")
            return redirect(url_for("risk_select"))

        risks = (
            Risk.query
            .filter(Risk.id.in_(ids))
            .order_by(Risk.id.asc())
            .all()
        )
        if len(risks) < 2:
            flash("Yeterli sayıda geçerli risk bulunamadı.", "danger")
            return redirect(url_for("risk_select"))

        # Aynı projeye ait olduklarından emin ol (değilse ilk projeye zorlayacağız)
        first = risks[0]
        pid   = first.project_id
        for r in risks:
            if r.project_id != pid:
                flash("Farklı projelere ait riskler birleştiriliyor. Yeni risk ilk projenin altında oluşturulacak.", "warning")
                break

        # Kategori: ilk dolu kategori, yoksa "Genel"
        cat = None
        for r in risks:
            if (r.category or "").strip():
                cat = r.category.strip()
                break
        cat = cat or "Genel"

        # Yeni title: formdan gelen veya ilk risk + "(Birleştirilmiş)"
        title_form = (request.form.get("title") or "").strip()
        new_title  = title_form or f"{first.title or 'Birleştirilmiş Risk'} (Birleştirilmiş)"

        # Açıklama: önce kısa bir üst bilgi, sonra tek tek risklerin detayları
        desc_lines = []
        desc_lines.append("Bu kayıt aşağıdaki risklerin birleştirilmesiyle oluşturulmuştur:\n")
        for r in risks:
            desc_lines.append(f"- [#{r.id}] {r.title or ''}")
        desc_lines.append("\n--- Ayrıntılı açıklamalar ---\n")
        for r in risks:
            if r.description:
                desc_lines.append(f"### Risk #{r.id}: {r.title or ''}")
                desc_lines.append(r.description)
                desc_lines.append("")  # boş satır

        final_desc = "\n".join(desc_lines).strip()

        # Mitigation alanı (Risk.mitigation text): eskilerin mitigation'larını birleştir
        mit_lines = []
        for r in risks:
            if (r.mitigation or "").strip():
                mit_lines.append(f"- [#{r.id}] {r.mitigation.strip()}")
        mitigation_merged = "\n".join(mit_lines) if mit_lines else None

        owner       = session.get("username")
        responsible = (request.form.get("responsible") or "").strip() or first.responsible
        duration    = first.duration

        # Tarih aralığı: seçilen risklerin min(start), max(end)
        def _norm_ym_pair(sm, em):
            s = _parse_ym(sm)
            e = _parse_ym(em)
            if s and not e:
                e = s
            if e and not s:
                s = e
            if s and e and s > e:
                s, e = e, s
            return s, e

        min_ym, max_ym = None, None
        for r in risks:
            s, e = _norm_ym_pair(r.start_month, r.end_month)
            if s and ((min_ym is None) or (s < min_ym)):
                min_ym = s
            if e and ((max_ym is None) or (e > max_ym)):
                max_ym = e

        def _ym_or_none(t):
            return _ym_to_str(*t) if t else None

        start_month = _ym_or_none(min_ym)
        end_month   = _ym_or_none(max_ym)

        # Yeni risk kaydı
        new_risk = Risk(
            title       = new_title,
            category    = cat,
            description = final_desc,
            mitigation  = mitigation_merged,
            responsible = responsible,
            duration    = duration,
            start_month = start_month,
            end_month   = end_month,
            owner       = owner,
            project_id  = pid,
            status      = "Merged",
        )
        db.session.add(new_risk)
        db.session.flush()  # id lazım

        # Eski risklerin değerlendirmelerini yeni riske taşı
        for r in risks:
            for ev in getattr(r, "evaluations", []):
                db.session.add(Evaluation(
                    risk_id    = new_risk.id,
                    evaluator  = ev.evaluator,
                    probability= ev.probability,
                    severity   = ev.severity,
                    detection  = ev.detection,
                    comment    = f"[Eski #{r.id}] {ev.comment or ''}",
                ))

        # Eski risklerin Mitigation satırlarını yeni riske kopyala
        for r in risks:
            m_rows = Mitigation.query.filter_by(risk_id=r.id).all()
            for m in m_rows:
                db.session.add(Mitigation(
                    risk_id  = new_risk.id,
                    text     = f"[Eski #{r.id}] {m.text}",
                    owner    = m.owner,
                    status   = m.status,
                    due_date = m.due_date,
                ))

        # Eski risklere sistem notu + status güncelle
        now_txt = datetime.utcnow().isoformat(timespec="seconds") + " UTC"
        merged_ids_str = ", ".join(f"#{r.id}" for r in risks)
        for r in risks:
            r.status = "Merged"
            db.session.add(Comment(
                risk_id = r.id,
                text    = f"Bu risk, yeni birleştirilmiş kayıt altında toplandı: {merged_ids_str} ({now_txt})",
                is_system=True
            ))

        # Yeni risk için de açıklayıcı sistem notu
        db.session.add(Comment(
            risk_id = new_risk.id,
            text    = f"Birleştirilmiş risk oluşturuldu; kaynak riskler: {merged_ids_str} ({now_txt})",
            is_system=True
        ))

        db.session.commit()
        flash(f"{len(risks)} risk tek bir raporda birleştirildi (Yeni ID: {new_risk.id}).", "success")
        return redirect(url_for("risk_detail", risk_id=new_risk.id))
    @app.post("/risks/split/<int:risk_id>")
    def risk_split(risk_id: int):
        """Birleşik bir riski description içindeki --- bloklarına göre parçalara böler."""
        if session.get("role") != "admin":
            abort(403)

        r = Risk.query.get_or_404(risk_id)

        raw = (r.description or "").strip()
        if not raw:
            flash("Bu riskin açıklaması boş, ayıracak bir içerik yok.", "warning")
            return redirect(url_for("risk_detail", risk_id=r.id))

        # Birleştirmede kullandığımız ayrım: \n\n---\n\n
        parts = [p.strip() for p in raw.split("\n\n---\n\n") if p.strip()]

        if len(parts) < 2:
            flash("Bu kayıt birleştirilmiş formatta görünmüyor; ayırma yapılmadı.", "warning")
            return redirect(url_for("risk_detail", risk_id=r.id))

        created = 0

        for idx, part in enumerate(parts, start=1):
            lines = [ln for ln in part.splitlines() if ln.strip()]
            if not lines:
                continue

            first_line = lines[0].strip()
            body = "\n".join(lines[1:]).strip() or None

            # İlk satırda [#id] Başlık formatını yakala
            m = _re.match(r"\[#(\d+)\]\s*(.+)", first_line)
            if m:
                title = (m.group(2) or "").strip() or f"{r.title} · Bölüm {idx}"
            else:
                title = first_line or f"{r.title} · Bölüm {idx}"

            new_risk = Risk(
                title=title[:255],
                description=body,
                category=getattr(r, "category", None),
            )

            if hasattr(Risk, "project_id"):
                new_risk.project_id = getattr(r, "project_id", None)

            db.session.add(new_risk)
            created += 1

        if created == 0:
            flash("Ayırma sırasında yeni kayıt oluşturulamadı.", "warning")
            return redirect(url_for("risk_detail", risk_id=r.id))

        db.session.commit()
        flash(f"Risk {created} parçaya ayrıldı ve ayrı kayıtlar oluşturuldu.", "success")
        return redirect(url_for("risk_select"))
    
    
    @app.route("/risk-templates/<int:sid>")
    def risk_template_detail(sid):
        s = Suggestion.query.get_or_404(sid)
        return render_template("risk_template_detail.html", s=s)
        
        
    def _to_decimal(v, default="0"):
            try:
                if v is None or str(v).strip() == "":
                    return Decimal(default)
                return Decimal(str(v).replace(",", "."))
            except (InvalidOperation, ValueError):
                return Decimal(default)

    def _active_project_id():
        pid = session.get("active_project_id") or session.get("project_id")
        if pid is not None:
            try:
                return int(pid)
            except (TypeError, ValueError):
                # bozuk değer olursa temizleyebilirsin
                session.pop("active_project_id", None)
                return None

        acc_id = session.get("account_id")
        if acc_id:
            prj = (
                ProjectInfo.query
                .filter_by(account_id=acc_id)
                .order_by(ProjectInfo.id.desc())
                .first()
            )
            if prj:
                session["active_project_id"] = prj.id
                return prj.id

        return None


    def _annual_factor(freq: str) -> Decimal:
        # Tek Sefer: 1 bırakıyorum (istersen 0 yapıp “yıllık karşılaştırma”dan çıkarabilirsin)
        if freq == "Aylık":
            return Decimal("12")
        if freq == "Yıllık":
            return Decimal("1")
        return Decimal("1")
    # -------------------------------------------------
# Helpers
# -------------------------------------------------
    

    # -------------------------------------------------
    # COSTS (GET + POST)
    # -------------------------------------------------
    



    @app.route("/costs", methods=["GET", "POST"])
    def costs():
        project_id = _active_project_id()
        if not project_id:
            flash("Aktif proje bulunamadı. Önce proje seç.", "warning")
            return redirect(url_for("dashboard"))

        # -------------------------
        # POST
        # -------------------------
        if request.method == "POST":
            title = (request.form.get("title") or "").strip()
            if not title:
                flash("Başlık zorunlu.", "warning")
                return redirect(url_for("costs"))

            category = (request.form.get("category") or "").strip()
            unit = (request.form.get("unit") or "").strip()
            currency = (request.form.get("currency") or "TRY").strip() or "TRY"
            frequency = (request.form.get("frequency") or "Tek Sefer").strip() or "Tek Sefer"

            # Dropdown boş seçildiyse DB'ye None basmasın diye
            if not category:
                flash("Kategori zorunlu.", "warning")
                return redirect(url_for("costs"))
            if not unit:
                flash("Birim zorunlu.", "warning")
                return redirect(url_for("costs"))

            qty = _to_decimal(request.form.get("qty"), "1")
            unit_price = _to_decimal(request.form.get("unit_price"), "0")

            # Sunucu tarafı sağlam dursun
            if qty <= 0:
                flash("Miktar 0’dan büyük olmalı.", "warning")
                return redirect(url_for("costs"))
            if unit_price < 0:
                flash("Birim fiyat negatif olamaz.", "warning")
                return redirect(url_for("costs"))

            # risk_id: hem sayı mı, hem bu projeye mi ait?
            risk_id = None
            risk_id_raw = (request.form.get("risk_id") or "").strip()
            if risk_id_raw.isdigit():
                cand = int(risk_id_raw)
                exists = (
                    Risk.query
                    .filter(Risk.id == cand, Risk.project_id == project_id)
                    .first()
                )
                if exists:
                    risk_id = cand
                else:
                    flash("Seçilen risk bu projeye ait değil. Risk bağlanmadı.", "warning")

            item = CostItem(
                project_id=project_id,
                risk_id=risk_id,
                title=title,
                category=category,   # ✅ artık zorunlu
                unit=unit,           # ✅ artık zorunlu
                currency=currency,
                frequency=frequency,
                qty=qty,
                unit_price=unit_price,
                description=(request.form.get("description") or "").strip() or None,
            )

            # total hesap
            try:
                item.total = qty * unit_price
            except Exception:
                item.total = Decimal("0")

            db.session.add(item)
            db.session.commit()

            flash("Maliyet kaydedildi.", "success")
            return redirect(url_for("costs"))

        # -------------------------
        # GET
        # -------------------------
        costs = (
            CostItem.query
            .filter_by(project_id=project_id)
            .order_by(CostItem.id.desc())
            .all()
        )

        cost_templates = (
            CostTemplate.query
            .filter_by(project_id=project_id)
            .order_by(CostTemplate.id.desc())
            .all()
        )

        # ✅ projeye ait riskleri template’e gönder (dropdown için)
        risks = (
            Risk.query
            .filter(Risk.project_id == project_id)
            .order_by(Risk.id.desc())
            .all()
        )

        # Pareto (Decimal hesap)
        sorted_costs = sorted(costs, key=lambda c: (c.total or Decimal("0")), reverse=True)
        grand = sum((c.total or Decimal("0")) for c in sorted_costs) or Decimal("0")

        run = Decimal("0")
        pareto = []
        for c in sorted_costs:
            val = (c.total or Decimal("0"))
            run += val
            cum = (run / grand * Decimal("100")) if grand > 0 else Decimal("0")
            pareto.append({
                "label": c.title,
                "value": float(val),     # Chart.js float ister
                "cum_pct": float(cum),
            })

        # Pareto Front: riskleri tek seferde çek (N+1 yok)
        risk_ids = sorted({c.risk_id for c in costs if c.risk_id})
        risk_map = {}
        if risk_ids:
            risks_for_front = (
                Risk.query
                .filter(Risk.project_id == project_id, Risk.id.in_(risk_ids))
                .all()
            )
            risk_map = {r.id: r for r in risks_for_front}

        front = []
        for c in costs:
            if not c.risk_id:
                continue
            r = risk_map.get(c.risk_id)
            if not r:
                continue

            s = r.score()
            if s is None:
                continue

            x_total = (c.total or Decimal("0"))
            front.append({
                "x": float(x_total),
                "y": float(s),
                "label": c.title
            })

        return render_template(
            "costs.html",
            costs=costs,
            cost_templates=cost_templates,
            risks=risks,                      # ✅ eklendi
            cost_categories=COST_CATEGORIES,  # ✅ eklendi (kategori dropdown için)
            pareto_json=pareto,
            front_json=front,
        )

    # -------------------------------------------------
    # COST EDIT (GET)
    # -------------------------------------------------
    @app.get("/costs/<int:cost_id>/edit")
    def cost_edit(cost_id):
        project_id = _active_project_id()
        if not project_id:
            flash("Aktif proje yok.", "warning")
            return redirect(url_for("dashboard"))

        c = CostItem.query.filter_by(id=cost_id, project_id=project_id).first()
        if not c:
            flash("Maliyet bulunamadı.", "warning")
            return redirect(url_for("costs"))

        # ✅ projeye ait riskleri dropdown için gönder
        risks = (
            Risk.query
            .filter(Risk.project_id == project_id)
            .order_by(Risk.id.desc())
            .all()
        )

        # ✅ kategori dropdown listesi (cost_edit.html bunu kullanacak)
        cost_categories = ["İş Gücü", "Ekipman", "Yazılım", "Eğitim", "Hizmet", "Operasyon"]

        return render_template(
            "cost_edit.html",
            c=c,
            risks=risks,
            cost_categories=cost_categories,
        )

# -------------------------------------------------
# COSTS: BULK ATTACH (POST)
# -------------------------------------------------
    @app.post("/costs/attach")
    def costs_attach():
        project_id = _active_project_id()
        if not project_id:
            flash("Aktif proje yok.", "warning")
            return redirect(url_for("dashboard"))

        risk_id_raw = (request.form.get("risk_id") or "").strip()
        if not risk_id_raw.isdigit():
            flash("Risk seçilemedi.", "danger")
            return redirect(url_for("costs"))

        risk_id = int(risk_id_raw)

        # risk bu projeye ait mi?
        r = Risk.query.filter(Risk.id == risk_id, Risk.project_id == project_id).first()
        if not r:
            flash("Bu risk bu projeye ait değil.", "danger")
            return redirect(url_for("costs"))

        ids = request.form.getlist("cost_ids")
        ids = [int(x) for x in ids if str(x).isdigit()]
        if not ids:
            flash("Hiç maliyet seçmedin.", "warning")
            return redirect(url_for("costs", risk_id=risk_id))

        # sadece bu projeye ait costlar
        items = CostItem.query.filter(
            CostItem.project_id == project_id,
            CostItem.id.in_(ids)
        ).all()

        if not items:
            flash("Seçilen maliyetler bulunamadı.", "warning")
            return redirect(url_for("costs", risk_id=risk_id))

        for c in items:
            c.risk_id = risk_id  # bağla (mevcut bağlıysa da yeniden bağlar)

        db.session.commit()
        flash(f"{len(items)} maliyet bu riske bağlandı.", "success")
        return redirect(url_for("risk_detail", risk_id=risk_id))
    

    @app.post("/costs/<int:cost_id>/delete")
    def cost_delete(cost_id):
        # Yetki: sen admin ile kısıtlamak istiyorsan aç:
        if session.get("role") != "admin":
            abort(403)

        c = Cost.query.get_or_404(cost_id)

        # Ekstra güvenlik: gerçekten bir riske bağlı mı? (istersen)
        # if not c.risk_id: abort(400)

        db.session.delete(c)
        db.session.commit()
        flash("Maliyet silindi.", "success")

        return redirect(request.referrer or url_for("index"))

    # -------------------------------------------------
    # COST EDIT (POST)
    # -------------------------------------------------
    @app.post("/costs/<int:cost_id>/edit")
    def cost_edit_post(cost_id):
        project_id = _active_project_id()
        if not project_id:
            flash("Aktif proje yok.", "warning")
            return redirect(url_for("dashboard"))

        c = CostItem.query.filter_by(id=cost_id, project_id=project_id).first()
        if not c:
            flash("Maliyet bulunamadı.", "warning")
            return redirect(url_for("costs"))

        # Basit doğrulama (front-end ile paralel)
        title = (request.form.get("title") or "").strip()
        category = (request.form.get("category") or "").strip()
        unit = (request.form.get("unit") or "").strip()
        currency = (request.form.get("currency") or "TRY").strip() or "TRY"
        frequency = (request.form.get("frequency") or "Tek Sefer").strip() or "Tek Sefer"

        try:
            qty = _to_decimal(request.form.get("qty"), "0")
            unit_price = _to_decimal(request.form.get("unit_price"), "0")
        except Exception:
            flash("Miktar ve birim fiyat sayısal olmalı.", "danger")
            return redirect(url_for("cost_edit", cost_id=cost_id))

        if not title or not category or not unit or qty <= 0 or unit_price <= 0:
            flash("Zorunlu alanları doğru doldur.", "danger")
            return redirect(url_for("cost_edit", cost_id=cost_id))

        desc = (request.form.get("description") or "").strip() or None

        # ✅ EKLENDİ: risk_id güncelle (bu projeye ait mi kontrol et)
        risk_id = None
        risk_id_raw = (request.form.get("risk_id") or "").strip()
        if risk_id_raw.isdigit():
            cand = int(risk_id_raw)
            exists = (
                Risk.query
                .filter(Risk.id == cand, Risk.project_id == project_id)
                .first()
            )
            if exists:
                risk_id = cand
            else:
                flash("Seçilen risk bu projeye ait değil. Risk bağlanmadı.", "warning")

        # Alanları güncelle
        c.title = title
        c.category = category
        c.unit = unit
        c.currency = currency
        c.frequency = frequency
        c.qty = qty
        c.unit_price = unit_price
        c.total = qty * unit_price
        c.description = desc

        # ✅ burası: commit'ten önce risk'i yaz
        c.risk_id = risk_id

        db.session.commit()
        flash("Maliyet güncellendi.", "success")
        return redirect(url_for("costs"))


    # -------------------------------------------------
    # COST DELETE (POST)
    # -------------------------------------------------
    @app.post("/costs/<int:cost_id>/delete")
    def cost_delete(cost_id):
        project_id = _active_project_id()
        if not project_id:
            flash("Aktif proje yok.", "warning")
            return redirect(url_for("dashboard"))

        item = CostItem.query.filter_by(id=cost_id, project_id=project_id).first()
        if not item:
            flash("Maliyet bulunamadı.", "warning")
            return redirect(url_for("costs"))

        db.session.delete(item)
        db.session.commit()
        flash("Maliyet silindi.", "success")
        return redirect(url_for("costs"))


    # -------------------------------------------------
    # COST TEMPLATE CREATE (POST)
    # -------------------------------------------------
    @app.post("/cost-templates/create")
    def cost_template_create():
        project_id = _active_project_id()
        if not project_id:
            flash("Aktif proje yok.", "warning")
            return redirect(url_for("dashboard"))

        title = (request.form.get("title") or "").strip()
        category = (request.form.get("category") or "").strip()
        unit = (request.form.get("unit") or "").strip()
        currency = (request.form.get("currency") or "TRY").strip() or "TRY"
        frequency = (request.form.get("frequency") or "Tek Sefer").strip() or "Tek Sefer"
        desc = (request.form.get("description") or "").strip() or None

        if not title or not category or not unit:
            flash("Şablon için başlık/kategori/birim zorunlu.", "danger")
            return redirect(url_for("costs"))

        t = CostTemplate(
            project_id=project_id,
            title=title,
            category=category,
            unit=unit,
            currency=currency,
            frequency=frequency,
            description=desc,
        )
        db.session.add(t)
        db.session.commit()
        flash("Şablon eklendi.", "success")
        return redirect(url_for("costs"))


    # -------------------------------------------------
    # COST TEMPLATE EDIT (POST)
    # -------------------------------------------------
    @app.post("/cost-templates/<int:tpl_id>/edit")
    def cost_template_edit_post(tpl_id):
        project_id = _active_project_id()
        if not project_id:
            flash("Aktif proje yok.", "warning")
            return redirect(url_for("dashboard"))

        t = CostTemplate.query.filter_by(id=tpl_id, project_id=project_id).first()
        if not t:
            flash("Şablon bulunamadı.", "warning")
            return redirect(url_for("costs"))

        title = (request.form.get("title") or "").strip()
        category = (request.form.get("category") or "").strip()
        unit = (request.form.get("unit") or "").strip()
        currency = (request.form.get("currency") or "TRY").strip() or "TRY"
        frequency = (request.form.get("frequency") or "Tek Sefer").strip() or "Tek Sefer"
        desc = (request.form.get("description") or "").strip() or None

        if not title or not category or not unit:
            flash("Şablon için başlık/kategori/birim zorunlu.", "danger")
            return redirect(url_for("costs"))

        t.title = title
        t.category = category
        t.unit = unit
        t.currency = currency
        t.frequency = frequency
        t.description = desc

        db.session.commit()
        flash("Şablon güncellendi.", "success")
        return redirect(url_for("costs"))


    # -------------------------------------------------
    # COST TEMPLATE DELETE (POST)
    # -------------------------------------------------
    @app.post("/cost-templates/<int:tpl_id>/delete")
    def cost_template_delete(tpl_id):
        project_id = _active_project_id()
        if not project_id:
            flash("Aktif proje yok.", "warning")
            return redirect(url_for("dashboard"))

        t = CostTemplate.query.filter_by(id=tpl_id, project_id=project_id).first()
        if not t:
            flash("Şablon bulunamadı.", "warning")
            return redirect(url_for("costs"))

        db.session.delete(t)
        db.session.commit()
        flash("Şablon silindi.", "success")
        return redirect(url_for("costs"))
        
    

    # -------------------------------------------------
#  Risk sepetini temizle (eski endpointi geri getir)
# -------------------------------------------------
    @app.route("/risk/basket/remove", methods=["POST"])
    def risk_basket_remove():
        # Sepeti session'dan sil
        session.pop("picked_rows", None)

        # İstersen flash mesajı da gösterebilir:
        flash("Risk sepeti temizlendi.", "info")

        # Tekrar şablon seçme ekranına dön
        return redirect(url_for("risk_identify"))
    
# -------------------------------------------------
#  Risk sepetini temizle (eski endpointi geri getir)
# -------------------------------------------------

    @app.get("/analytics/pareto")
    def pareto_cost():
        pid = _get_active_project_id()
        currency = (request.args.get("currency") or "TRY").upper()
        limit = int(request.args.get("limit") or 50)

        # Risk filtresi (aktif proje)
        rq = Risk.query
        if pid:
            rq = rq.filter(Risk.project_id == pid)

        risk_ids = [r.id for r in rq.with_entities(Risk.id).all()]
        if not risk_ids:
            return jsonify({"currency": currency, "items": [], "total": 0, "note": "No risks in scope"})

        # Risk bazlı toplam maliyet
        rows = (
            db.session.query(
                CostItem.risk_id,
                func.coalesce(func.sum(CostItem.total), 0).label("sum_total"),
            )
            .filter(CostItem.risk_id.in_(risk_ids))
            .filter(func.coalesce(CostItem.currency, "TRY") == currency)
            .group_by(CostItem.risk_id)
            .order_by(func.coalesce(func.sum(CostItem.total), 0).desc())
            .limit(limit)
            .all()
        )

        if not rows:
            return jsonify({"currency": currency, "items": [], "total": 0, "note": "No cost items for this currency"})

        # Risk bilgilerini tek seferde çek
        rid_list = [rid for rid, _ in rows]
        risk_map = {r.id: r for r in Risk.query.filter(Risk.id.in_(rid_list)).all()}

        total = sum(float(t or 0) for _, t in rows) or 0.0
        running = 0.0

        items = []
        for rid, t in rows:
            v = float(t or 0)
            running += v
            r = risk_map.get(rid)

            pct = (v / total) * 100 if total else 0
            cum = (running / total) * 100 if total else 0

            items.append({
                "risk_id": rid,
                "title": (r.title if r else f"Risk #{rid}"),
                "category": (r.category if r else None),
                "owner": (r.responsible if r else None),
                "value": round(v, 2),
                "pct": round(pct, 2),
                "cum_pct": round(cum, 2),
            })

        # 80% cutoff
        cutoff_index = next((i for i, it in enumerate(items) if it["cum_pct"] >= 80), len(items)-1)
        top_80 = items[:cutoff_index+1]

        return jsonify({
            "currency": currency,
            "total": round(total, 2),
            "top_80_count": len(top_80),
            "items": items,
        })
    
    
    @app.get("/analytics/pareto/view")
    def pareto_view():
        currency = (request.args.get("currency") or "TRY").upper()
        return render_template("pareto_view.html", currency=currency)
    
    @app.get("/api/cost-items")
    def api_cost_items():
        project_id = _active_project_id()
        if not project_id:
            return jsonify({"items": []})

        q = (request.args.get("q") or "").strip()
        only_unlinked = request.args.get("unlinked", "1") == "1"
        limit = min(int(request.args.get("limit", 80)), 200)

        qry = CostItem.query.filter(CostItem.project_id == project_id)

        if only_unlinked:
            qry = qry.filter(CostItem.risk_id.is_(None))

        if q:
            like = f"%{q}%"
            qry = qry.filter(or_(
                CostItem.title.ilike(like),
                CostItem.category.ilike(like),
                CostItem.currency.ilike(like),
                CostItem.unit.ilike(like),
            ))

        items = []
        for c in qry.order_by(CostItem.id.desc()).limit(limit).all():
            items.append({
                "id": c.id,
                "title": c.title,
                "category": c.category,
                "unit": c.unit,
                "currency": c.currency,
                "frequency": c.frequency,
                "qty": float(c.qty or 0),
                "unit_price": float(c.unit_price or 0),
                "total": float(c.total or 0),
                "risk_id": c.risk_id,
            })

        return jsonify({"items": items})
    
    @app.post("/api/risks/<int:risk_id>/cost-items/attach")
    def api_attach_cost_items(risk_id):
        project_id = _active_project_id()
        if not project_id:
            return jsonify({"ok": False, "error": "Aktif proje yok"}), 400

        r = Risk.query.filter(Risk.id == risk_id, Risk.project_id == project_id).first()
        if not r:
            return jsonify({"ok": False, "error": "Risk bulunamadı / proje dışı"}), 404

        data = request.get_json(force=True) or {}
        ids = data.get("cost_ids") or []
        mode = (data.get("mode") or "move").lower()  # move | copy

        if not ids:
            return jsonify({"ok": False, "error": "cost_ids boş"}), 400

        costs = (CostItem.query
                .filter(CostItem.project_id == project_id, CostItem.id.in_(ids))
                .all())

        moved = 0
        copied = 0

        if mode == "move":
            # güvenlik: istersen sadece boşta olanları taşı
            for c in costs:
                if c.risk_id is None:
                    c.risk_id = risk_id
                    moved += 1
            db.session.commit()
            return jsonify({"ok": True, "moved": moved, "copied": 0})

        if mode == "copy":
            for c in costs:
                newc = CostItem(
                    project_id=project_id,
                    risk_id=risk_id,
                    title=c.title,
                    category=c.category,
                    unit=c.unit,
                    currency=c.currency,
                    frequency=c.frequency,
                    qty=c.qty,
                    unit_price=c.unit_price,
                    description=c.description,
                    total=c.total,
                )
                db.session.add(newc)
                copied += 1
            db.session.commit()
            return jsonify({"ok": True, "moved": 0, "copied": copied})

        return jsonify({"ok": False, "error": "mode move|copy olmalı"}), 400

        
 




    @app.get("/analytics/pareto/ai")
    def pareto_cost_ai():
        # ----------------------------
        # helpers
        # ----------------------------
        def clamp_int(v, lo, hi, default):
            try:
                x = int(v)
                return max(lo, min(hi, x))
            except Exception:
                return default

        def clamp_float(v, lo, hi, default):
            try:
                x = float(v)
                if x != x:  # NaN
                    return default
                return max(lo, min(hi, x))
            except Exception:
                return default

        def norm_currency(s):
            s = (s or "TRY").strip().upper()
            return s if s in ("TRY", "USD", "EUR") else "TRY"

        def stext(x):
            return (x or "").strip()

        def as_float(x, default=0.0):
            try:
                return float(x)
            except Exception:
                return default

        def as_int(x, default=None):
            try:
                return int(x)
            except Exception:
                return default

        def compute_top80_count(items, total):
            """
            1) cum_pct varsa oradan
            2) yoksa value üzerinden 80% kümülatif yap
            """
            if not items or total <= 0:
                return 0

            # cum_pct var mı?
            has_cum = any(("cum_pct" in it) and (it.get("cum_pct") is not None) for it in items)
            if has_cum:
                for i, it in enumerate(items, start=1):
                    if as_float(it.get("cum_pct"), 0.0) >= 80.0:
                        return i
                return len(items)

            # cum_pct yoksa value ile hesapla (items zaten değer azalan sıralı varsayımı)
            cum = 0.0
            for i, it in enumerate(items, start=1):
                cum += as_float(it.get("value"), 0.0)
                if (cum / total) >= 0.80:
                    return i
            return len(items)

        def hhi(shares_0_1):
            # Herfindahl-Hirschman Index (0..1)
            return sum((s * s) for s in shares_0_1 if s > 0)

        # ----------------------------
        # params
        # ----------------------------
        currency = norm_currency(request.args.get("currency"))
        top_n = clamp_int(request.args.get("top_n"), 3, 50, 10)               # UI için
        scenario_cut = clamp_float(request.args.get("cut"), 0.0, 0.9, 0.10)   # 0.10 = %10
        scenario_scope = (request.args.get("scope") or "top3").strip().lower()  # top3 | top80 | topcat

        # Cache key
        cache_key = (currency, top_n, scenario_cut, scenario_scope)
        now = time.time()
        cached = _PARETO_AI_CACHE.get(cache_key)
        if cached and (now - cached[0]) <= _CACHE_TTL_SEC:
            return jsonify(cached[1])

        # ----------------------------
        # Pareto verisini içeriden al
        # ----------------------------
        resp = pareto_cost()  # senin mevcut fonksiyonun
        try:
            data = resp.get_json() if hasattr(resp, "get_json") else (resp or {})
        except Exception:
            data = {}

        items = (data.get("items") or [])
        total = as_float(data.get("total"), 0.0)
        top_80_count = as_int(data.get("top_80_count"), 0) or 0

        if not items or total <= 0:
            payload = {
                "currency": currency,
                "summary": "Bu para birimi için maliyet verisi yok (veya toplam 0).",
                "insights": [],
                "top_risks": [],
                "actions": [],
                "meta": {"total": total, "top_80_count": 0, "scenario_cut": scenario_cut, "scope": scenario_scope},
            }
            _PARETO_AI_CACHE[cache_key] = (now, payload)
            return jsonify(payload)

        # top_80_count güvenli hale getir
        if top_80_count <= 0:
            top_80_count = compute_top80_count(items, total)
            if top_80_count <= 0:
                top_80_count = min(len(items), 5)

        top_80_items = items[:top_80_count]
        top_risks_raw = items[:min(top_80_count, top_n)]

        # items içindeki id alanı bazen id bazen risk_id olabiliyor
        def item_rid(it):
            rid = it.get("id")
            if rid is None:
                rid = it.get("risk_id")
            return as_int(rid, None)

        top_ids = [item_rid(it) for it in top_80_items]
        top_ids = [x for x in top_ids if isinstance(x, int)]

        # ----------------------------
        # Kategori katkıları (top80)
        # ----------------------------
        cats = defaultdict(float)
        for it in top_80_items:
            c = stext(it.get("category")) or "GENEL"
            cats[c] += as_float(it.get("value"), 0.0)

        top_cats = sorted(cats.items(), key=lambda x: x[1], reverse=True)
        top_cats3 = top_cats[:3]
        top_cat = top_cats[0][0] if top_cats else "GENEL"
        top_cat_ratio = (top_cats[0][1] / sum(cats.values())) if cats else 0.0

        # Konsantrasyon (kategori) HHI
        cat_total = sum(cats.values()) or 1.0
        cat_shares = [v / cat_total for _, v in cats.items()]
        cat_hhi = hhi(cat_shares)

        # Konsantrasyon (risk) HHI (top80 risklerin değer paylaşımı)
        risk_total = sum(as_float(it.get("value"), 0.0) for it in top_80_items) or 1.0
        risk_shares = [as_float(it.get("value"), 0.0) / risk_total for it in top_80_items]
        risk_hhi = hhi(risk_shares)

        # ----------------------------
        # CostItem istatistikleri
        # ----------------------------
        cost_stats = {}                   # risk_id -> {n_items, sum_total}
        freq_stats = defaultdict(float)   # freq -> sum_total
        per_risk_freq = defaultdict(lambda: defaultdict(float))  # risk_id -> freq -> sum_total
        top_costitems_by_risk = defaultdict(list)  # risk_id -> [{name,total,frequency}...]

        if top_ids:
            # risk bazında adet + toplam
            rows = (
                db.session.query(
                    CostItem.risk_id,
                    func.count(CostItem.id).label("n_items"),
                    func.coalesce(func.sum(CostItem.total), 0).label("sum_total"),
                )
                .filter(CostItem.risk_id.in_(top_ids))
                .filter(func.upper(func.coalesce(CostItem.currency, "TRY")) == currency)
                .group_by(CostItem.risk_id)
                .all()
            )
            cost_stats = {int(rid): {"n_items": int(n), "sum_total": float(s)} for rid, n, s in rows}

            # periyot dağılımı (global)
            freq_rows = (
                db.session.query(
                    func.lower(func.coalesce(CostItem.frequency, "belirsiz")).label("freq"),
                    func.coalesce(func.sum(CostItem.total), 0).label("sum_total"),
                )
                .filter(CostItem.risk_id.in_(top_ids))
                .filter(func.upper(func.coalesce(CostItem.currency, "TRY")) == currency)
                .group_by("freq")
                .all()
            )
            for f, s in freq_rows:
                freq_stats[f] += float(s)

            # per risk periyot
            prf_rows = (
                db.session.query(
                    CostItem.risk_id,
                    func.lower(func.coalesce(CostItem.frequency, "belirsiz")).label("freq"),
                    func.coalesce(func.sum(CostItem.total), 0).label("sum_total"),
                )
                .filter(CostItem.risk_id.in_(top_ids))
                .filter(func.upper(func.coalesce(CostItem.currency, "TRY")) == currency)
                .group_by(CostItem.risk_id, "freq")
                .all()
            )
            for rid, f, s in prf_rows:
                per_risk_freq[int(rid)][f] += float(s)

            # En pahalı costitem kalemleri (top 3)
            # NOTE: CostItem.name alanı yoksa, title/description vb ile değiştir.
            ci_rows = (
                db.session.query(
                    CostItem.risk_id,
                    func.coalesce(getattr(CostItem, "name", None), "").label("name"),
                    CostItem.total,
                    func.lower(func.coalesce(CostItem.frequency, "belirsiz")).label("freq"),
                )
                .filter(CostItem.risk_id.in_(top_ids))
                .filter(func.upper(func.coalesce(CostItem.currency, "TRY")) == currency)
                .order_by(CostItem.risk_id.asc(), CostItem.total.desc())
                .all()
            )

            # Her risk için ilk 3’ü al
            seen = defaultdict(int)
            for rid, name, total_ci, freq in ci_rows:
                rid = int(rid)
                if seen[rid] >= 3:
                    continue
                seen[rid] += 1
                top_costitems_by_risk[rid].append({
                    "name": (name or "").strip() or "Kalem",
                    "total": float(total_ci or 0),
                    "frequency": freq or "belirsiz"
                })

        # ----------------------------
        # Quick win adayları: çok kalemli + yüksek toplam
        # ----------------------------
        quick_win_candidates = []
        for it in top_risks_raw:
            rid = item_rid(it)
            if rid is None:
                continue
            st = cost_stats.get(rid)
            if not st:
                continue
            avg = (st["sum_total"] / max(st["n_items"], 1))
            if st["n_items"] >= 5 or (st["n_items"] >= 3 and avg <= (0.05 * total)):
                quick_win_candidates.append({
                    "id": rid,
                    "title": it.get("title") or f"Risk #{rid}",
                    "n_items": st["n_items"],
                    "sum_total": st["sum_total"],
                    "avg_item": avg,
                })

        quick_win_candidates = sorted(
            quick_win_candidates, key=lambda x: (x["n_items"], x["sum_total"]), reverse=True
        )[:3]

        # ----------------------------
        # Scenario: scope bazlı tasarruf
        # ----------------------------
        if scenario_scope == "top80":
            base_scope_items = top_80_items
            scope_label = f"Top-80 bandı ({top_80_count} risk)"
        elif scenario_scope == "topcat":
            base_scope_items = [it for it in top_80_items if (stext(it.get("category")) or "GENEL") == top_cat]
            scope_label = f"Top kategori ({top_cat})"
        else:
            base_scope_items = top_80_items[:3]
            scope_label = "Top 3 risk"

        scope_total = sum(as_float(x.get("value"), 0.0) for x in base_scope_items)
        scenario_saving = scope_total * scenario_cut
        scenario_after = max(0.0, total - scenario_saving)

        # ----------------------------
        # Top riskleri zenginleştir (UI için)
        # ----------------------------
        top_risks = []
        for it in top_risks_raw:
            rid = item_rid(it)
            title = it.get("title") or (f"Risk #{rid}" if rid is not None else "Risk")
            value = as_float(it.get("value"), 0.0)
            pct = it.get("pct")
            cum_pct = it.get("cum_pct")

            st = cost_stats.get(rid) if rid is not None else None
            top_risks.append({
                "id": rid,
                "risk_id": rid,  # uyumluluk
                "title": title,
                "category": it.get("category") or "GENEL",
                "value": value,
                "pct": as_float(pct, (value / total * 100 if total else 0.0)),
                "cum_pct": as_float(cum_pct, None) if cum_pct is not None else None,
                "url": f"/reports/{rid}" if rid is not None else None,
                "cost_items": {
                    "n_items": st["n_items"] if st else 0,
                    "sum_total": st["sum_total"] if st else 0.0,
                    "freq_breakdown": dict(per_risk_freq.get(rid, {})) if rid is not None else {},
                    "top_cost_items": top_costitems_by_risk.get(rid, []) if rid is not None else [],
                }
            })

        # ----------------------------
        # Insights (kart listesi)
        # ----------------------------
        insights = []

        insights.append({
            "type": "pareto",
            "title": "80/20 Özeti",
            "text": f"Toplam {total:,.2f} {currency} maliyetin %80’ini yaklaşık {top_80_count} risk üretiyor."
        })

        if top_risks:
            t0 = top_risks[0]
            insights.append({
                "type": "top",
                "title": "En Büyük Katkı",
                "text": f"En büyük katkı: “{t0['title']}” ({t0['value']:,.2f} {currency}, ~%{t0['pct']:.1f})."
            })

        if top_cats3:
            insights.append({
                "type": "category",
                "title": "Kategori Dağılımı",
                "text": "İlk 3 kategori: " + ", ".join([f"{k} ({v:,.2f} {currency})" for k, v in top_cats3]) + "."
            })

        if top_cat_ratio >= 0.60 and top_cats:
            insights.append({
                "type": "concentration",
                "title": "Tek Kategori Bağımlılığı",
                "text": f"Yoğunlaşma yüksek: {top_cat} top-80 maliyetinin ~%{int(top_cat_ratio*100)}’ini taşıyor."
            })

        if freq_stats:
            f_top = sorted(freq_stats.items(), key=lambda x: x[1], reverse=True)[:2]
            insights.append({
                "type": "frequency",
                "title": "Periyot Yoğunluğu",
                "text": "En maliyetli periyotlar: " + ", ".join([f"{k} ({v:,.2f} {currency})" for k, v in f_top]) + "."
            })

        if quick_win_candidates:
            insights.append({
                "type": "quickwin",
                "title": "Quick Win Adayları",
                "text": "Çok kalemli riskler: " + ", ".join([f"{x['title']} ({x['n_items']} kalem)" for x in quick_win_candidates]) + "."
            })

        insights.append({
            "type": "scenario",
            "title": "Tasarruf Senaryosu",
            "text": f"Senaryo ({scope_label}): maliyeti %{int(scenario_cut*100)} düşürürsen ~{scenario_saving:,.2f} {currency} tasarruf, yeni toplam ~{scenario_after:,.2f} {currency}."
        })

        insights.append({
            "type": "metric",
            "title": "Yoğunlaşma Metrikleri",
            "text": f"Kategori HHI: {cat_hhi:.3f} · Risk HHI: {risk_hhi:.3f} (yüksek = az sayıda öğe domine ediyor)"
        })

        # Summary: ilk 3 insight’tan tek paragraf
        summary = " ".join([i["text"] for i in insights[:3]])

        # ----------------------------
        # Actions (structured) - EXPANDED
        # ----------------------------
        actions = []
        _seen = set()

        def add_action(a: dict):
            """Dedup by title, keep order."""
            t = (a.get("title") or "").strip().lower()
            if not t:
                return
            if t in _seen:
                return
            _seen.add(t)
            actions.append(a)

        def cut_pct_label(cut):
            try:
                return int(float(cut) * 100)
            except Exception:
                return 10

        cut_pct = cut_pct_label(scenario_cut)

        # 0) Always: first 48h playbook
        add_action({
            "type": "triage",
            "priority": "high",
            "title": "İlk 48 saat: Top maliyet sürücülerini kilitle",
            "details": "Top 10 CostItem kalemini çıkar: owner, neden, tekrar sıklığı, onay noktası. 'Bugün durdurulabilir mi?' filtresi uygula.",
            "kpi": "48 saatte Top 10 kalem sahipliği %100",
            "url": None
        })

        # 1) Top risks: deep dive first 1-3
        if top_risks:
            t0 = top_risks[0]
            add_action({
                "type": "workshop",
                "priority": "high",
                "title": f"{t0['title']} için kök neden + en pahalı kalem temizliği",
                "details": "30 dk mini çalıştay: en pahalı 2-3 CostItem kalemini incele, tekrar edenleri konsolide et, gereksizleri kaldır.",
                "kpi": f"{currency} maliyeti 30 günde -%{cut_pct}",
                "url": t0.get("url")
            })

            for t in top_risks[1:3]:
                add_action({
                    "type": "workshop",
                    "priority": "high",
                    "title": f"{t.get('title','Risk')} için Stop/Start/Continue maliyet kararı",
                    "details": "Riskin maliyet kalemlerini Stop/Start/Continue etiketle. Stop olanlar için 7 gün içinde kapatma planı yaz.",
                    "kpi": "7 günde Stop kalemlerin %80'i kapatıldı",
                    "url": t.get("url")
                })

        # 2) Quick wins (up to 8)
        for x in (quick_win_candidates or [])[:8]:
            add_action({
                "type": "consolidate",
                "priority": "medium",
                "title": f"{x['title']} kalem konsolidasyonu ({x['n_items']} kalem)",
                "details": "CostItem’ları kategori+periyot bazında grupla. Aynı işi yapan kalemleri standardize et ve birleştir.",
                "kpi": f"Kalem sayısı -%20, toplam maliyet -%{cut_pct}",
                "url": f"/reports/{x['id']}"
            })

        # 3) Category controls if concentrated
        if top_cat_ratio >= 0.60 and top_cats:
            add_action({
                "type": "controls",
                "priority": "high",
                "title": f"{top_cat} kategorisine kontrol listesi + 2. onay",
                "details": "Üst limit, ikinci onay, teklif karşılaştırma ve standart kalem şablonu ekle. Serbest kalem açmayı kısıtla.",
                "kpi": f"{top_cat} maliyeti 90 günde -%{cut_pct}",
                "url": None
            })
            add_action({
                "type": "standardize",
                "priority": "medium",
                "title": f"{top_cat} için standart CostItem kataloğu",
                "details": "En sık geçen kalemleri (isim, birim, periyot, açıklama) standardize et. Duplicate isimleri temizle.",
                "kpi": "Yeni kalem açma oranı -%50",
                "url": None
            })

        # 4) HHI-based governance
        if (cat_hhi is not None) and (cat_hhi >= 0.25):
            add_action({
                "type": "portfolio",
                "priority": "medium",
                "title": "Kategori yoğunlaşması yüksek: portföy dengeleme",
                "details": "Tek kategoriye yığılma varsa sürücü bazlı alternatif mitigasyonlar ve maliyet dağıtım planı oluştur.",
                "kpi": "Cat HHI 60 günde -%10",
                "url": None
            })

        if (risk_hhi is not None) and (risk_hhi >= 0.18):
            add_action({
                "type": "portfolio",
                "priority": "medium",
                "title": "Risk yoğunlaşması yüksek: ilk 5 riske owner + haftalık review",
                "details": "İlk 5 risk için sorumlu ata. Haftalık 15 dk review: gerçekleşen maliyet, plan, sapma nedeni.",
                "kpi": "İlk 5 riskte haftalık sapma raporu %100",
                "url": None
            })

        # 5) Frequency-based actions from enriched top_risks
        def add_freq_actions(r):
            ci = (r or {}).get("cost_items") or {}
            fb = ci.get("freq_breakdown") or {}
            if not isinstance(fb, dict) or not fb:
                return

            pairs = sorted([(k, float(v or 0)) for k, v in fb.items()], key=lambda x: x[1], reverse=True)[:2]
            for freq, v in pairs:
                if v <= 0:
                    continue
                f = str(freq).lower()
                ttl = r.get("title") or "Risk"
                url = r.get("url")

                if "daily" in f or "gün" in f:
                    add_action({
                        "type": "cadence",
                        "priority": "high",
                        "title": f"{ttl}: günlük kalemlere limit + anomali uyarısı",
                        "details": "Günlük tekrarlayan kalemler için limit/uyarı kur. % artış olursa ikinci onaya düşsün.",
                        "kpi": "Günlük anomalileri yakalama %90",
                        "url": url
                    })
                elif "weekly" in f or "hafta" in f:
                    add_action({
                        "type": "cadence",
                        "priority": "medium",
                        "title": f"{ttl}: haftalık kalemlerde birleştirme/planlama optimizasyonu",
                        "details": "Haftalık kalemleri tek güne/küme teslimata topla. Aynı işi yapanları konsolide et.",
                        "kpi": f"Haftalık kalem maliyeti 30 günde -%{min(cut_pct, 15)}",
                        "url": url
                    })
                else:
                    add_action({
                        "type": "cadence",
                        "priority": "low",
                        "title": f"{ttl}: periyot standardizasyonu ({freq})",
                        "details": "Periyot tanımlarını sadeleştir. Raporlama için standarda çek.",
                        "kpi": "Periyot çeşitliliği -%30",
                        "url": url
                    })

        for r in (top_risks or [])[:5]:
            add_freq_actions(r)

        # 6) Data quality + alerts
        add_action({
            "type": "data",
            "priority": "medium",
            "title": "CostItem isimleri: duplicate/benzer isim temizliği",
            "details": "Benzer isimli kalemleri tek standarda indir (örn. 'Nakliye', 'Nakliye bedeli'). Tag sistemi ekle.",
            "kpi": "Duplicate isim oranı -%60",
            "url": None
        })

        add_action({
            "type": "alerts",
            "priority": "medium",
            "title": "Eşik bazlı uyarı: % artış / bütçe sapması",
            "details": "Kalem bazında %10+ artış veya bütçe sapması olunca uyarı üret ve owner onayına düşür.",
            "kpi": "Sapma yakalama: haftalık %100",
            "url": None
        })

        # 7) Long-tail hygiene if top80 exists
        if top_80_count and top_80_count > 0:
            add_action({
                "type": "hygiene",
                "priority": "low",
                "title": "Uzun kuyruk: Top-80 dışını paket yaklaşımıyla yönet",
                "details": "Top-80 dışındaki küçük riskleri kategori/periyot altında paketleyip birleşik metriklerle takip et.",
                "kpi": "Top-80 dışı rapor zamanı -%30",
                "url": None
            })

        # 8) Keep your mitigation baseline (always)
        add_action({
            "type": "mitigation",
            "priority": "low",
            "title": "Top-80 riskler için ölçülebilir mitigation hedefi",
            "details": f"Mitigation planına metrik koy: ‘{currency} maliyeti 90 günde %{cut_pct} azalt’. Haftalık takip metriklerini yaz.",
            "kpi": "Haftalık takip: gerçekleşen tasarruf / plan",
            "url": None
        })

        # UI’yi boğma diye (sonsuz aksiyon üretmek kolay)
        actions = actions[:14]

        payload = {
            "currency": currency,
            "summary": summary,
            "insights": insights,
            "top_risks": top_risks,
            "actions": actions,
            "meta": {
                "total": total,
                "top_80_count": top_80_count,
                "scenario_cut": scenario_cut,
                "scope": scenario_scope,
                "scope_label": scope_label,
                "scenario_saving": scenario_saving,
                "scenario_after": scenario_after,
                "top_category": top_cat,
                "top_category_ratio": top_cat_ratio,
                "cat_hhi": cat_hhi,
                "risk_hhi": risk_hhi,
            }
        }

        _PARETO_AI_CACHE[cache_key] = (now, payload)
        return jsonify(payload)





    

    




    
    return app
def _truthy(v) -> bool:
    return str(v or "").strip().lower() in ("on", "true", "1", "yes")


# -------------------------------------------------
#  Uygulama Başlatma
# -------------------------------------------------
if __name__ == "__main__":
    app = create_app()
    app.run(debug=True)
